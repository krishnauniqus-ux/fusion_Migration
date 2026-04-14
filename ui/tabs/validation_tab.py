"""
validation_tab.py - Validation tab - EXACT from vnew.py
"""
import gc
import streamlit as st
import pandas as pd
import io
import re
import unicodedata
from rapidfuzz import fuzz
from features.validation.engine import ValidationEngine
from core.models import ColumnRule
from config.settings import (
    VALIDATION_GC_EVERY_N_ROWS,
    VALIDATION_PROGRESS_MIN_ROWS,
    VALIDATION_PROGRESS_MAX_FRACTION,
    ERROR_UI_MAX_ROWS,
)


def _normalize_for_duplicate_check(value) -> str:
    """Normalize values so visually equal Excel text compares equal."""
    if value is None or pd.isna(value):
        return ""
    text = str(value)
    text = unicodedata.normalize("NFKC", text)
    text = re.sub(r"[\u200b\u200c\u200d\ufeff]", "", text)
    text = text.replace("\u00a0", " ").replace("\t", " ")
    text = re.sub(r"\s+", " ", text).strip()
    return text.casefold()


def _sanitize_for_excel_cell(value):
    """Strip control chars rejected by Excel/XML; same rules as openpyxl Cell.check_string."""
    if isinstance(value, str):
        from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
        return ILLEGAL_CHARACTERS_RE.sub("", value)
    return value


def _can_fuzz_ratio_reach_85(len_a: int, len_b: int) -> bool:
    """If False, RapidFuzz/Levenshtein ratio cannot be >= 85 (safe to skip pair)."""
    if len_a < 0 or len_b < 0:
        return False
    s = len_a + len_b
    if s == 0:
        return False
    return (2.0 * min(len_a, len_b)) / s >= 0.85


def render_validation_tab():
    st.markdown('<div class="section-title">Success Step 5: Validation & Export</div>', 
                unsafe_allow_html=True)
    
    if not st.session_state.mappings:
        st.warning("Warning Please configure column mappings first")
        return
    
    # Tracking Column Selection - BEFORE Execute Button
    st.markdown('<div class="section-container">', unsafe_allow_html=True)

    
    # Get source dataframe for tracking column options
    source_df = st.session_state.source_file.data.get(
        st.session_state.source_file.selected_sheet,
        pd.DataFrame()
    )
    
    available_columns = ["Row Number (Default)"] + list(source_df.columns)
    
    # Initialize tracking column in session state if not exists
    if 'error_tracking_column' not in st.session_state:
        st.session_state.error_tracking_column = "Row Number (Default)"
    
    # Tracking column selector - NO AUTO-RERUN
    tracking_col = st.selectbox(
        "Select Tracking ID Column for Error Reports",
        available_columns,
        index=available_columns.index(st.session_state.error_tracking_column) if st.session_state.error_tracking_column in available_columns else 0,
        help="This column will be shown in validation errors to help you locate the exact row in your source file. Select BEFORE executing validation.",
        key="tracking_col_selector_before"
    )
    
    # Update session state WITHOUT rerun
    st.session_state.error_tracking_column = tracking_col
    
    
    # Execute validation button
    if st.button(" Execute Validation & Transform", type="primary", use_container_width=True):
        with st.spinner("Processing data... This may take a moment"):
            execute_validation_pipeline()
    
    # Display results if available
    if st.session_state.validation_results:
        results = st.session_state.validation_results
        
        # Summary cards
        st.markdown('<div class="section-container">', unsafe_allow_html=True)
        
        cols = st.columns(4)
        with cols[0]:
            st.metric(
                "Total Rows",
                results.get('total_rows', 0)
            )
        with cols[1]:
            valid_pct = results.get('valid_percentage', 0)
            st.metric(
                "Valid Rows",
                results.get('valid_rows', 0),
                delta=f"{valid_pct:.1f}%"
            )
        with cols[2]:
            error_pct = results.get('error_percentage', 0)
            st.metric(
                "Rows with Errors",
                results.get('error_rows', 0),
                delta=f"{error_pct:.1f}%",
                delta_color="inverse"
            )
        with cols[3]:
            st.metric(
                "Total Errors",
                results.get('total_errors', 0)
            )
        
        st.markdown('</div>', unsafe_allow_html=True)
        
        # Error details
        if results.get('errors'):
            st.markdown('<div class="section-container">', unsafe_allow_html=True)
            st.markdown("** Validation Errors**")
            
            error_df = pd.DataFrame(results['errors'])
            
            # Filter dropdown only - NO tracking column selector here
            col1, col2 = st.columns(2)
            
            with col1:
                # Show current tracking column (read-only info)
                tracking_col_name = st.session_state.get('error_tracking_column', 'Row Number (Default)')
                st.info(f"📍 Tracking ID: **{tracking_col_name}**")
            
            with col2:
                filter_col = st.selectbox(
                    "Filter by Column",
                    ["All"] + list(error_df['target_column'].unique()),
                    key="error_filter_column"
                )
            
            if filter_col != "All":
                error_df = error_df[error_df['target_column'] == filter_col]
            
            # Dynamic column configuration based on tracking column
            tracking_col_name = st.session_state.get('error_tracking_column', 'Row Number (Default)')
            
            # Rename tracking_value column to actual tracking column name
            if tracking_col_name != "Row Number (Default)" and 'tracking_value' in error_df.columns:
                error_df = error_df.rename(columns={'tracking_value': tracking_col_name})
                # Reorder columns to show tracking column first
                cols = [tracking_col_name, 'row', 'source_column', 'target_column', 'value', 'error']
                # Only include columns that exist in the dataframe
                cols = [col for col in cols if col in error_df.columns]
                error_df = error_df[cols]
            
            total_err_rows = len(error_df)
            error_df_view = (
                error_df.head(ERROR_UI_MAX_ROWS)
                if total_err_rows > ERROR_UI_MAX_ROWS
                else error_df
            )
            
            if tracking_col_name == "Row Number (Default)":
                column_config = {
                    "row": st.column_config.NumberColumn("Row #"),
                    "source_column": st.column_config.TextColumn("Source"),
                    "target_column": st.column_config.TextColumn("Target"),
                    "value": st.column_config.TextColumn("Value"),
                    "error": st.column_config.TextColumn("Error")
                }
            else:
                column_config = {
                    tracking_col_name: st.column_config.TextColumn(f"{tracking_col_name}"),
                    "row": st.column_config.NumberColumn("Row #"),
                    "source_column": st.column_config.TextColumn("Source"),
                    "target_column": st.column_config.TextColumn("Target"),
                    "value": st.column_config.TextColumn("Value"),
                    "error": st.column_config.TextColumn("Error")
                }
            
            if total_err_rows > ERROR_UI_MAX_ROWS:
                st.caption(
                    f"Showing first {ERROR_UI_MAX_ROWS:,} of {total_err_rows:,} error rows "
                    "(all errors are kept for export / rejected-records workflows)."
                )
            st.dataframe(
                error_df_view,
                use_container_width=True,
                height=300,
                column_config=column_config
            )
            st.markdown('</div>', unsafe_allow_html=True)
        
        # Transformed data preview
        if st.session_state.transformed_data is not None:
            st.markdown('<div class="section-container">', unsafe_allow_html=True)
            st.markdown("** Transformed Data Preview**")
            
            df = st.session_state.transformed_data
            
            # Show sample with color coding for errors
            st.dataframe(
                df.head(100),
                use_container_width=True,
                height=400
            )
            
            # Export options
            st.markdown("** Export Options**")
            
            col1, col2, col3 = st.columns(3)
            
            with col1:
                # Excel export - preserve entire template structure, replace only selected sheet
                if st.session_state.template_file.raw_data:
                    if st.button("📥 Prepare Excel Export", use_container_width=True, type="primary"):
                        with st.spinner("Preparing Excel export..."):
                            try:
                                # Import openpyxl for preserving formatting
                                from openpyxl import load_workbook
                                from openpyxl.utils.dataframe import dataframe_to_rows
                                
                                progress_bar = st.progress(0)
                                status_text = st.empty()
                                
                                status_text.text("Loading template workbook...")
                                progress_bar.progress(0.1)
                                
                                # Load the original template workbook to preserve all sheets and formatting
                                wb = load_workbook(io.BytesIO(st.session_state.template_file.raw_data))
                                
                                # Get the selected sheet name
                                selected_sheet = st.session_state.template_file.selected_sheet
                                
                                # Check if the selected sheet exists in the workbook
                                if selected_sheet in wb.sheetnames:
                                    status_text.text("Preparing sheet data...")
                                    progress_bar.progress(0.2)
                                    
                                    # Preserve original data and only update mapped columns
                                    ws = wb[selected_sheet]
                                    
                                    # Get template columns and mapped columns
                                    template_cols = st.session_state.template_file.columns.get(selected_sheet, [])
                                    mapped_cols = [m.target_column for m in st.session_state.mappings if m.is_active and m.target_sheet == selected_sheet]
                                    
                                    # Get the actual header row position (0-indexed from session state)
                                    header_row = st.session_state.template_file.header_row
                                    data_start_row = header_row + 2  # Row after headers (1-indexed)
                                    
                                    # Create column index mapping
                                    col_indices = {}
                                    for col_idx, col_name in enumerate(template_cols, 1):
                                        col_indices[col_name] = col_idx
                                    
                                    status_text.text("Clearing old data...")
                                    progress_bar.progress(0.3)
                                    
                                    # Find the last row and column
                                    max_row = ws.max_row
                                    max_col = ws.max_column
                                    
                                    # Clear data only in mapped columns from data_start_row onwards
                                    for row in range(data_start_row, max_row + 1):
                                        for mapped_col in mapped_cols:
                                            if mapped_col in col_indices:
                                                col_idx = col_indices[mapped_col]
                                                if col_idx <= max_col:
                                                    cell = ws.cell(row=row, column=col_idx)
                                                    cell.value = None
                                                    # Clear any comments/notes in mapped columns
                                                    if cell.comment:
                                                        cell.comment = None
                                    
                                    # Clear comments from header row in mapped columns
                                    header_row_1_indexed = header_row + 1
                                    for mapped_col in mapped_cols:
                                        if mapped_col in col_indices:
                                            col_idx = col_indices[mapped_col]
                                            if col_idx <= max_col:
                                                cell = ws.cell(row=header_row_1_indexed, column=col_idx)
                                                if cell.comment:
                                                    cell.comment = None
                                    
                                    # Hide all comments in the worksheet
                                    ws.show_comments = False
                                    
                                    status_text.text("Writing transformed data...")
                                    progress_bar.progress(0.5)
                                    
                                    # Add the mapped data only to mapped columns starting from data_start_row
                                    if len(df) > 0:
                                        total_rows = len(df)
                                        xport_prog_every = max(100, total_rows // 200)
                                        # Write data starting from data_start_row, only to mapped columns
                                        for r_idx, row in enumerate(df.itertuples(index=False), data_start_row):
                                            written = r_idx - data_start_row
                                            if written % xport_prog_every == 0:
                                                progress = 0.5 + 0.4 * (written / total_rows)
                                                progress_bar.progress(progress)
                                                status_text.text(f"Writing row {written + 1} of {total_rows}...")
                                            if written > 0 and written % 15000 == 0:
                                                gc.collect()
                                            
                                            for mapped_col in mapped_cols:
                                                if mapped_col in col_indices:
                                                    col_idx = col_indices[mapped_col]
                                                    if col_idx <= max_col:
                                                        # Get the value for this mapped column
                                                        if hasattr(row, mapped_col):
                                                            value = getattr(row, mapped_col)
                                                        else:
                                                            # Fallback to column position
                                                            col_pos = template_cols.index(mapped_col) if mapped_col in template_cols else None
                                                            value = row[col_pos] if col_pos is not None and col_pos < len(row) else None
                                                        
                                                        cell = ws.cell(row=r_idx, column=col_idx)
                                                        cell.value = _sanitize_for_excel_cell(value)
                                                        # Ensure no comments on new cells
                                                        if cell.comment:
                                                            cell.comment = None
                                    
                                    # Also hide comments for the entire workbook
                                    wb.show_comments = False
                                    
                                    status_text.text("Saving workbook...")
                                    progress_bar.progress(0.9)
                                    
                                    # Save the modified workbook
                                    output = io.BytesIO()
                                    wb.save(output)
                                    
                                    progress_bar.progress(1.0)
                                    status_text.text("✅ Export ready!")
                                    
                                    # Use original filename
                                    original_filename = st.session_state.template_file.name
                                    if not original_filename.lower().endswith('.xlsx'):
                                        original_filename = original_filename.rsplit('.', 1)[0] + '_export.xlsx'
                                    
                                    st.download_button(
                                        label="📥 Download Excel (.xlsx)",
                                        data=output.getvalue(),
                                        file_name=original_filename,
                                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                        use_container_width=True
                                    )
                                else:
                                    # Fallback if sheet not found
                                    st.error(f"❌ Selected sheet '{selected_sheet}' not found in template")
                                    output = io.BytesIO()
                                    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                                        df.to_excel(writer, sheet_name='Mapped_Data', index=False)
                                        if results.get('errors'):
                                            error_df = pd.DataFrame(results['errors'])
                                            error_df.to_excel(writer, sheet_name='Validation_Errors', index=False)
                                    
                                    st.download_button(
                                        label="📥 Download Excel (.xlsx)",
                                        data=output.getvalue(),
                                        file_name=f"template_export_{st.session_state.session_id}.xlsx",
                                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                        use_container_width=True
                                    )
                                    
                            except ImportError:
                                st.warning("⚠ openpyxl not available, using fallback export")
                                # Fallback without formatting preservation
                                output = io.BytesIO()
                                with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                                    df.to_excel(writer, sheet_name='Mapped_Data', index=False)
                                    if results.get('errors'):
                                        error_df_export = pd.DataFrame(results['errors'])
                                        # Rename tracking_value column to actual tracking column name
                                        tracking_col_name = st.session_state.get('error_tracking_column', 'Row Number (Default)')
                                        if tracking_col_name != "Row Number (Default)" and 'tracking_value' in error_df_export.columns:
                                            error_df_export = error_df_export.rename(columns={'tracking_value': tracking_col_name})
                                        error_df_export.to_excel(writer, sheet_name='Validation_Errors', index=False)
                                
                                st.download_button(
                                    label="📥 Download Excel (.xlsx)",
                                    data=output.getvalue(),
                                    file_name=f"template_export_{st.session_state.session_id}.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    use_container_width=True
                                )
                else:
                    # No original template available
                    st.warning("⚠ Original template not available for format preservation")
                    output = io.BytesIO()
                    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                        df.to_excel(writer, sheet_name='Mapped_Data', index=False)
                        if results.get('errors'):
                            error_df_export = pd.DataFrame(results['errors'])
                            # Rename tracking_value column to actual tracking column name
                            tracking_col_name = st.session_state.get('error_tracking_column', 'Row Number (Default)')
                            if tracking_col_name != "Row Number (Default)" and 'tracking_value' in error_df_export.columns:
                                error_df_export = error_df_export.rename(columns={'tracking_value': tracking_col_name})
                            error_df_export.to_excel(writer, sheet_name='Validation_Errors', index=False)
                    
                    st.download_button(
                        label="📥 Download Excel (.xlsx)",
                        data=output.getvalue(),
                        file_name=f"template_export_{st.session_state.session_id}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
            
            with col2:
                # CSV export
                csv = df.to_csv(index=False)
                st.download_button(
                    label=" Download CSV",
                    data=csv,
                    file_name=f"mapped_data_{st.session_state.session_id}.csv",
                    mime="text/csv",
                    use_container_width=True
                )
            
            with col3:
                # Rejected records export
                if results.get('errors'):
                    error_df_export = pd.DataFrame(results['errors'])
                    
                    # Rename tracking_value column to actual tracking column name
                    tracking_col_name = st.session_state.get('error_tracking_column', 'Row Number (Default)')
                    if tracking_col_name != "Row Number (Default)" and 'tracking_value' in error_df_export.columns:
                        error_df_export = error_df_export.rename(columns={'tracking_value': tracking_col_name})
                    
                    output = io.BytesIO()
                    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                        error_df_export.to_excel(writer, sheet_name='Rejected_Records', index=False)

                    source_filename = st.session_state.source_file.name or "source_data.xlsx"
                    if source_filename.lower().endswith('.xlsx'):
                        base_name = source_filename[:-5]
                    elif '.' in source_filename:
                        base_name = source_filename.rsplit('.', 1)[0]
                    else:
                        base_name = source_filename
                    rejected_filename = f"{base_name}_Validation_Error.xlsx"
                    
                    st.download_button(
                        label=" Download Rejected Records",
                        data=output.getvalue(),
                        file_name=rejected_filename,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
                else:
                    st.info(" No rejected records to export")
            
            st.markdown('</div>', unsafe_allow_html=True)
    
    st.markdown("---")
    if st.button("⬅️ Back to Mapping", use_container_width=True, key='btn_validate_to_mapping'):
        st.session_state.active_tab = 4
        st.session_state.tab_change_requested = True
        st.rerun()

def execute_validation_pipeline():
    """Execute the full validation and transformation pipeline.

    Key guarantees:
    - Every source row produces exactly one output row (no data loss).
    - Transform rules (case, no_special_chars) are applied to the output.
    - All other rules only generate errors — they never change the data.
    - All errors are stored without any cap/limit.
    - Duplicate/unique post-processing correctly updates the error-row count.
    """
    try:
        source_df = st.session_state.source_file.data.get(
            st.session_state.source_file.selected_sheet,
            pd.DataFrame()
        )

        template_cols = st.session_state.template_file.columns.get(
            st.session_state.template_file.selected_sheet,
            []
        )

        # Get tracking column from session state
        tracking_column = st.session_state.get('error_tracking_column', 'Row Number (Default)')
        use_tracking_column = tracking_column != "Row Number (Default)" and tracking_column in source_df.columns

        errors: list = []
        error_row_set: set = set()

        total_rows = len(source_df)
        transformed_rows: list = [None] * total_rows if total_rows else []

        progress_bar = st.progress(0)
        status_text = st.empty()
        prog_every = max(
            VALIDATION_PROGRESS_MIN_ROWS,
            total_rows // VALIDATION_PROGRESS_MAX_FRACTION if total_rows else 1,
        )

        # Detect duplicate target-column mappings — warn and keep only last
        seen_targets: dict = {}
        for mapping in st.session_state.mappings:
            if not mapping.is_active:
                continue
            tc = mapping.target_column
            if tc in seen_targets:
                st.warning(
                    f"⚠️ Target column '{tc}' is mapped from multiple source "
                    f"columns ('{seen_targets[tc]}' and '{mapping.source_column}'). "
                    f"Only the last mapping ('{mapping.source_column}') will be used."
                )
            seen_targets[tc] = mapping.source_column

        # Track original values for duplicate and unique checks
        column_values: dict = {col: [] for col in template_cols}

        _col_names = list(source_df.columns)
        _use_itertuples = len(_col_names) == len(set(_col_names))

        # itertuples is much faster than iterrows; fall back if duplicate column names.
        if _use_itertuples:
            _row_iter = enumerate(source_df.itertuples(index=False, name=None))
        else:
            _row_iter = enumerate(r for _, r in source_df.iterrows())

        for row_num, row in _row_iter:
            if _use_itertuples:
                row = dict(zip(_col_names, row))

            if row_num % prog_every == 0 or row_num == total_rows - 1:
                progress = min((row_num + 1) / total_rows, 1.0) if total_rows else 1.0
                progress_bar.progress(progress)
                status_text.text(
                    f"Processing row {row_num + 1} of {total_rows} "
                    f"({progress * 100:.1f}%)"
                )

            if (
                VALIDATION_GC_EVERY_N_ROWS
                and row_num > 0
                and row_num % VALIDATION_GC_EVERY_N_ROWS == 0
            ):
                gc.collect()

            display_row = row_num + 1

            if use_tracking_column:
                tracking_value = (
                    str(row[tracking_column])
                    if tracking_column in row
                    else str(display_row)
                )
            else:
                tracking_value = None

            new_row: dict = {}

            for mapping in st.session_state.mappings:
                if not mapping.is_active:
                    continue

                source_val = row.get(mapping.source_column, None)
                target_col = mapping.target_column
                rule = st.session_state.column_rules.get(
                    target_col,
                    ColumnRule(column_name=target_col)
                )

                # Engine returns (is_valid, errors, output_val).
                # Transform rules (case, no_special_chars) mutate output_val.
                # All other rules only produce errors — output is untouched.
                _is_valid, val_errors, output_val = ValidationEngine.validate_value(
                    source_val, rule
                )

                filtered_errors = [
                    e for e in val_errors
                    if e not in ("UNIQUE_CHECK", "DUPLICATE_CHECK")
                ]

                if filtered_errors:
                    error_row_set.add(display_row)
                    for err in filtered_errors:
                        error_entry = {
                            'row': display_row,
                            'source_column': mapping.source_column,
                            'target_column': target_col,
                            'value': source_val,
                            'error': err
                        }
                        if use_tracking_column:
                            error_entry['tracking_value'] = tracking_value
                        errors.append(error_entry)

                new_row[target_col] = output_val

                # Track ORIGINAL values for duplicate/unique checks
                if target_col in column_values:
                    column_values[target_col].append({
                        'tracking_value': tracking_value if use_tracking_column else None,
                        'row': display_row,
                        'value': source_val,
                        'source_column': mapping.source_column
                    })

            # Handle unmapped template columns
            for col in template_cols:
                if col not in new_row:
                    rule = st.session_state.column_rules.get(col, ColumnRule(column_name=col))
                    new_row[col] = rule.default_value if rule.default_value else None

                    if rule.is_mandatory:
                        error_row_set.add(display_row)
                        error_entry = {
                            'row': display_row,
                            'source_column': 'N/A',
                            'target_column': col,
                            'value': None,
                            'error': 'Mandatory column not mapped'
                        }
                        if use_tracking_column:
                            error_entry['tracking_value'] = tracking_value
                        errors.append(error_entry)

            transformed_rows[row_num] = new_row

        # ----- Post-processing: duplicates & unique checks -----
        status_text.text("Checking for duplicates and unique values...")
        progress_bar.progress(0.95)

        for col, values_list in column_values.items():
            if not values_list:
                continue

            rule = st.session_state.column_rules.get(col, ColumnRule(column_name=col))

            if rule.check_duplicates:
                value_counts: dict = {}
                for item in values_list:
                    normalized_val = _normalize_for_duplicate_check(item['value'])
                    if normalized_val:
                        value_counts.setdefault(normalized_val, []).append(item)

                for _normalized_val, occurrences in value_counts.items():
                    if len(occurrences) > 1:
                        for item in occurrences:
                            error_row_set.add(item['row'])
                            error_entry = {
                                'row': item['row'],
                                'source_column': item['source_column'],
                                'target_column': col,
                                'value': item['value'],
                                'error': f'Duplicate value (appears {len(occurrences)} times)'
                            }
                            if use_tracking_column and item.get('tracking_value'):
                                error_entry['tracking_value'] = item['tracking_value']
                            errors.append(error_entry)

            if rule.unique_value:
                seen_values: dict = {}
                for item in values_list:
                    normalized_val = _normalize_for_duplicate_check(item['value'])
                    if normalized_val:
                        if normalized_val in seen_values:
                            error_row_set.add(item['row'])
                            error_entry = {
                                'row': item['row'],
                                'source_column': item['source_column'],
                                'target_column': col,
                                'value': item['value'],
                                'error': (
                                    "Value must be unique "
                                    f"(first seen at row {seen_values[normalized_val]})"
                                )
                            }
                            if use_tracking_column and item.get('tracking_value'):
                                error_entry['tracking_value'] = item['tracking_value']
                            errors.append(error_entry)
                        else:
                            seen_values[normalized_val] = item['row']

            if rule.similar_match:
                normalized_items = []
                for item in values_list:
                    normalized_val = _normalize_for_duplicate_check(item['value'])
                    if normalized_val:
                        normalized_items.append((item, normalized_val))

                if normalized_items:
                    status_text.text(
                        f"Similar match on '{col}' ({len(normalized_items):,} values)..."
                    )

                # Keep memory bounded: summarize per row, don't append every pair.
                row_similar_summary = {}
                for i in range(len(normalized_items)):
                    item_i, text_i = normalized_items[i]
                    li = len(text_i)
                    for j in range(i + 1, len(normalized_items)):
                        item_j, text_j = normalized_items[j]
                        lj = len(text_j)
                        if not _can_fuzz_ratio_reach_85(li, lj):
                            continue
                        score = fuzz.ratio(text_i, text_j)
                        if score >= 85:
                            error_row_set.add(item_i['row'])
                            error_row_set.add(item_j['row'])
                            for current, other in ((item_i, item_j), (item_j, item_i)):
                                row_key = (current['row'], current['source_column'])
                                existing = row_similar_summary.get(row_key)
                                if existing is None:
                                    row_similar_summary[row_key] = {
                                        'item': current,
                                        'best_score': score,
                                        'best_match_row': other['row'],
                                        'matches': 1
                                    }
                                else:
                                    existing['matches'] += 1
                                    if score > existing['best_score']:
                                        existing['best_score'] = score
                                        existing['best_match_row'] = other['row']

                for summary in row_similar_summary.values():
                    item = summary['item']
                    error_entry = {
                        'row': item['row'],
                        'source_column': item['source_column'],
                        'target_column': col,
                        'value': item['value'],
                        'error': (
                            f"Similar match found: {summary['matches']} row(s), "
                            f"highest {summary['best_score']:.1f}% with row "
                            f"{summary['best_match_row']} (threshold: 85%)"
                        )
                    }
                    if use_tracking_column and item.get('tracking_value'):
                        error_entry['tracking_value'] = item['tracking_value']
                    errors.append(error_entry)

        gc.collect()

        # Build transformed DataFrame in one shot
        transformed_data = pd.DataFrame(transformed_rows, columns=template_cols)

        # Derive counts from the authoritative error_row_set
        error_rows = len(error_row_set)
        valid_rows = total_rows - error_rows

        progress_bar.progress(1.0)
        status_text.text(f"✅ Completed processing {total_rows} rows")

        st.session_state.transformed_data = transformed_data
        st.session_state.validation_results = {
            'total_rows': total_rows,
            'valid_rows': valid_rows,
            'error_rows': error_rows,
            'valid_percentage': (valid_rows / total_rows * 100) if total_rows > 0 else 0,
            'error_percentage': (error_rows / total_rows * 100) if total_rows > 0 else 0,
            'total_errors': len(errors),
            'errors': errors
        }

        st.success(
            f"✅ Validation complete! Processed {total_rows:,} rows "
            f"with {valid_rows:,} valid and {error_rows:,} with errors"
        )

    except Exception as e:
        st.error(f"❌ Error during processing: {str(e)}")
        import traceback
        st.error(traceback.format_exc())
