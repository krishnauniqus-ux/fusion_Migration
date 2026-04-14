"""
Oracle Fusion Universal Enterprise Field Mapper — v7.0 REAL-TIME TRANSFORMATION EDITION
Production-Ready | Enterprise Validation Engine | AI Mandatory Detection | Real-Time Transformation
Enhanced: Dynamic Transformation Panel + Live Preview + Conditional Logic
"""

import streamlit as st
import pandas as pd
import json, re, os, hashlib, warnings
from typing import Dict, List, Tuple, Optional, Any
from dataclasses import dataclass, field, asdict
from enum import Enum
from datetime import datetime
from io import BytesIO
import numpy as np
from difflib import SequenceMatcher
import pickle
warnings.filterwarnings("ignore")

st.set_page_config(
    page_title="Oracle Fusion Mapper Enterprise",
    page_icon="🔮",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ══════════════════════════════════════════════════════════════════════════════
# FILE PERSISTENCE HELPER - Modern Streamlit Approach
# ══════════════════════════════════════════════════════════════════════════════
def persist_uploaded_file(uploaded_file, session_key):
    """
    Persist uploaded file in session state so it survives page refresh.
    Modern approach using Streamlit's session state.
    """
    if uploaded_file is not None:
        # Read file bytes and store in session state
        file_bytes = uploaded_file.read()
        uploaded_file.seek(0)  # Reset file pointer
        
        st.session_state[f"{session_key}_bytes"] = file_bytes
        st.session_state[f"{session_key}_name"] = uploaded_file.name
        st.session_state[f"{session_key}_type"] = uploaded_file.type
        return True
    return False

def get_persisted_file(session_key):
    """
    Retrieve persisted file from session state.
    Returns a BytesIO object that mimics uploaded file.
    """
    if f"{session_key}_bytes" in st.session_state:
        from io import BytesIO
        file_bytes = st.session_state[f"{session_key}_bytes"]
        file_name = st.session_state.get(f"{session_key}_name", "file")
        return BytesIO(file_bytes), file_name
    return None, None

# ══════════════════════════════════════════════════════════════════════════════
# CONFIG - Read from Streamlit Secrets
# ══════════════════════════════════════════════════════════════════════════════
class Config:
    """Configuration loaded from .streamlit/secrets.toml"""
    try:
        # Try to load from Streamlit secrets first
        AZURE_OPENAI_ENDPOINT   = st.secrets.get("azure_openai", {}).get("endpoint", "")
        AZURE_OPENAI_KEY        = st.secrets.get("azure_openai", {}).get("api_key", "")
        AZURE_OPENAI_DEPLOYMENT = st.secrets.get("azure_openai", {}).get("deployment", "")
    except:
        # Fallback to environment variables if secrets not available
        AZURE_OPENAI_ENDPOINT   = os.getenv("AZURE_OPENAI_ENDPOINT", "")
        AZURE_OPENAI_KEY        = os.getenv("AZURE_OPENAI_KEY", "")
        AZURE_OPENAI_DEPLOYMENT = os.getenv("AZURE_OPENAI_DEPLOYMENT", "")

# ══════════════════════════════════════════════════════════════════════════════
# ENUMS & DATA CLASSES
# ══════════════════════════════════════════════════════════════════════════════
class DataTypeEnum(Enum):
    TEXT = "Text"
    NUMBER = "Number"
    DATE = "Date"
    EMAIL = "Email"
    CURRENCY = "Currency"
    BOOLEAN = "Boolean"
    CUSTOM_REGEX = "Custom Regex"

class MandatoryStatus(Enum):
    MANDATORY = "Mandatory"
    OPTIONAL = "Optional"

class ValidationStatus(Enum):
    NOT_TESTED = "Not Tested"
    PASSED = "Passed"
    FAILED = "Failed"

class ValidationType(Enum):
    DATE_FORMAT          = "date_format"
    REGEX_PATTERN        = "regex_pattern"
    NUMERIC_RANGE        = "numeric_range"
    LENGTH_LIMIT         = "length_limit"
    LIST_OF_VALUES       = "list_of_values"
    REQUIRED             = "required"
    UNIQUE               = "unique"
    DATA_TYPE_CHECK      = "data_type_check"
    EMAIL_FORMAT         = "email_format"
    PHONE_FORMAT         = "phone_format"
    CONDITIONAL_REQUIRED = "conditional_required"
    CUSTOM               = "custom_formula"

class TemplatePattern(Enum):
    FLAT_HEADER        = "flat_header"
    FOUR_ROW_METADATA  = "four_row_metadata"
    THREE_ROW_METADATA = "three_row_metadata"
    AI_DETECTED        = "ai_detected"

# Enterprise Date Formats
DATE_FORMATS = {
    "Standard Date Formats": [
        "DD/MM/YYYY",
        "DD-MM-YYYY",
        "MM/DD/YYYY",
        "MM-DD-YYYY",
        "YYYY-MM-DD",
        "YYYY/MM/DD"
    ],
    "Short Year Formats": [
        "DD/MM/YY",
        "DD-MM-YY",
        "MM/DD/YY",
        "MM-DD-YY"
    ],
    "ISO Formats (Enterprise Standard)": [
        "YYYY-MM-DD (ISO 8601)",
        "YYYY-MM-DDTHH:MM:SS",
        "YYYY-MM-DDTHH:MM:SSZ",
        "YYYY-MM-DDTHH:MM:SS+05:30"
    ],
    "Date With Time (24 Hour)": [
        "DD/MM/YYYY HH:MM",
        "DD/MM/YYYY HH:MM:SS",
        "YYYY-MM-DD HH:MM:SS"
    ],
    "Date With Time (12 Hour)": [
        "DD/MM/YYYY HH:MM AM/PM",
        "MM/DD/YYYY HH:MM AM/PM"
    ],
    "Timestamp Formats": [
        "Unix Timestamp (seconds)",
        "Unix Timestamp (milliseconds)"
    ],
    "Enterprise Database Formats": [
        "Oracle: DD-MON-YYYY",
        "SQL Server: YYYY-MM-DD HH:MM:SS.SSS",
        "ISO UTC Format"
    ]
}

# Flatten for dropdown
ALL_DATE_FORMATS = []
for category, formats in DATE_FORMATS.items():
    ALL_DATE_FORMATS.extend(formats)

@dataclass
class ConditionalRule:
    """IF-THEN conditional transformation rule"""
    rule_id: str
    condition_field: str  # Source field to check
    operator: str  # equals, contains, >, <, >=, <=, !=, in
    condition_value: str
    target_field: str
    result_value: str
    logic_connector: str = "AND"  # AND, OR, None (for single rule)

@dataclass
class ValueMapping:
    """Source value to target value mapping"""
    source_value: str
    target_value: str

@dataclass
class TransformationConfig:
    """Comprehensive transformation configuration for mapping workspace"""
    # Core Logic
    transformation_type: str = "direct" # direct, concatenate, expression, condition
    
    # Text Transformations
    trim_type: str = "none"  # none, left, right, anywhere
    case_type: str = "none"  # none, upper, lower, title
    remove_special_chars: bool = False
    replace_text_source: str = ""
    replace_text_target: str = ""
    substring_start: Optional[int] = None
    substring_length: Optional[int] = None
    split_delimiter: str = ""
    split_index: int = 0
    extract_name_part: str = "none" # none, first, last
    
    # Concatenation
    concatenate_fields: List[str] = field(default_factory=list) # List of source field IDs
    concatenate_separator: str = " "
    
    # Conditions (IF/ELSE)
    if_field: str = ""
    if_operator: str = "equals"
    if_value: str = ""
    then_result: str = ""
    else_result: str = ""
    handle_null: str = "keep" # keep, default, error
    default_value: str = ""
    
    # Expressions & Calculations
    formula_expression: str = "" # e.g. "{src_1} * 1.05 + 10"
    
    # Date Advanced
    input_date_format: str = ""
    output_date_format: str = "YYYY-MM-DD"
    timezone_source: str = "UTC"
    timezone_target: str = "UTC"
    add_days: int = 0
    subtract_days: int = 0
    
    # Chain Persistence
    transformation_chain: List[Dict[str, Any]] = field(default_factory=list)

@dataclass
class EnterpriseValidationRule:
    """Enterprise-level validation rule with comprehensive configuration"""
    field_id: str
    field_name: str
    data_type: str = "Text"
    rule_name: str = ""  # User-defined rule name
    
    # Text Constraints
    min_length: Optional[int] = None
    max_length: Optional[int] = None
    no_special_chars: bool = False
    regex_pattern: str = ""
    allowed_values: Optional[List[str]] = None
    
    # Number Constraints
    min_value: Optional[float] = None
    max_value: Optional[float] = None
    decimal_allowed: bool = True
    decimal_precision: Optional[int] = None
    duplicate_check: bool = False
    
    # Date Constraints
    accept_all_formats: bool = False
    date_format: str = "YYYY-MM-DD"
    min_date: str = ""
    max_date: str = ""
    date_cannot_be_future: bool = False
    date_cannot_be_past: bool = False
    
    # Custom Error
    custom_error_message: str = ""
    
    # Meta / Status
    validation_status: str = "Not Tested"
    test_result: str = ""
    is_ai_detected: bool = False
    mandatory_status: str = "Optional" # Mandatory, Optional

@dataclass
class ValidationRule:
    rule_id:          str
    field_name:       str
    validation_type:  ValidationType
    rule_config:      Dict[str, Any]
    error_message:    str
    severity:         str = "error"
    auto_fix:         bool = False
    is_ai_generated:  bool = False
    confidence_score: float = 0.0
    source_context:   str = ""

@dataclass
class FieldMetadata:
    field_id:         str
    name:             str
    data_type:        str
    sample_values:    List[str]
    is_mandatory:     bool = False
    mandatory_status: str = "Optional"
    description:      str = ""
    max_length:       Optional[int] = None
    db_column_name:   str = ""
    db_data_type:     str = ""
    validation_rules: List[ValidationRule] = field(default_factory=list)
    enterprise_rule:  Optional[EnterpriseValidationRule] = None
    statistics:       Dict[str, Any]       = field(default_factory=dict)
    column_index:     int = 0
    ai_mandatory_confidence: float = 0.0

@dataclass
class SheetProfile:
    sheet_name:       str
    pattern:          TemplatePattern
    header_row_index: int
    data_start_index: int
    metadata_rows:    List[int]
    fields:           Dict[str, FieldMetadata] = field(default_factory=dict)

@dataclass
class FieldMapping:
    mapping_id:         str
    target_field_id:    str
    target_field_name:  str
    source_field:       str
    source_field_id:    str
    
    # ENHANCED: Enterprise transformations
    transformation:     str = "none"
    transform_params:   Dict[str, Any] = field(default_factory=dict)
    
    # ENHANCED: Transformation chain (multiple transformations)
    transformation_chain: List[Dict[str, Any]] = field(default_factory=list)
    
    # ENHANCED: Conditional logic
    conditional_logic:  str = ""  # Python expression for conditional transformation
    
    # Rules inherited from template (non-editable)
    template_rules:     List[Any] = field(default_factory=list)
    
    # Validation results from mapping
    mapping_validation_status: str = "Pending"  # Pending, Valid, Invalid
    mapping_validation_errors: List[str] = field(default_factory=list)
    
    is_active:          bool = True
    mapping_confidence: float = 0.0
    ai_suggested:       bool = False
    user_notes:         str = ""
    
    # Preview cache
    preview_sample:     Dict[str, Any] = field(default_factory=dict)
    
    # Legacy fields (kept for compatibility)
    validation_passed:  bool = False
    validation_errors:  List[str] = field(default_factory=list)
    transform_config:   Optional[TransformationConfig] = None

@dataclass
class DataQualityReport:
    total_records:      int
    field_reports:      Dict[str, Dict[str, Any]]
    critical_errors:    int
    warnings:           int
    auto_fixes_applied: int
    compliance_score:   float

# ══════════════════════════════════════════════════════════════════════════════
# NEW: TEMPLATE DRIVEN DATA QUALITY RULE SYSTEM
# ══════════════════════════════════════════════════════════════════════════════

class QualityRuleType(Enum):
    COMPLETENESS = "Completeness"
    UNIQUENESS = "Uniqueness"
    FORMAT = "Format"
    STANDARDIZATION = "Standardization"
    VALIDITY = "Validity"
    CHARACTER_LENGTH = "Character Length"
    DATE_FORMAT = "Date Format"
    REGEX = "Regex"

@dataclass
class DataQualityRule:
    """Template-driven data quality rule"""
    rule_id: str
    column_name: str
    field_id: str
    rule_type: QualityRuleType
    rule_description: str
    is_active: bool = True
    
    # Completeness Rules
    not_null: bool = False
    not_blank: bool = False
    
    # Uniqueness Rules
    unique_values: bool = False
    no_duplicates: bool = False
    
    # Format Rules
    alphanumeric_only: bool = False
    no_special_chars: bool = False
    no_latin_chars: bool = False
    
    # Standardization Rules
    to_uppercase: bool = False
    to_lowercase: bool = False
    trim_spaces: bool = False
    remove_special_chars: bool = False
    replace_underscore_with_space: bool = False
    
    # Character Length Rules
    min_length: Optional[int] = None
    max_length: Optional[int] = None
    
    # Date Format Rules
    date_format_pattern: str = "YYYY/MM/DD"
    
    # Regex Rules
    regex_pattern: str = ""
    regex_replacement: str = ""
    regex_description: str = ""
    
    # Validation Results
    validation_status: str = "Not Tested"
    last_test_result: str = ""
    test_sample_count: int = 0
    pass_count: int = 0
    fail_count: int = 0

@dataclass
class CountryStandardization:
    """Country name standardization mapping"""
    mapping_id: str
    source_values: List[str]  # e.g., ["uae", "u.a.e", "united arab emirates"]
    target_value: str  # e.g., "UAE"
    is_active: bool = True

@dataclass
class TransformationPreview:
    """Preview of data transformation"""
    source_value: str
    transformed_value: str
    validation_status: str  # Valid, Invalid, Corrected
    applied_rules: List[str]
    errors: List[str]
    warnings: List[str]

# ══════════════════════════════════════════════════════════════════════════════
# NEW: DATA QUALITY RULE SYSTEM
# ══════════════════════════════════════════════════════════════════════════════

# ══════════════════════════════════════════════════════════════════════════════
# PATTERN DETECTOR
# ══════════════════════════════════════════════════════════════════════════════
class PatternDetector:
    LOV_SHEETS = {
        "lovs","lov","payment terms","paymentterms","bu_name","buname",
        "suppliers","customers","locations","assetcategory","asset category",
        "custdataextractquery","fa_mc_mass_rates","instructions",
        "ra_interface_salescredits_all","ar_interface_conts_all","fa_massadd_distributions",
    }

    def is_reference_sheet(self, name: str) -> bool:
        return name.lower().strip() in self.LOV_SHEETS

    def detect(self, df: pd.DataFrame) -> Tuple[TemplatePattern, int, int, List[int]]:
        if df.empty: return TemplatePattern.FLAT_HEADER, 0, 1, []
        rows = df.values

        def header_score(r: int) -> float:
            row = [str(v).strip() for v in rows[r] if v is not None and str(v).strip() not in ('nan','None','')]
            if not row: return 0.0
            avg_len     = sum(len(v) for v in row) / len(row)
            short_ratio = sum(1 for v in row if len(v) <= 60) / len(row)
            star_ratio  = sum(1 for v in row if v.startswith('*') or v.endswith('*')) / len(row)
            upper_ratio = sum(1 for v in row if v[0].isupper()) / len(row)
            col_ratio   = len(row) / max(df.shape[1], 1)
            sc = short_ratio*0.3 + upper_ratio*0.2 + star_ratio*0.25 + col_ratio*0.25
            if avg_len > 80: sc -= 0.4
            return sc

        def is_maxchar(r: int) -> bool:
            row = [str(v).strip() for v in rows[r] if v is not None and str(v).strip() not in ('nan','None','')]
            if not row: return False
            pats = [re.compile(r'^\d+\s*(chars?|characters?|num)',re.I),
                    re.compile(r'^(varchar|number|date)',re.I),
                    re.compile(r'^yyyy/mm/dd$',re.I)]
            hits = sum(1 for v in row if any(p.match(v) for p in pats))
            return hits / len(row) > 0.3

        def is_legend(r: int) -> bool:
            row = [str(v).strip() for v in rows[r] if v is not None and str(v).strip() not in ('nan','None','')]
            if not row: return True
            return sum(1 for v in row if len(v) > 60) / len(row) > 0.4

        n = len(rows)
        if n >= 4 and (is_legend(0) or is_maxchar(2)) and header_score(3) > 0.25:
            return TemplatePattern.FOUR_ROW_METADATA, 3, 4, [0,1,2]
        if n >= 3 and is_maxchar(1) and header_score(2) > 0.25:
            return TemplatePattern.THREE_ROW_METADATA, 2, 3, [0,1]
        if header_score(0) > 0.25:
            return TemplatePattern.FLAT_HEADER, 0, 1, []
        best = max(range(min(6,n)), key=header_score)
        return TemplatePattern.AI_DETECTED, best, best+1, list(range(best))

# ══════════════════════════════════════════════════════════════════════════════
# LOV EXTRACTOR
# ══════════════════════════════════════════════════════════════════════════════
class LOVExtractor:
    COL_HINTS = {
        "invoice type":"invoice_types","transaction types":"transaction_types",
        "transaction type":"transaction_types","payment method":"payment_methods",
        "line type":"line_types","line types":"line_types",
        "account class":"account_classes","sources":"sources",
        "batch sources":"batch_sources","pay group":"pay_groups",
        "currency_code":"currencies","currency code":"currencies",
        "payment term name":"payment_terms","asset books":"asset_books",
        "assettype":"asset_types","asset type":"asset_types",
        "posting status":"posting_statuses","depreciation method":"depreciation_methods",
        "prorate convention":"prorate_conventions","salvagetype":"salvage_types",
        "bu_name":"bu_names","bu_id":"bu_ids",
        "supplier name":"supplier_names","account_name":"customer_names",
    }

    def extract(self, wb) -> Dict[str, List[str]]:
        lovs: Dict[str, List[str]] = {}
        detector = PatternDetector()
        for sname in wb.sheetnames:
            if not detector.is_reference_sheet(sname): continue
            ws = wb[sname]
            self._from_sheet(ws, sname, lovs)
        return lovs

    def _from_sheet(self, ws, sname: str, lovs: Dict[str, List[str]]):
        maxr = ws.max_row; maxc = ws.max_column
        hrow = None
        for r in range(1, min(5, maxr+1)):
            vals = [ws.cell(r,c).value for c in range(1,maxc+1)]
            nv = [v for v in vals if v is not None and str(v).strip() not in ('','nan')]
            if nv: hrow = r; break
        if hrow is None: return

        for c in range(1, maxc+1):
            hv = ws.cell(hrow,c).value
            if not hv or str(hv).strip() in ('','nan'): continue
            hstr = str(hv).strip()
            key = None
            for hint, k in self.COL_HINTS.items():
                if hint.lower() in hstr.lower(): key = k; break
            if not key:
                key = re.sub(r'[^a-z0-9_]','_', hstr.lower()).strip('_')
            vals = []
            for r in range(hrow+1, maxr+1):
                v = ws.cell(r,c).value
                if v is not None and str(v).strip() not in ('','nan','None'):
                    vals.append(str(v).strip())
            if vals:
                lovs[key] = list(dict.fromkeys(lovs.get(key,[])+vals))

        clean = re.sub(r'[^a-z0-9_]','_',sname.lower()).strip('_')
        if clean not in lovs:
            cvals = []
            for r in range(hrow+1, maxr+1):
                v = ws.cell(r,1).value
                if v is not None and str(v).strip() not in ('','nan','None'):
                    cvals.append(str(v).strip())
            if cvals: lovs[clean] = cvals

# ══════════════════════════════════════════════════════════════════════════════
# AI ENGINE - Azure OpenAI Only
# ══════════════════════════════════════════════════════════════════════════════
class AIEngine:
    def __init__(self):
        self._client = None
        self._ready = False
        self._provider = None
        self._last_test = None
        self._test_result = False
        self._init()

    def _init(self):
        """Initialize Azure OpenAI client"""
        if Config.AZURE_OPENAI_KEY and Config.AZURE_OPENAI_ENDPOINT and Config.AZURE_OPENAI_DEPLOYMENT:
            try:
                from openai import AzureOpenAI
                self._client = AzureOpenAI(
                    api_key=Config.AZURE_OPENAI_KEY,
                    azure_endpoint=Config.AZURE_OPENAI_ENDPOINT,
                    api_version="2024-02-15-preview"
                )
                self._ready = True
                self._provider = "azure"
            except Exception as e:
                self._ready = False
                self._provider = None
        else:
            self._ready = False
            self._provider = None
    
    def test_connection(self) -> bool:
        """
        Test actual connection to Azure OpenAI.
        Returns True if connection works, False otherwise.
        """
        if not self._ready:
            return False
        
        try:
            # Make a minimal API call to test connection
            response = self._client.chat.completions.create(
                model=Config.AZURE_OPENAI_DEPLOYMENT,
                messages=[{"role": "user", "content": "test"}],
                max_tokens=5,
                temperature=0
            )
            self._test_result = True
            self._last_test = datetime.now()
            return True
        except Exception as e:
            self._test_result = False
            self._last_test = datetime.now()
            print(f"AI Connection Test Failed: {e}")
            return False

    def configure(self, endpoint: str, api_key: str, deployment: str) -> bool:
        """Configure Azure OpenAI credentials and connect"""
        try:
            Config.AZURE_OPENAI_ENDPOINT = endpoint
            Config.AZURE_OPENAI_KEY = api_key
            Config.AZURE_OPENAI_DEPLOYMENT = deployment
            
            from openai import AzureOpenAI
            self._client = AzureOpenAI(
                api_key=api_key,
                azure_endpoint=endpoint,
                api_version="2024-02-15-preview"
            )
            self._ready = True
            self._provider = "azure"
            return True
        except Exception as e:
            self._ready = False
            self._provider = None
            return False

    @property
    def available(self) -> bool:
        return self._ready

    @property
    def provider(self) -> str:
        return self._provider or "none"

    def chat(self, system: str, user: str, max_tokens: int = 3000) -> str:
        if not self.available: return ""
        try:
            r = self._client.chat.completions.create(
                model=Config.AZURE_OPENAI_DEPLOYMENT,
                messages=[{"role":"system","content":system},{"role":"user","content":user}],
                temperature=0.1,
                max_tokens=max_tokens
            )
            return r.choices[0].message.content
        except Exception as e:
            return f"__AI_ERROR__:{e}"
        return ""

    def detect_mandatory_fields(self, fields_info: List[Dict]) -> Dict[str, Dict]:
        """AI-based mandatory field detection"""
        prompt = f"""You are an Oracle Fusion ERP expert. Analyze these template fields and determine if each is Mandatory or Optional.

Fields to analyze:
{json.dumps(fields_info[:50], indent=2)}

Look for indicators:
1. Field name starts or ends with '*'
2. Description contains: "mandatory", "required", "must be provided", "cannot be null"
3. Field name patterns: ID fields, Date fields, Amount fields typically mandatory
4. Oracle Fusion standard mandatory fields

Return ONLY valid JSON:
{{
  "Field Name": {{
    "status": "Mandatory" or "Optional",
    "confidence": 0.0 to 1.0,
    "reason": "brief explanation"
  }}
}}
"""
        raw = self.chat("Oracle Fusion ERP mandatory field detection expert. Return ONLY valid JSON.", prompt, 2500)
        if not raw or "__AI_ERROR__" in raw: return {}
        try:
            clean = raw.replace("```json","").replace("```","").strip()
            return json.loads(clean)
        except:
            return {}

    def suggest_mappings(self, src: List[FieldMetadata], tgt: List[FieldMetadata], tpl_type: str) -> List[Dict]:
        ss = [{"id":f.field_id,"name":f.name,"type":f.data_type,"samples":f.sample_values[:3]} for f in src]
        ts = [{"id":f.field_id,"name":f.name,"type":f.data_type,"mandatory":f.is_mandatory,"desc":f.description[:80]} for f in tgt]
        prompt = f"""Oracle Fusion template: {tpl_type}
Source fields (legacy data): {json.dumps(ss[:30],indent=2)}
Target fields (Fusion template): {json.dumps(ts[:30],indent=2)}

Map source → target. Available transformations: none, uppercase, lowercase, trim,
date_yyyy_mm_dd, date_standardize, numeric_clean, remove_thousand_separator,
make_positive, make_negative, document_type_to_invoice_type,
document_type_to_transaction_type, static_external, static_standard,
truncate_30, truncate_100, email_normalize, phone_format, proper_case, extract_number

Return ONLY JSON array:
[{{"target_id":"..","source_id":"..","confidence":0.0,"reasoning":"..","transformation":"none"}}]"""
        raw = self.chat("Oracle Fusion data migration expert. Return ONLY valid JSON array.", prompt, 3000)
        if not raw or "__AI_ERROR__" in raw: return []
        try:
            return json.loads(raw.replace("```json","").replace("```","").strip())
        except:
            return []

    def fallback_match(self, src: List[FieldMetadata], tgt: List[FieldMetadata]) -> List[Dict]:
        KNOWN = {
            "vendor name":"supplier name","vendor id":"supplier number",
            "document type":"invoice type","external document no.":"invoice number",
            "amount":"invoice amount","date":"invoice date",
            "posting date":"invoice date","description":"description",
            "entry no.":"invoice id","customer name":"bill-to customer account name",
        }
        XFORM = {"invoice date":"date_yyyy_mm_dd","invoice amount":"numeric_clean",
                 "invoice type":"document_type_to_invoice_type"}
        res = []
        for t in tgt:
            tl = t.name.lower()
            best = None; bconf = 0.0; bxf = "none"
            if tl in KNOWN:
                for s in src:
                    if s.name.lower() == KNOWN[tl]:
                        best = s; bconf = 0.92; bxf = XFORM.get(tl,"none"); break
            if not best:
                for s in src:
                    sc = SequenceMatcher(None, tl, s.name.lower()).ratio()
                    if t.data_type == s.data_type: sc += 0.1
                    if sc > bconf and sc > 0.45: bconf = sc; best = s
            if best:
                res.append({"target_id":t.field_id,"source_id":best.field_id,
                            "confidence":round(min(bconf,1.0),2),
                            "reasoning":"Heuristic / name similarity","transformation":bxf})
        return res

# ══════════════════════════════════════════════════════════════════════════════
# ENTERPRISE VALIDATION ENGINE - ENHANCED WITH DATE VALIDATION
# ══════════════════════════════════════════════════════════════════════════════
class EnterpriseValidator:
    """Centralized enterprise validation engine with comprehensive date support"""
    
    # Date format mapping from display format to strptime format
    DATE_FORMAT_MAPPING = {
        "DD/MM/YYYY": "%d/%m/%Y",
        "DD-MM-YYYY": "%d-%m-%Y",
        "MM/DD/YYYY": "%m/%d/%Y",
        "MM-DD-YYYY": "%m-%d-%Y",
        "YYYY-MM-DD": "%Y-%m-%d",
        "YYYY/MM/DD": "%Y/%m/%d",
        "DD/MM/YY": "%d/%m/%y",
        "DD-MM-YY": "%d-%m-%y",
        "MM/DD/YY": "%m/%d/%y",
        "MM-DD-YY": "%m-%d-%y",
        "YYYY-MM-DD (ISO 8601)": "%Y-%m-%d",
        "DD/MM/YYYY HH:MM": "%d/%m/%Y %H:%M",
        "DD/MM/YYYY HH:MM:SS": "%d/%m/%Y %H:%M:%S",
        "YYYY-MM-DD HH:MM:SS": "%Y-%m-%d %H:%M:%S",
        "DD/MM/YYYY HH:MM AM/PM": "%d/%m/%Y %I:%M %p",
        "MM/DD/YYYY HH:MM AM/PM": "%m/%d/%Y %I:%M %p",
        "Oracle: DD-MON-YYYY": "%d-%b-%Y",
        "SQL Server: YYYY-MM-DD HH:MM:SS.SSS": "%Y-%m-%d %H:%M:%S.%f",
    }
    
    @staticmethod
    def validate_field_value(value: Any, rule: EnterpriseValidationRule) -> Tuple[bool, str]:
        """Validate a single value against enterprise rule"""
        if value is None or (isinstance(value, float) and np.isnan(value)):
            value_str = ""
        else:
            value_str = str(value).strip()
        
        # Mandatory Check is usually handled by MappingValidator, 
        # but internal logic might need to check nulls.
        
        # Text Constraints
        if rule.data_type == "Text":
            if rule.no_special_chars and value_str:
                if not re.match(r'^[a-zA-Z0-9\s]*$', value_str):
                    return False, f"Field '{rule.field_name}': Special characters are not allowed"
            
            if rule.min_length and len(value_str) < rule.min_length:
                return False, f"Field '{rule.field_name}': Length {len(value_str)} is below minimum {rule.min_length}"
            
            if rule.max_length and len(value_str) > rule.max_length:
                return False, f"Field '{rule.field_name}': Length {len(value_str)} exceeds maximum {rule.max_length}"

            if rule.allowed_values and value_str:
                if value_str not in rule.allowed_values:
                    return False, f"Field '{rule.field_name}': Value '{value_str}' not in allowed list"
        
        # Regex (Global)
        if rule.regex_pattern and value_str:
            try:
                if not re.match(rule.regex_pattern, value_str):
                    return False, rule.custom_error_message or f"Field '{rule.field_name}': Pattern mismatch"
            except:
                return False, f"Field '{rule.field_name}': Invalid regex pattern"

        # Number Constraints
        if rule.data_type == "Number":
            if value_str:
                try:
                    # Clean number string
                    clean_str = re.sub(r'[^\d.\-]', '', value_str.replace(',',''))
                    num_val = float(clean_str)
                    
                    if not rule.decimal_allowed and "." in clean_str and float(clean_str).is_integer() == False:
                        return False, f"Field '{rule.field_name}': Decimals are not allowed"
                    
                    if rule.min_value is not None and num_val < rule.min_value:
                        return False, f"Field '{rule.field_name}': Value {num_val} is below minimum {rule.min_value}"
                    if rule.max_value is not None and num_val > rule.max_value:
                        return False, f"Field '{rule.field_name}': Value {num_val} exceeds maximum {rule.max_value}"
                        
                except Exception as e:
                    return False, f"Field '{rule.field_name}': Invalid number format"

        # Date Constraints
        if rule.data_type == "Date":
            if value_str:
                if rule.accept_all_formats:
                    # Simple check if it looks like a date
                    date_obj = None
                    for fmt in ["%Y-%m-%d", "%Y/%m/%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y"]:
                        try:
                            date_obj = datetime.strptime(value_str, fmt)
                            break
                        except: pass
                    if not date_obj:
                        return False, f"Field '{rule.field_name}': Could not identify date format"
                else:
                    date_obj = EnterpriseValidator._parse_date(value_str, rule.date_format)
                    if date_obj is None:
                        return False, f"Field '{rule.field_name}': Expected format {rule.date_format}"
                
                if date_obj:
                    if rule.date_cannot_be_future and date_obj > datetime.now():
                        return False, f"Field '{rule.field_name}': Future dates not allowed"
                    if rule.date_cannot_be_past and date_obj < datetime.now().replace(hour=0, minute=0, second=0):
                        return False, f"Field '{rule.field_name}': Past dates not allowed"

        return True, ""
        
        return True, ""
    
    @staticmethod
    def _parse_date(date_str: str, date_format: str) -> Optional[datetime]:
        """Parse date string according to specified format"""
        if not date_str:
            return None
        
        # Get strptime format
        strptime_format = EnterpriseValidator.DATE_FORMAT_MAPPING.get(date_format)
        
        # Handle special formats
        if date_format == "Unix Timestamp (seconds)":
            try:
                return datetime.fromtimestamp(int(date_str))
            except:
                return None
        elif date_format == "Unix Timestamp (milliseconds)":
            try:
                return datetime.fromtimestamp(int(date_str) / 1000)
            except:
                return None
        elif date_format.startswith("YYYY-MM-DDTHH:MM:SS"):
            # ISO format variants
            try:
                # Remove timezone info for parsing
                clean_str = date_str.split('+')[0].split('Z')[0]
                return datetime.fromisoformat(clean_str)
            except:
                return None
        elif date_format == "ISO UTC Format":
            try:
                return datetime.fromisoformat(date_str.replace('Z', '+00:00'))
            except:
                return None
        
        if strptime_format:
            try:
                parsed_date = datetime.strptime(date_str, strptime_format)
                # Validate actual calendar date (e.g., no 31/02/2026)
                # strptime already does this, but let's be explicit
                return parsed_date
            except ValueError:
                return None
        
        return None
    
    @staticmethod
    def test_sample_value(sample_value: str, rule: EnterpriseValidationRule) -> Tuple[bool, str]:
        """Test a sample value against rule for UI testing"""
        return EnterpriseValidator.validate_field_value(sample_value, rule)

# ══════════════════════════════════════════════════════════════════════════════
# NEW: DATA QUALITY ENGINE
# ══════════════════════════════════════════════════════════════════════════════

class DataQualityEngine:
    """Template-driven data quality rule engine"""
    
    def __init__(self):
        self.quality_rules: Dict[str, List[DataQualityRule]] = {}  # field_id -> rules
        self.country_mappings: List[CountryStandardization] = []
        self._init_default_country_mappings()
    
    def _init_default_country_mappings(self):
        """Initialize default country standardization mappings"""
        default_mappings = [
            CountryStandardization("uae_mapping", 
                ["uae", "u.a.e", "united arab emirates", "emirates", "u.a.e.", "UAE"], "UAE"),
            CountryStandardization("usa_mapping",
                ["usa", "u.s.a", "united states", "america", "u.s.a.", "US"], "USA"),
            CountryStandardization("uk_mapping",
                ["uk", "u.k", "united kingdom", "britain", "england", "GB"], "UK"),
            CountryStandardization("india_mapping",
                ["india", "ind", "hindustan", "bharat", "IN"], "INDIA"),
            CountryStandardization("dubai_mapping",
                ["dubai", "dxb", "dubai emirate"], "DUBAI"),
            CountryStandardization("saudi_mapping",
                ["saudi arabia", "saudi", "ksa", "kingdom of saudi arabia", "SA"], "SAUDI ARABIA"),
        ]
        self.country_mappings.extend(default_mappings)
    
    def add_quality_rule(self, field_id: str, rule: DataQualityRule):
        """Add a quality rule for a field"""
        if field_id not in self.quality_rules:
            self.quality_rules[field_id] = []
        self.quality_rules[field_id].append(rule)
    
    def get_quality_rules(self, field_id: str) -> List[DataQualityRule]:
        """Get all quality rules for a field"""
        return self.quality_rules.get(field_id, [])
    
    def remove_quality_rule(self, field_id: str, rule_id: str):
        """Remove a specific quality rule"""
        if field_id in self.quality_rules:
            self.quality_rules[field_id] = [r for r in self.quality_rules[field_id] if r.rule_id != rule_id]
    
    def apply_transformations(self, value: Any, field_id: str) -> Tuple[Any, List[str], List[str]]:
        """Apply all transformation rules to a value"""
        if value is None:
            return value, [], []
        
        transformed_value = str(value)
        applied_rules = []
        errors = []
        
        rules = self.get_quality_rules(field_id)
        
        for rule in rules:
            if not rule.is_active:
                continue
                
            try:
                # Apply transformations in order
                if rule.trim_spaces:
                    transformed_value = transformed_value.strip()
                    applied_rules.append("Trim Spaces")
                
                if rule.to_uppercase:
                    transformed_value = transformed_value.upper()
                    applied_rules.append("Convert to Uppercase")
                
                if rule.to_lowercase:
                    transformed_value = transformed_value.lower()
                    applied_rules.append("Convert to Lowercase")
                
                if rule.remove_special_chars:
                    transformed_value = re.sub(r'[^A-Za-z0-9\s]', '', transformed_value)
                    applied_rules.append("Remove Special Characters")
                
                if rule.replace_underscore_with_space:
                    transformed_value = transformed_value.replace('_', ' ')
                    applied_rules.append("Replace Underscore with Space")
                
                if rule.no_latin_chars:
                    # Remove Latin characters like é, ñ, ü, etc.
                    import unicodedata
                    transformed_value = ''.join(c for c in unicodedata.normalize('NFD', transformed_value) 
                                              if unicodedata.category(c) != 'Mn')
                    applied_rules.append("Remove Latin Characters")
                
                if rule.regex_pattern and rule.regex_replacement:
                    transformed_value = re.sub(rule.regex_pattern, rule.regex_replacement, transformed_value)
                    applied_rules.append(f"Regex: {rule.regex_description or 'Custom Pattern'}")
                
                # Apply country standardization
                transformed_value = self._apply_country_standardization(transformed_value, applied_rules)
                
            except Exception as e:
                errors.append(f"Transformation error in {rule.rule_type.value}: {str(e)}")
        
        return transformed_value, applied_rules, errors
    
    def _apply_country_standardization(self, value: str, applied_rules: List[str]) -> str:
        """Apply country standardization mappings"""
        value_lower = value.lower().strip()
        
        for mapping in self.country_mappings:
            if not mapping.is_active:
                continue
            
            for source_val in mapping.source_values:
                if value_lower == source_val.lower():
                    applied_rules.append(f"Country Standardization: {value} → {mapping.target_value}")
                    return mapping.target_value
        
        return value
    
    def validate_value(self, value: Any, field_id: str) -> Tuple[bool, List[str], List[str]]:
        """Validate a value against all rules"""
        if value is None:
            value_str = ""
        else:
            value_str = str(value).strip()
        
        errors = []
        warnings = []
        rules = self.get_quality_rules(field_id)
        
        for rule in rules:
            if not rule.is_active:
                continue
            
            try:
                # Completeness validation
                if rule.not_null and value is None:
                    errors.append(f"Value cannot be null")
                
                if rule.not_blank and not value_str:
                    errors.append(f"Value cannot be blank")
                
                # Uniqueness validation (Note: This requires dataset context for full validation)
                if rule.unique_values or rule.no_duplicates:
                    warnings.append(f"Uniqueness validation requires full dataset context")
                
                # Format validation
                if rule.alphanumeric_only and value_str:
                    if not re.match(r'^[A-Za-z0-9\s]*$', value_str):
                        errors.append(f"Value must be alphanumeric only")
                
                if rule.no_special_chars and value_str:
                    if re.search(r'[^A-Za-z0-9\s]', value_str):
                        warnings.append(f"Value contains special characters")
                
                # Character length validation
                if rule.min_length and len(value_str) < rule.min_length:
                    errors.append(f"Value length {len(value_str)} is below minimum {rule.min_length}")
                
                if rule.max_length and len(value_str) > rule.max_length:
                    errors.append(f"Value length {len(value_str)} exceeds maximum {rule.max_length}")
                
                # Date format validation
                if rule.date_format_pattern and value_str:
                    if not self._validate_date_format(value_str, rule.date_format_pattern):
                        errors.append(f"Date format does not match {rule.date_format_pattern}")
                
                # Regex validation
                if rule.regex_pattern and value_str:
                    if not re.match(rule.regex_pattern, value_str):
                        errors.append(f"Value does not match pattern: {rule.regex_description or rule.regex_pattern}")
                
            except Exception as e:
                errors.append(f"Validation error: {str(e)}")
        
        is_valid = len(errors) == 0
        return is_valid, errors, warnings
    
    def _validate_date_format(self, value: str, format_pattern: str) -> bool:
        """Validate date format"""
        format_map = {
            "YYYY/MM/DD": "%Y/%m/%d",
            "DD/MM/YYYY": "%d/%m/%Y",
            "MM/DD/YYYY": "%m/%d/%Y",
            "YYYY-MM-DD": "%Y-%m-%d",
            "DD-MM-YYYY": "%d-%m-%Y"
        }
        
        python_format = format_map.get(format_pattern, "%Y/%m/%d")
        
        try:
            datetime.strptime(value, python_format)
            return True
        except:
            return False
    
    def generate_preview(self, values: List[Any], field_id: str) -> List[TransformationPreview]:
        """Generate transformation preview for a list of values"""
        previews = []
        
        for value in values[:10]:  # Limit to first 10 values
            if value is None or str(value).strip() in ('', 'nan', 'None'):
                continue
            
            # Apply transformations
            transformed_value, applied_rules, transform_errors = self.apply_transformations(value, field_id)
            
            # Validate transformed value
            is_valid, validation_errors, warnings = self.validate_value(transformed_value, field_id)
            
            # Determine status
            if validation_errors:
                status = "Invalid"
            elif applied_rules:
                status = "Corrected"
            else:
                status = "Valid"
            
            all_errors = transform_errors + validation_errors
            
            preview = TransformationPreview(
                source_value=str(value),
                transformed_value=str(transformed_value),
                validation_status=status,
                applied_rules=applied_rules,
                errors=all_errors,
                warnings=warnings
            )
            previews.append(preview)
        
        return previews
    
    def validate_dataset(self, df: pd.DataFrame, field_mappings: Dict[str, Any]) -> Dict[str, Dict]:
        """Validate entire dataset against quality rules"""
        validation_results = {}
        
        for field_id, mapping in field_mappings.items():
            if not mapping.is_active or mapping.source_field not in df.columns:
                continue
            
            field_results = {
                'total_records': len(df),
                'valid_records': 0,
                'invalid_records': 0,
                'errors': [],
                'warnings': [],
                'unique_violations': []
            }
            
            rules = self.get_quality_rules(field_id)
            column_data = df[mapping.source_field].dropna()
            
            # Check uniqueness rules first (dataset-level validation)
            for rule in rules:
                if not rule.is_active:
                    continue
                
                if rule.unique_values or rule.no_duplicates:
                    duplicates = column_data[column_data.duplicated(keep=False)]
                    if not duplicates.empty:
                        field_results['unique_violations'] = duplicates.tolist()
                        field_results['errors'].append(f"Found {len(duplicates)} duplicate values")
            
            # Validate each value
            for idx, value in column_data.items():
                is_valid, errors, warnings = self.validate_value(value, field_id)
                
                if is_valid:
                    field_results['valid_records'] += 1
                else:
                    field_results['invalid_records'] += 1
                    field_results['errors'].extend(errors)
                    field_results['warnings'].extend(warnings)
            
            validation_results[field_id] = field_results
        
        return validation_results
    
    def add_country_mapping(self, mapping: CountryStandardization):
        """Add a custom country mapping"""
        self.country_mappings.append(mapping)
    
    def remove_country_mapping(self, mapping_id: str):
        """Remove a country mapping"""
        self.country_mappings = [m for m in self.country_mappings if m.mapping_id != mapping_id]
    
    def remove_country_mapping(self, mapping_id: str):
        """Remove a country mapping"""
        self.country_mappings = [m for m in self.country_mappings if m.mapping_id != mapping_id]

# ══════════════════════════════════════════════════════════════════════════════
# REAL-TIME TRANSFORMATION ENGINE
# ══════════════════════════════════════════════════════════════════════════════
# ══════════════════════════════════════════════════════════════════════════════
# TRANSFORMATION ENGINE - ENHANCED
# ══════════════════════════════════════════════════════════════════════════════
class EnterpriseTransformation:
    """Enterprise-grade transformation engine with chain support"""
    
    @classmethod
    def apply_config(cls, value: Any, config: TransformationConfig, source_row: Dict[str, Any] = None) -> Any:
        """Apply comprehensive configuration to a value"""
        result = value
        
        # 1. Core Logic Selection
        if config.transformation_type == "concatenate" and source_row:
            vals = []
            for fid in config.concatenate_fields:
                vals.append(str(source_row.get(fid, "")))
            result = config.concatenate_separator.join(vals)
        
        elif config.transformation_type == "expression" and config.formula_expression:
            result = cls._apply_formula(config.formula_expression, source_row)
        
        elif config.transformation_type == "condition":
            result = cls._apply_condition(result, config)

        # 2. Text Transformations
        if config.trim_type == "left": result = str(result).lstrip()
        elif config.trim_type == "right": result = str(result).rstrip()
        elif config.trim_type == "anywhere": result = re.sub(r'\s+', ' ', str(result)).strip()
        
        if config.case_type == "upper": result = str(result).upper()
        elif config.case_type == "lower": result = str(result).lower()
        elif config.case_type == "title": result = str(result).title()
        
        if config.remove_special_chars:
            result = re.sub(r'[^a-zA-Z0-9\s]', '', str(result))
            
        if config.replace_text_source:
            result = str(result).replace(config.replace_text_source, config.replace_text_target)
            
        if config.substring_start is not None:
            s_idx = config.substring_start
            e_idx = s_idx + (config.substring_length or 999)
            result = str(result)[s_idx:e_idx]
            
        if config.split_delimiter:
            parts = str(result).split(config.split_delimiter)
            if 0 <= config.split_index < len(parts):
                result = parts[config.split_index]
                
        if config.extract_name_part != "none":
            names = str(result).split()
            if names:
                if config.extract_name_part == "first": result = names[0]
                elif config.extract_name_part == "last": result = names[-1]

        # 3. Date Advanced
        if config.input_date_format or config.add_days or config.subtract_days:
            result = cls._apply_date_advanced(result, config)

        # 4. Null Handling
        if result is None or str(result).strip() in ('', 'nan', 'None'):
            if config.handle_null == "default":
                result = config.default_value
            elif config.handle_null == "error":
                return "[NULL_ERROR]"
        
        return result

    @staticmethod
    def _apply_formula(formula: str, row: Dict[str, Any]) -> Any:
        try:
            processed = formula
            if row:
                for fid, val in row.items():
                    processed = processed.replace(f"{{{fid}}}", str(val))
            return eval(processed, {"__builtins__": {}}, {"abs": abs, "round": round, "min": min, "max": max})
        except:
            return f"[EXPRESSION_ERROR]"

    @staticmethod
    def _apply_condition(val: Any, config: TransformationConfig) -> Any:
        try:
            v_check = str(val)
            if config.if_operator == "equals" and v_check == config.if_value: return config.then_result
            if config.if_operator == "contains" and config.if_value in v_check: return config.then_result
            return config.else_result
        except: return val

    @classmethod
    def _apply_date_advanced(cls, val: Any, config: TransformationConfig) -> Any:
        try:
            date_obj = None
            if config.input_date_format:
                for fmt in ["%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%Y/%m/%d"]:
                    try:
                        date_obj = datetime.strptime(str(val).strip(), fmt)
                        break
                    except: pass
            if not date_obj: date_obj = datetime.now()

            from datetime import timedelta
            if config.add_days: date_obj += timedelta(days=config.add_days)
            if config.subtract_days: date_obj -= timedelta(days=config.subtract_days)
            
            out_fmt = EnterpriseValidator.DATE_FORMAT_MAPPING.get(config.output_date_format, "%Y-%m-%d")
            return date_obj.strftime(out_fmt)
        except: return val
    
    @classmethod
    def apply_chain(cls, value: Any, chain: List[Dict[str, Any]]) -> Any:
        """Apply a chain of transformations sequentially"""
        result = value
        for step in chain:
            transform_name = step.get("transform", "none")
            params = step.get("params", {})
            result = cls.apply(result, transform_name, params)
        return result
    
    @classmethod
    def apply(cls, value: Any, transform: str, params: Dict = None) -> Any:
        """Apply single transformation (Legacy Support)"""
        if value is None: return ""
        if transform == "uppercase": return str(value).upper()
        if transform == "lowercase": return str(value).lower()
        if transform == "trim": return str(value).strip()
        return value
    
    @staticmethod
    def _t_none(x, **k): return x
    @staticmethod
    def _t_trim(x, **k): return str(x).strip()
    @staticmethod
    def _t_uppercase(x, **k): return str(x).upper()
    @staticmethod
    def _t_lowercase(x, **k): return str(x).lower()
    @staticmethod
    def _t_proper_case(x, **k): return str(x).title()
    @staticmethod
    def _t_remove_extra_spaces(x, **k): return re.sub(r'\s+', ' ', str(x)).strip()
    @staticmethod
    def _t_remove_special_chars(x, **k): return re.sub(r'[^a-zA-Z0-9\s\-_.]', '', str(x))
    @staticmethod
    def _t_numeric_clean(x, **k): return re.sub(r'[^\d.\-]', '', str(x).replace(',', ''))
    @staticmethod
    def _t_remove_thousand_separator(x, **k): return str(x).replace(',', '')
    
    @staticmethod
    def _t_currency_2_decimal(x, **k):
        try: return round(float(str(x).replace(',', '').replace('$', '')), 2)
        except: return x
    
    @staticmethod
    def _t_currency_4_decimal(x, **k):
        try: return round(float(str(x).replace(',', '').replace('$', '')), 4)
        except: return x
    
    @staticmethod
    def _t_make_positive(x, **k):
        try: return abs(float(str(x).replace(',', '')))
        except: return x
    
    @staticmethod
    def _t_make_negative(x, **k):
        try: return -abs(float(str(x).replace(',', '')))
        except: return x
    
    @staticmethod
    def _t_null_replace(x, null_replace="", **k):
        if x is None or str(x).strip() in ('', 'nan', 'None', 'NaT', 'NaN'):
            return null_replace
        return x
    
    @staticmethod
    def _t_regex_replace(x, pattern="", replacement="", **k):
        try:
            return re.sub(pattern, replacement, str(x))
        except:
            return x
    
    @staticmethod
    def _t_lookup_map(x, mapping=None, default="", **k):
        mapping = mapping or {}
        return mapping.get(str(x).strip(), default)
    
    @staticmethod
    def _t_conditional_transform(x, condition="", true_transform="none", false_transform="none", **k):
        """Apply transformation based on condition"""
        try:
            if eval(condition, {"__builtins__": {}}, {"x": x}):
                return EnterpriseTransformation.apply(x, true_transform, k)
            else:
                return EnterpriseTransformation.apply(x, false_transform, k)
        except:
            return x
    
    @staticmethod
    def _t_extract_number(x, **k):
        m = re.findall(r'-?\d+\.?\d*', str(x).replace(',', ''))
        return m[0] if m else "0"
    
    @staticmethod
    def _t_pad_left_zeros(x, length=10, **k): return str(x).zfill(int(length))
    
    @staticmethod
    def _t_truncate_30(x, **k): return str(x)[:30]
    @staticmethod
    def _t_truncate_50(x, **k): return str(x)[:50]
    @staticmethod
    def _t_truncate_100(x, **k): return str(x)[:100]
    
    @staticmethod
    def _t_email_normalize(x, **k): return str(x).lower().strip()
    
    @staticmethod
    def _t_phone_standardize(x, **k):
        digits = re.sub(r'\D', '', str(x))
        return digits
    
    @staticmethod
    def _t_boolean_standardize(x, **k):
        v = str(x).lower().strip()
        return "TRUE" if v in ('true', 'yes', '1', 'y', 't') else "FALSE" if v in ('false', 'no', '0', 'n', 'f') else str(x)
    
    @staticmethod
    def _t_static_external(x, **k): return "External"
    @staticmethod
    def _t_static_standard(x, **k): return "STANDARD"
    @staticmethod
    def _t_static_aed(x, **k): return "AED"
    @staticmethod
    def _t_static_net30(x, **k): return "Net 30"
    @staticmethod
    def _t_static_yes(x, **k): return "Y"
    @staticmethod
    def _t_static_no(x, **k): return "N"
    
    @classmethod
    def _t_date_standardize(cls, x, **k): return cls._t_date_yyyy_mm_dd(x, **k)
    
    @classmethod
    def _t_date_yyyy_mm_dd(cls, x, **k):
        for f in ["%Y/%m/%d", "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y",
                  "%d-%m-%Y", "%m-%d-%Y", "%d-%b-%Y", "%d %b %Y", "%Y%m%d"]:
            try:
                return datetime.strptime(str(x).strip(), f).strftime("%Y-%m-%d")
            except:
                pass
        return str(x)
    
    @classmethod
    def _t_date_dd_mm_yyyy(cls, x, **k):
        for f in ["%Y/%m/%d", "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y",
                  "%d-%m-%Y", "%m-%d-%Y", "%d-%b-%Y", "%d %b %Y", "%Y%m%d"]:
            try:
                return datetime.strptime(str(x).strip(), f).strftime("%d/%m/%Y")
            except:
                pass
        return str(x)
    
    @classmethod
    def _t_date_mm_dd_yyyy(cls, x, **k):
        for f in ["%Y/%m/%d", "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y",
                  "%d-%m-%Y", "%m-%d-%Y", "%d-%b-%Y", "%d %b %Y", "%Y%m%d"]:
            try:
                return datetime.strptime(str(x).strip(), f).strftime("%m/%d/%Y")
            except:
                pass
        return str(x)

# ══════════════════════════════════════════════════════════════════════════════
# MAPPING VALIDATOR - STRICT ENTERPRISE VALIDATION
# ══════════════════════════════════════════════════════════════════════════════
class MappingValidator:
    """Validates source-to-template mapping against all enterprise rules"""
    
    def __init__(self, validator: EnterpriseValidator):
        self.validator = validator
    
    def validate_mapping(self, mapping: FieldMapping, 
                        source_value: Any,
                        all_source_data: Dict[str, Any] = None) -> Tuple[bool, List[str], Any]:
        """
        Validate a mapping attempt
        Returns: (is_valid: bool, error_messages: List[str], transformed_value: Any)
        """
        errors = []
        
        # Step 1: Apply transformations
        if mapping.transformation_chain:
            transformed = EnterpriseTransformation.apply_chain(source_value, mapping.transformation_chain)
        else:
            transformed = EnterpriseTransformation.apply(source_value, mapping.transformation, mapping.transform_params)
        
        # Step 2: Check mandatory condition
        for rule in mapping.template_rules:
            if isinstance(rule, EnterpriseValidationRule):
                if hasattr(rule, 'mandatory_status') and rule.mandatory_status == "Mandatory":
                    if self._is_empty(transformed):
                        errors.append(f"❌ Mandatory field violation: Value is required but empty")
                        return False, errors, transformed
        
        # Step 3: Validate against all template rules
        for rule in mapping.template_rules:
            if isinstance(rule, EnterpriseValidationRule):
                passed, msg = self.validator.validate_field_value(transformed, rule)
                if not passed:
                    errors.append(f"❌ {rule.field_name}: {msg}")
                    return False, errors, transformed
        
        if errors:
            return False, errors, transformed
        
        return True, ["✅ All validations passed"], transformed
    
    def _is_empty(self, value: Any) -> bool:
        if value is None: return True
        if isinstance(value, float) and np.isnan(value): return True
        return str(value).strip() in ('', 'nan', 'None', 'NaT', 'NaN', '<NA>')
    
    def preview_transformation(self, mapping: FieldMapping, source_value: Any, all_source_data: Dict[str, Any] = None) -> Dict[str, Any]:
        """Generate preview of transformation and validation"""
        is_valid, messages, transformed = self.validate_mapping(mapping, source_value, all_source_data)
        
        return {
            "source_value": source_value,
            "after_transformation": transformed,
            "final_value": transformed,
            "validation_status": "✅ Valid" if is_valid else "❌ Invalid",
            "validation_messages": messages,
            "is_valid": is_valid
        }

# ══════════════════════════════════════════════════════════════════════════════
# SHEET ANALYSER - Enhanced with AI Mandatory Detection
# ══════════════════════════════════════════════════════════════════════════════
class SheetAnalyser:
    def __init__(self, lovs: Dict[str,List[str]], ai: AIEngine):
        self.lovs = lovs
        self.ai = ai
        self.detector = PatternDetector()

    def analyse(self, raw_df: pd.DataFrame, sheet_name: str,
                wb=None, forced_header: Optional[int]=None) -> SheetProfile:
        if raw_df is None or raw_df.empty:
            return SheetProfile(sheet_name, TemplatePattern.FLAT_HEADER, 0, 1, [])

        if forced_header is not None:
            pat = TemplatePattern.AI_DETECTED
            hdr = forced_header; ds = forced_header+1; meta = list(range(forced_header))
        else:
            pat, hdr, ds, meta = self.detector.detect(raw_df)

        fields: Dict[str,FieldMetadata] = {}
        hrow = raw_df.iloc[hdr]
        
        # Prepare data for AI mandatory detection
        fields_for_ai = []
        for ci, cv in enumerate(hrow):
            if cv is None or str(cv).strip() in ('','nan','None'): continue
            cs = str(cv).strip()
            name = cs.strip('*').strip()
            
            # Get description from metadata rows
            desc = ""
            for mr in meta:
                if mr < len(raw_df) and ci < raw_df.shape[1]:
                    v = raw_df.iloc[mr, ci]
                    if v is not None and str(v).strip() not in ('','nan','None') and len(str(v)) > 10:
                        desc = str(v).strip()[:250]
                        break
            
            fields_for_ai.append({
                "field_name": name,
                "original_name": cs,
                "has_asterisk": cs.startswith('*') or cs.endswith('*'),
                "description": desc
            })
        
        # AI Mandatory Detection
        ai_mandatory_results = {}
        if self.ai.available and fields_for_ai:
            ai_mandatory_results = self.ai.detect_mandatory_fields(fields_for_ai)
        
        # Build fields with AI results
        seen_names: Dict[str,int] = {}
        for ci, cv in enumerate(hrow):
            if cv is None or str(cv).strip() in ('','nan','None'): continue
            cs = str(cv).strip()
            has_asterisk = cs.startswith('*') or cs.endswith('*')
            name = cs.strip('*').strip()

            # Deduplicate
            if name in seen_names:
                seen_names[name] += 1
                unique_name = f"{name} ({seen_names[name]})"
            else:
                seen_names[name] = 1
                unique_name = name

            # Get AI mandatory status
            ai_result = ai_mandatory_results.get(unique_name, ai_mandatory_results.get(name, {}))
            ai_status = ai_result.get("status", "")
            ai_confidence = ai_result.get("confidence", 0.0)
            ai_reason = ai_result.get("reason", "")
            
            # Determine final mandatory status
            if has_asterisk:
                mandatory_status = "Mandatory"
                is_mandatory = True
                confidence = 1.0
            elif ai_status == "Mandatory":
                mandatory_status = "Mandatory"
                is_mandatory = True
                confidence = ai_confidence
            elif ai_status == "Optional":
                mandatory_status = "Optional"
                is_mandatory = False
                confidence = ai_confidence
            else:
                mandatory_status = "Optional"
                is_mandatory = False
                confidence = 0.0
            
            # Get other metadata
            desc = ""
            for mr in meta:
                if mr < len(raw_df) and ci < raw_df.shape[1]:
                    v = raw_df.iloc[mr, ci]
                    if v is not None and str(v).strip() not in ('','nan','None') and len(str(v)) > 10:
                        desc = str(v).strip()[:250]
                        break
            
            # Get samples
            srows = raw_df.iloc[ds:ds+5] if ds < len(raw_df) else pd.DataFrame()
            samples = []
            for _, row in srows.iterrows():
                if ci < len(row):
                    v = str(row.iloc[ci]).strip()
                    if v not in ('','nan','None'): samples.append(v)
            
            # Detect data type
            data_type = self._detect_data_type(samples, desc)
            
            fid = f"tpl_{ci}_{hashlib.md5(unique_name.encode()).hexdigest()[:8]}"
            
            # Create enterprise rule
            enterprise_rule = EnterpriseValidationRule(
                field_id=fid,
                field_name=unique_name,
                data_type=data_type,
                is_ai_detected=bool(ai_status),
                mandatory_status=mandatory_status
            )
            
            fields[fid] = FieldMetadata(
                field_id=fid,
                name=unique_name,
                data_type=data_type,
                sample_values=samples[:5],
                is_mandatory=is_mandatory,
                mandatory_status=mandatory_status,
                description=desc,
                enterprise_rule=enterprise_rule,
                column_index=ci,
                ai_mandatory_confidence=confidence
            )

        return SheetProfile(sheet_name=sheet_name, pattern=pat,
                            header_row_index=hdr, data_start_index=ds,
                            metadata_rows=meta, fields=fields)
    
    def _detect_data_type(self, samples, desc) -> str:
        """Detect data type from samples"""
        if not samples: return "Text"
        
        desc_lower = desc.lower() if desc else ""
        
        # Check description hints
        if "date" in desc_lower: return "Date"
        if "email" in desc_lower: return "Email"
        if any(k in desc_lower for k in ["amount", "price", "cost", "currency"]): return "Currency"
        if "number" in desc_lower: return "Number"
        
        # Check samples
        date_pattern = re.compile(r'^\d{1,4}[/\-]\d{1,2}[/\-]\d{2,4}$')
        date_matches = sum(1 for s in samples if date_pattern.match(s.strip()))
        if date_matches / len(samples) > 0.6: return "Date"
        
        numeric_matches = 0
        for s in samples:
            try:
                float(s.replace(",","").replace("$",""))
                numeric_matches += 1
            except:
                pass
        if numeric_matches / len(samples) > 0.7: return "Number"
        
        return "Text"

# ══════════════════════════════════════════════════════════════════════════════
# TRANSFORMATIONS (Legacy - kept for compatibility)
# ══════════════════════════════════════════════════════════════════════════════
class TLib:
    T = {
        "none":"No transformation",
        "trim":"Trim whitespace",
        "uppercase":"→ UPPERCASE",
        "lowercase":"→ lowercase",
        "proper_case":"→ Title Case",
        "remove_extra_spaces":"Collapse spaces",
        "date_standardize":"Auto-detect → YYYY/MM/DD",
        "date_yyyy_mm_dd":"→ YYYY/MM/DD",
        "date_dd_mm_yyyy":"→ DD/MM/YYYY",
        "date_mm_dd_yyyy":"→ MM/DD/YYYY",
        "numeric_clean":"Remove non-numeric chars",
        "remove_thousand_separator":"Remove , separator",
        "make_positive":"abs() → positive",
        "make_negative":"Negate → negative",
        "currency_standard":"Round to 2 decimals",
        "decimal_2":"Round to 2 decimals",
        "decimal_4":"Round to 4 decimals",
        "email_normalize":"Lowercase + trim",
        "phone_format":"Standardise phone",
        "truncate_30":"Truncate to 30 chars",
        "truncate_50":"Truncate to 50 chars",
        "truncate_100":"Truncate to 100 chars",
        "pad_left_zeros":"Pad zeros to 10",
        "extract_number":"Extract numeric portion",
        "null_to_empty":"NULL → empty string",
        "empty_to_null":"Empty → NULL",
        "boolean_standardize":"→ TRUE/FALSE",
    }

    @classmethod
    def apply(cls, v: Any, t: str, p: Dict = None) -> Any:
        if v is None: return v
        if isinstance(v, float) and np.isnan(v): return v
        fn = getattr(cls, f"_t_{t}", None)
        if fn:
            try: return fn(v, **(p or {}))
            except: return v
        return v

    @staticmethod
    def _t_none(x,**k): return x
    @staticmethod
    def _t_trim(x,**k): return str(x).strip()
    @staticmethod
    def _t_uppercase(x,**k): return str(x).upper()
    @staticmethod
    def _t_lowercase(x,**k): return str(x).lower()
    @staticmethod
    def _t_proper_case(x,**k): return str(x).title()
    @staticmethod
    def _t_remove_extra_spaces(x,**k): return re.sub(r'\s+',' ',str(x)).strip()
    @staticmethod
    def _t_numeric_clean(x,**k): return re.sub(r'[^\d.\-]','',str(x).replace(',',''))
    @staticmethod
    def _t_remove_thousand_separator(x,**k): return str(x).replace(',','')
    @staticmethod
    def _t_make_positive(x,**k):
        try: return abs(float(str(x).replace(',','')))
        except: return x
    @staticmethod
    def _t_make_negative(x,**k):
        try: return -abs(float(str(x).replace(',','')))
        except: return x
    @staticmethod
    def _t_currency_standard(x,**k): return TLib._t_decimal_2(x,**k)
    @staticmethod
    def _t_decimal_2(x,**k):
        try: return round(float(str(x).replace(',','')), 2)
        except: return x
    @staticmethod
    def _t_decimal_4(x,**k):
        try: return round(float(str(x).replace(',','')), 4)
        except: return x
    @staticmethod
    def _t_email_normalize(x,**k): return str(x).lower().strip()
    @staticmethod
    def _t_truncate_30(x,**k): return str(x)[:30]
    @staticmethod
    def _t_truncate_50(x,**k): return str(x)[:50]
    @staticmethod
    def _t_truncate_100(x,**k): return str(x)[:100]
    @staticmethod
    def _t_pad_left_zeros(x,length=10,**k): return str(x).zfill(int(length))
    @staticmethod
    def _t_extract_number(x,**k):
        m = re.findall(r'-?\d+\.?\d*', str(x).replace(',',''))
        return m[0] if m else "0"
    @staticmethod
    def _t_null_to_empty(x,**k): return "" if x is None else x
    @staticmethod
    def _t_empty_to_null(x,**k): return None if str(x).strip() == "" else x
    @staticmethod
    def _t_boolean_standardize(x,**k):
        v=str(x).lower().strip()
        return "TRUE" if v in('true','yes','1','y') else "FALSE" if v in('false','no','0','n') else str(x)
    @classmethod
    def _t_date_standardize(cls,x,**k): return cls._pd(str(x),"%Y/%m/%d")
    @classmethod
    def _t_date_yyyy_mm_dd(cls,x,**k): return cls._pd(str(x),"%Y/%m/%d")
    @classmethod
    def _t_date_dd_mm_yyyy(cls,x,**k): return cls._pd(str(x),"%d/%m/%Y")
    @classmethod
    def _t_date_mm_dd_yyyy(cls,x,**k): return cls._pd(str(x),"%m/%d/%Y")
    @staticmethod
    def _pd(v,fmt):
        for f in ["%Y/%m/%d","%Y-%m-%d","%d/%m/%Y","%m/%d/%Y",
                  "%d-%m-%Y","%m-%d-%Y","%d-%b-%Y","%d %b %Y","%Y%m%d"]:
            try: return datetime.strptime(str(v).strip(),f).strftime(fmt)
            except: pass
        return str(v)

# ══════════════════════════════════════════════════════════════════════════════
# MAPPING ENGINE - ENHANCED WITH REAL-TIME TRANSFORMATION
# ══════════════════════════════════════════════════════════════════════════════
class MappingEngine:
    def __init__(self):
        self.ai = AIEngine()
        self.lovs:      Dict[str,List[str]] = {}
        self.profiles:  Dict[str,SheetProfile] = {}
        self.active:    str = ""
        self.tpl_type:  str = "Oracle Fusion"
        self.all_sheets:  List[str] = []
        self.data_sheets: List[str] = []
        self.source_df:   Optional[pd.DataFrame] = None
        self.src_fields:  Dict[str,FieldMetadata] = {}
        self.mappings:    Dict[str,FieldMapping] = {}
        self.last_report: Optional[DataQualityReport] = None
        self._validator = EnterpriseValidator()
        self._mapping_validator = MappingValidator(self._validator)
        self._transformer = EnterpriseTransformation()
        self._analyser: Optional[SheetAnalyser] = None
        self._suggestions: List[Dict] = []
        self._wb = None
        # ✅ NEW: Data Quality Engine
        self._quality_engine = DataQualityEngine()
    
    def __getstate__(self):
        """Exclude non-picklable items from state"""
        state = self.__dict__.copy()
        state['_wb'] = None
        state['_analyser'] = None
        state['_validator'] = None
        state['_mapping_validator'] = None
        state['ai'] = None
        return state

    def __setstate__(self, state):
        """Restore state and re-initialize components"""
        self.__dict__.update(state)
        # Re-initialize non-picklable components
        self.ai = AIEngine()
        self._validator = EnterpriseValidator()
        self._mapping_validator = MappingValidator(self._validator)
        self._transformer = EnterpriseTransformation()
        self._analyser = SheetAnalyser(self.lovs, self.ai)
        # ✅ NEW: Re-initialize quality engine
        if not hasattr(self, '_quality_engine'):
            self._quality_engine = DataQualityEngine()

    @property
    def tpl_fields(self) -> Dict[str,FieldMetadata]:
        p = self.profiles.get(self.active)
        return p.fields if p else {}

    def load_template(self, f) -> Dict[str,Any]:
        import openpyxl, io
        raw = f.read(); f.seek(0)
        st.session_state["_tpl_bytes"] = raw
        
        try:
            wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True, read_only=False)
        except Exception as e:
            wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True, read_only=False)
        
        self._wb = wb
        
        visible_sheets = []
        for sheet_name in wb.sheetnames:
            try:
                ws = wb[sheet_name]
                if not hasattr(ws, 'sheet_state') or ws.sheet_state != 'hidden':
                    visible_sheets.append(sheet_name)
            except Exception:
                visible_sheets.append(sheet_name)
        
        self.all_sheets = visible_sheets if visible_sheets else wb.sheetnames
        self.all_sheets = list(dict.fromkeys(self.all_sheets))
        
        det = PatternDetector()
        self.data_sheets = [s for s in self.all_sheets if not det.is_reference_sheet(s)]
        
        self.lovs = LOVExtractor().extract(wb)
        self._analyser = SheetAnalyser(self.lovs, self.ai)
        
        fname = getattr(f,'name','').lower()
        if "supplier" in fname:                       self.tpl_type = "Supplier Master"
        elif "ar" in fname or "receivable" in fname:  self.tpl_type = "AR Invoices"
        elif "asset" in fname or "fa" in fname:       self.tpl_type = "Fixed Assets"
        elif "ap" in fname or "payable" in fname:     self.tpl_type = "AP Invoices"
        else: self.tpl_type = self.all_sheets[0] if self.all_sheets else "Oracle Fusion"
        
        if self.all_sheets: self.active = self.all_sheets[0]
        
        return {
            "all_sheets": self.all_sheets,
            "data_sheets": self.data_sheets,
            "lovs": {k:len(v) for k,v in self.lovs.items()},
            "type": self.tpl_type
        }

    def analyse_sheet(self, sheet: str, forced_hdr: Optional[int]=None) -> SheetProfile:
        import io, openpyxl
        raw = st.session_state.get("_tpl_bytes")
        if not raw:
            return SheetProfile(sheet, TemplatePattern.FLAT_HEADER, 0, 1, [])
        
        # Restore workbook if missing (after state load)
        if self._wb is None:
            try:
                self._wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
            except: pass
            
        df = pd.read_excel(io.BytesIO(raw), sheet_name=sheet, header=None, engine="openpyxl")
        prof = self._analyser.analyse(df, sheet, wb=self._wb, forced_header=forced_hdr)
        self.profiles[sheet] = prof
        if sheet == self.active: self.mappings = {}
        return prof

    def load_source(self, f, sheet: str, hdr: int, cols: List[str]) -> Dict[str,Any]:
        import io
        raw = f.read(); f.seek(0)
        st.session_state["_src_bytes"] = raw
        st.session_state["_src_meta"] = {"sheet": sheet, "hdr": hdr, "cols": cols,
                                          "is_csv": getattr(f,'name','').lower().endswith('.csv')}
        is_csv = st.session_state["_src_meta"]["is_csv"]
        if is_csv:
            df = pd.read_csv(io.BytesIO(raw))
        else:
            df = pd.read_excel(io.BytesIO(raw), sheet_name=sheet, header=hdr, engine="openpyxl")
        if cols: df = df[[c for c in cols if c in df.columns]]
        self.source_df = df
        self.src_fields = {}
        for i, col in enumerate(df.columns):
            smp = df[col].dropna().head(10).astype(str).tolist()
            fid = f"src_{i}_{hashlib.md5(str(col).encode()).hexdigest()[:8]}"
            dt = "Text"
            st_ = {"total":len(df[col]),"nulls":int(df[col].isna().sum()),
                   "unique":int(df[col].nunique()),
                   "null_pct":round(df[col].isna().sum()/len(df[col])*100,1)}
            self.src_fields[fid] = FieldMetadata(fid,str(col),dt,smp[:5],statistics=st_,column_index=i)
        return {"fields":len(self.src_fields),"records":len(df)}

    def suggest(self) -> List[Dict]:
        tgt = list(self.tpl_fields.values()); src = list(self.src_fields.values())
        if not tgt or not src: return []
        if self.ai.available:
            s = self.ai.suggest_mappings(src, tgt, self.tpl_type)
            if s: self._suggestions = s; return s
        fb = self.ai.fallback_match(src, tgt)
        self._suggestions = fb; return fb

    def save_map(self, tid, sid, xf="none", params=None, notes="", confidence=0.0, ai_suggested=False,
                 transform_config: TransformationConfig = None) -> FieldMapping:
        tf = self.tpl_fields.get(tid); sf = self.src_fields.get(sid)
        if not tf or not sf: raise ValueError(f"Invalid field IDs: {tid}, {sid}")
        
        # In v4 logic, template_rules is used by MappingValidator
        template_rules = []
        if tf.enterprise_rule:
            template_rules.append(tf.enterprise_rule)
        
        if transform_config is None:
            transform_config = TransformationConfig()
        
        m = FieldMapping(
            f"map_{tid}", tid, tf.name, sf.name, sid, 
            transformation=xf,
            transform_params=params or {},
            transformation_chain=[], # Initialize empty chain
            template_rules=template_rules,
            user_notes=notes,
            mapping_confidence=confidence,
            ai_suggested=ai_suggested,
            transform_config=transform_config
        )
        self.mappings[tid] = m
        return m

    def update_mapping_transform(self, tid: str, transform_config: TransformationConfig):
        """Update transformation configuration for existing mapping"""
        if tid in self.mappings:
            self.mappings[tid].transform_config = transform_config
            # Regenerate preview
            if self.source_df is not None:
                # Use MappingValidator for preview in v4 logic
                sample_val = self.source_df[self.mappings[tid].source_field].dropna().iloc[0] if not self.source_df[self.mappings[tid].source_field].dropna().empty else ""
                self.mappings[tid].preview_sample = self._mapping_validator.preview_transformation(
                    self.mappings[tid],
                    sample_val
                )

    def remove_map(self, tid):
        if tid in self.mappings: del self.mappings[tid]

    def validate_mapping(self, mapping: FieldMapping, sample_size: int = 5) -> Tuple[bool, List[str]]:
        if self.source_df is None or mapping.source_field not in self.source_df.columns:
            return False, ["Source data error"]
        
        errors = []
        samples = self.source_df[mapping.source_field].dropna().head(sample_size)
        
        for idx, raw_val in samples.items():
            is_valid, messages, _ = self._mapping_validator.validate_mapping(mapping, raw_val)
            if not is_valid:
                errors.extend(messages)
                break
        
        mapping.mapping_validation_status = "Valid" if not errors else "Invalid"
        mapping.mapping_validation_errors = errors
        return not errors, errors

    def generate(self, fix=True) -> pd.DataFrame:
            """Legacy method - generates DataFrame only (deprecated for template preservation)"""
            if self.source_df is None: raise ValueError("Load source data first")
            out = {}
            for tid, fm in self.tpl_fields.items():
                m = self.mappings.get(tid)
                if m and m.is_active and m.source_field in self.source_df.columns:
                    # Apply transformation chain in v4 logic
                    if m.transformation_chain:
                        transformed_series = self.source_df[m.source_field].apply(
                            lambda v: EnterpriseTransformation.apply_chain(v, m.transformation_chain)
                        )
                    else:
                        transformed_series = self.source_df[m.source_field].apply(
                            lambda v: EnterpriseTransformation.apply(v, m.transformation, m.transform_params)
                        )

                    # ✅ NEW: Apply quality rules transformation
                    quality_transformed_series = transformed_series.apply(
                        lambda v: self._quality_engine.apply_transformations(v, tid)[0]
                    )
                    out[fm.name] = quality_transformed_series
                else:
                    out[fm.name] = pd.Series([""]*len(self.source_df))
            df = pd.DataFrame(out)
            cols = [f.name for f in self.tpl_fields.values() if f.name in df.columns]
            return df[cols]

    
    def generate_with_template_preservation(self) -> bytes:
        """
        🔥 NEW: Generate output preserving EXACT template format
        - Preserves all sheets (only modifies active sheet)
        - Preserves all formatting, colors, styles
        - Preserves formulas, merged cells, column widths
        - Only fills data rows based on mappings
        """
        import openpyxl, io
        from openpyxl.utils import get_column_letter
        from copy import copy
        
        if self.source_df is None:
            raise ValueError("Load source data first")
        
        # Load original template workbook
        raw = st.session_state.get("_tpl_bytes")
        if not raw:
            raise ValueError("Template file not found in session")
        
        # Load workbook preserving ALL formatting
        wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=False)
        
        # Get active sheet profile
        profile = self.profiles.get(self.active)
        if not profile:
            raise ValueError(f"Sheet {self.active} not analyzed")
        
        ws = wb[self.active]
        
        # Get data start row (where we'll write data)
        data_start_row = profile.data_start_index + 1  # +1 for 1-based indexing
        header_row = profile.header_row_index + 1
        
        # Clear existing data rows (keep header and metadata)
        max_row = ws.max_row
        if max_row > data_start_row:
            ws.delete_rows(data_start_row, max_row - data_start_row + 1)
        
        # Build column mapping: field_id -> column_index
        field_to_col = {}
        for field_id, field_meta in self.tpl_fields.items():
            field_to_col[field_id] = field_meta.column_index + 1  # +1 for 1-based
        
        # Process each source row
        for src_row_idx in range(len(self.source_df)):
            target_row_idx = data_start_row + src_row_idx
            
            # For each template field
            for field_id, field_meta in self.tpl_fields.items():
                col_idx = field_to_col.get(field_id)
                if not col_idx:
                    continue
                
                # Get mapping for this field
                mapping = self.mappings.get(field_id)
                
                if mapping and mapping.is_active:
                    # Get source value
                    source_col = mapping.source_field
                    if source_col in self.source_df.columns:
                        source_value = self.source_df.iloc[src_row_idx][source_col]
                        
                        # Apply transformations
                        if mapping.transform_config:
                            source_row_dict = self.source_df.iloc[src_row_idx].to_dict()
                            transformed_value = EnterpriseTransformation.apply_config(
                                source_value, 
                                mapping.transform_config, 
                                source_row_dict
                            )
                        elif mapping.transformation_chain:
                            transformed_value = EnterpriseTransformation.apply_chain(
                                source_value, 
                                mapping.transformation_chain
                            )
                        else:
                            transformed_value = EnterpriseTransformation.apply(
                                source_value, 
                                mapping.transformation, 
                                mapping.transform_params
                            )
                        
                        # ✅ NEW: Apply quality rules transformation
                        transformed_value, applied_rules, errors = self._quality_engine.apply_transformations(
                            transformed_value, field_id
                        )
                        
                        # Write to cell
                        cell = ws.cell(row=target_row_idx, column=col_idx)
                        cell.value = transformed_value
                        
                        # Copy style from header row to maintain formatting
                        if header_row < target_row_idx:
                            header_cell = ws.cell(row=header_row, column=col_idx)
                            if header_cell.has_style:
                                cell.font = copy(header_cell.font)
                                cell.border = copy(header_cell.border)
                                cell.fill = copy(header_cell.fill)
                                cell.number_format = copy(header_cell.number_format)
                                cell.protection = copy(header_cell.protection)
                                cell.alignment = copy(header_cell.alignment)
                else:
                    # No mapping - leave empty or use default
                    cell = ws.cell(row=target_row_idx, column=col_idx)
                    cell.value = ""
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

# ══════════════════════════════════════════════════════════════════════════════
# STATE PERSISTENCE (DISK-BASED)
# ══════════════════════════════════════════════════════════════════════════════
STATE_FILE = "fusion_state.pkl"

def save_app_state():
    """🔥 ENHANCED: Serialize entire application state to disk with robust error handling"""
    try:
        eng = st.session_state.get("eng")
        if not eng: 
            return
        
        # Comprehensive state capture
        state = {
            "eng": eng,
            "tpl_ok": st.session_state.get("tpl_ok", False),
            "src_ok": st.session_state.get("src_ok", False),
            "suggestions": st.session_state.get("suggestions", []),
            "_tpl_bytes": st.session_state.get("_tpl_bytes"),
            "_src_bytes": st.session_state.get("_src_bytes"),
            "_src_meta": st.session_state.get("_src_meta"),
            "ai_config": {
                "endpoint": Config.AZURE_OPENAI_ENDPOINT,
                "key": Config.AZURE_OPENAI_KEY,
                "deployment": Config.AZURE_OPENAI_DEPLOYMENT
            },
            # 🔥 NEW: Save all mapping-related session state
            "enterprise_rules": st.session_state.get("enterprise_rules", {}),
            "transform_configs": st.session_state.get("transform_configs", {}),
            "preview_data": st.session_state.get("preview_data", {}),
            # Metadata for verification
            "_save_timestamp": datetime.now().isoformat(),
            "_version": "v7.0_template_preservation"
        }
        
        # Use atomic write: Save to temp then rename
        temp_file = STATE_FILE + ".tmp"
        with open(temp_file, "wb") as f:
            pickle.dump(state, f, protocol=pickle.HIGHEST_PROTOCOL)
        
        # Verify temp file was written
        if os.path.exists(temp_file) and os.path.getsize(temp_file) > 0:
            # Atomic rename (safe on all platforms)
            if os.path.exists(STATE_FILE):
                # Keep backup of previous state
                backup_file = STATE_FILE + ".backup"
                try:
                    if os.path.exists(backup_file):
                        os.remove(backup_file)
                    os.rename(STATE_FILE, backup_file)
                except:
                    pass
            
            os.rename(temp_file, STATE_FILE)
            
            # ✅ Track last auto-save time for UI indicator
            st.session_state["_last_auto_save"] = datetime.now()
            
            return True
        else:
            print(f"DEBUG: Temp file not created or empty")
            return False
            
    except Exception as e:
        # Don't show in sidebar as it might vanish on rerun
        print(f"DEBUG: Save failed: {e}")
        import traceback
        traceback.print_exc()
        return False

def load_app_state():
    """🔥 FORCE LOAD: Restore application state from disk with multiple fallbacks"""
    
    # ✅ Try main state file first
    if os.path.exists(STATE_FILE):
        try:
            with open(STATE_FILE, "rb") as f:
                state = pickle.load(f)
            
            # ✅ FORCE RESTORE: Restore ALL state including files
            for k, v in state.items():
                if k == "ai_config" and v:
                    # Restore AI config (but secrets.toml takes precedence)
                    if not Config.AZURE_OPENAI_ENDPOINT:
                        Config.AZURE_OPENAI_ENDPOINT = v.get("endpoint", "")
                    if not Config.AZURE_OPENAI_KEY:
                        Config.AZURE_OPENAI_KEY = v.get("key", "")
                    if not Config.AZURE_OPENAI_DEPLOYMENT:
                        Config.AZURE_OPENAI_DEPLOYMENT = v.get("deployment", "")
                elif k.startswith("_") and k not in ["_tpl_bytes", "_src_bytes", "_src_meta", "_save_timestamp", "_version"]:
                    # Skip internal metadata except file bytes
                    continue
                else:
                    st.session_state[k] = v
            
            # ✅ CRITICAL: Restore file bytes so uploads persist
            if "_tpl_bytes" in state and state["_tpl_bytes"]:
                st.session_state["_tpl_bytes"] = state["_tpl_bytes"]
                st.session_state["tpl_ok"] = True
            
            if "_src_bytes" in state and state["_src_bytes"]:
                st.session_state["_src_bytes"] = state["_src_bytes"]
                st.session_state["src_ok"] = True
            
            if "_src_meta" in state:
                st.session_state["_src_meta"] = state["_src_meta"]
            
            # Log successful load
            save_time = state.get("_save_timestamp", "unknown")
            version = state.get("_version", "unknown")
            print(f"✅ State loaded successfully (saved: {save_time}, version: {version})")
            return True
            
        except Exception as e:
            print(f"❌ Main state load failed: {e}")
            import traceback
            traceback.print_exc()
            
            # Try backup file
            backup_file = STATE_FILE + ".backup"
            if os.path.exists(backup_file):
                try:
                    with open(backup_file, "rb") as f:
                        state = pickle.load(f)
                    
                    for k, v in state.items():
                        if k == "ai_config" and v:
                            Config.AZURE_OPENAI_ENDPOINT = v.get("endpoint", "")
                            Config.AZURE_OPENAI_KEY = v.get("key", "")
                            Config.AZURE_OPENAI_DEPLOYMENT = v.get("deployment", "")
                        elif not k.startswith("_") or k in ["_tpl_bytes", "_src_bytes", "_src_meta"]:
                            st.session_state[k] = v
                    
                    st.sidebar.success("✅ Restored from backup")
                    return True
                except Exception as e2:
                    st.sidebar.error(f"❌ Backup load also failed: {e2}")
    
    return False

def clear_application():
    """Clear all application data and persistence file"""
    keys_to_clear = [
        "eng", "tpl_ok", "src_ok", "suggestions", "show_test_field",
        "enterprise_rules", "_tpl_bytes", "_src_bytes", "_src_meta",
        "selected_target_field", "mapping_panel_open", "transform_configs"
    ]
    for key in keys_to_clear:
        if key in st.session_state:
            del st.session_state[key]
    
    if os.path.exists(STATE_FILE):
        try: os.remove(STATE_FILE)
        except: pass

# ══════════════════════════════════════════════════════════════════════════════
# UI STYLES - ENHANCED FOR REAL-TIME INTERFACE
# ══════════════════════════════════════════════════════════════════════════════
def apply_css():
    st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=IBM+Plex+Mono:wght@400;600&family=DM+Sans:wght@400;500;600;700&display=swap');

:root {
    --bg: #f8fafc;
    --surface: #ffffff;
    --border: #e2e8f0;
    --text: #0f172a;
    --text-muted: #64748b;
    --primary: #2563eb;
    --primary-light: #eff6ff;
    --success: #16a34a;
    --success-light: #f0fdf4;
    --error: #dc2626;
    --error-light: #fef2f2;
    --warning: #d97706;
    --warning-light: #fffbeb;
    --ai: #7c3aed;
    --ai-light: #f5f3ff;
    --radius: 12px;
    --shadow-sm: 0 1px 2px 0 rgba(0, 0, 0, 0.05);
    --shadow-md: 0 4px 6px -1px rgba(0, 0, 0, 0.1), 0 2px 4px -1px rgba(0, 0, 0, 0.06);
    --animate-speed: 0.3s;
}

.animate-fade { animation: fadeIn var(--animate-speed) ease-out; }
@keyframes fadeIn { from { opacity: 0; transform: translateY(5px); } to { opacity: 1; transform: translateY(0); } }

.live-indicator {
    display: inline-flex;
    align-items: center;
    gap: 6px;
    padding: 4px 12px;
    background: #f0fdf4;
    color: #166534;
    border-radius: 999px;
    font-size: 0.7rem;
    font-weight: 700;
    text-transform: uppercase;
    letter-spacing: 0.05em;
    border: 1px solid #bbf7d0;
}

.live-indicator::before {
    content: "";
    width: 6px;
    height: 6px;
    background: #22c55e;
    border-radius: 50%;
    display: inline-block;
    animation: pulse 2s infinite;
}

@keyframes pulse {
    0% { transform: scale(0.95); box-shadow: 0 0 0 0 rgba(34, 197, 94, 0.7); }
    70% { transform: scale(1); box-shadow: 0 0 0 6px rgba(34, 197, 94, 0); }
    100% { transform: scale(0.95); box-shadow: 0 0 0 0 rgba(34, 197, 94, 0); }
}

.validation-badge {
    font-size: 0.65rem;
    font-weight: 700;
    padding: 2px 6px;
    border-radius: 4px;
    text-transform: uppercase;
}
.badge-passed { background: #dcfce7; color: #166534; }
.badge-failed { background: #fee2e2; color: #991b1b; }

.mapping-row {
    padding: 12px 16px;
    border-radius: 8px;
    border: 1px solid var(--border);
    background: white;
    margin-bottom: 8px;
    transition: all 0.2s;
    display: flex;
    align-items: center;
    gap: 16px;
}

.mapping-row:hover {
    box-shadow: var(--shadow-md);
    border-color: var(--primary);
}

.mapping-row.valid { background-color: var(--success-light); border-color: #86efac; }
.mapping-row.invalid { background-color: var(--error-light); border-color: #fca5a5; }
.mapping-row.mandatory-missing { background-color: var(--warning-light); border-color: #fde047; }

.field-label {
    font-weight: 600;
    font-size: 0.95rem;
    color: var(--text);
}

.field-sublabel {
    font-size: 0.75rem;
    color: var(--text-muted);
}

.badge {
    padding: 2px 8px;
    border-radius: 4px;
    font-size: 0.7rem;
    font-weight: 600;
    text-transform: uppercase;
}

.badge-blue { background: #dbeafe; color: #1e40af; }
.badge-green { background: #dcfce7; color: #166534; }
.badge-red { background: #fee2e2; color: #991b1b; }
.badge-gray { background: #f1f5f9; color: #475569; }
.badge-orange { background: #ffedd5; color: #9a3412; }

.preview-box {
    font-family: 'IBM Plex Mono', monospace;
    font-size: 0.8rem;
    padding: 4px 8px;
    border-radius: 6px;
    background: #f8fafc;
    border: 1px solid #e2e8f0;
    max-width: 200px;
    overflow: hidden;
    text-overflow: ellipsis;
    white-space: nowrap;
}

.stTabs [data-baseweb="tab"] {
    height: 40px;
    white-space: pre;
    background-color: transparent;
    border-radius: 6px;
    color: var(--text-muted);
    font-weight: 500;
    transition: all 0.2s;
}

.stTabs [data-baseweb="tab"]:hover {
    background-color: var(--primary-light);
    color: var(--primary);
}

.stTabs [aria-selected="true"] {
    background-color: var(--primary) !important;
    color: white !important;
}

/* Excel Style Grid */
.excel-grid-container {
    background: white;
    border: 1px solid var(--border);
    border-radius: var(--radius);
    overflow: hidden;
    margin-bottom: 24px;
    box-shadow: var(--shadow-sm);
}

.excel-grid-header {
    background: #f8fafc;
    border-bottom: 2px solid var(--border);
    display: flex;
    position: sticky;
    top: 0;
    z-index: 10;
}

.excel-grid-row {
    display: flex;
    border-bottom: 1px solid var(--border);
    transition: background 0.1s;
}

.excel-grid-row:hover {
    background: #f1f5f9;
}

.excel-grid-cell {
    padding: 12px 16px;
    border-right: 1px solid var(--border);
    font-size: 0.85rem;
    overflow: hidden;
    text-overflow: ellipsis;
    white-space: nowrap;
    display: flex;
    align-items: center;
}

.excel-grid-cell:last-child {
    border-right: none;
}

.excel-header-cell {
    font-weight: 700;
    color: #475569;
    text-transform: uppercase;
    font-size: 0.72rem;
    letter-spacing: 0.05em;
    background: #f1f5f9;
}

/* Mapping Panel Layout */
.mapping-grid {
    display: grid;
    grid-template-columns: 1fr 1.5fr 1fr;
    gap: 20px;
    height: calc(100vh - 250px);
}

.mapping-section {
    background: white;
    border: 1px solid var(--border);
    border-radius: var(--radius);
    display: flex;
    flex-direction: column;
    overflow: hidden;
    box-shadow: var(--shadow-sm);
}

.section-header {
    padding: 16px;
    background: #f8fafc;
    border-bottom: 1px solid var(--border);
    font-weight: 700;
    color: #334155;
    display: flex;
    align-items: center;
    justify-content: space-between;
}

.section-content {
    flex: 1;
    overflow-y: auto;
    padding: 16px;
}

/* Field Cards */
.field-card {
    padding: 12px;
    border: 1px solid var(--border);
    border-radius: 8px;
    margin-bottom: 8px;
    cursor: pointer;
    transition: all 0.2s;
    background: white;
}

.field-card:hover {
    border-color: var(--primary);
    box-shadow: 0 2px 4px rgba(37, 99, 235, 0.1);
    transform: translateY(-1px);
}

.field-card.selected {
    border-color: var(--primary);
    background: var(--primary-light);
    border-width: 2px;
}

.field-card.mandatory {
    border-left: 4px solid var(--error);
}

.field-card-title {
    font-weight: 600;
    font-size: 0.9rem;
    margin-bottom: 4px;
    display: flex;
    align-items: center;
    gap: 6px;
}

.field-card-meta {
    font-size: 0.75rem;
    color: var(--text-muted);
}

/* Transformation UI */
.trans-pill {
    display: inline-flex;
    align-items: center;
    padding: 4px 10px;
    background: #f1f5f9;
    border-radius: 999px;
    font-size: 0.75rem;
    font-weight: 500;
    color: #475569;
    margin-right: 6px;
    margin-bottom: 6px;
    border: 1px solid #e2e8f0;
}

.trans-pill.active {
    background: var(--primary-light);
    color: var(--primary);
    border-color: #bfdbfe;
}

/* Live Preview Table */
.preview-table {
    width: 100%;
    border-collapse: collapse;
    font-size: 0.8rem;
    background: #f8fafc;
    border-radius: 8px;
    overflow: hidden;
}

.preview-table th {
    text-align: left;
    padding: 8px 12px;
    background: #e2e8f0;
    color: #475569;
    font-weight: 700;
}

.preview-table td {
    padding: 8px 12px;
    border-bottom: 1px solid #e2e8f0;
}

.preview-table tr:last-child td {
    border-bottom: none;
}

.preview-val-old { color: var(--text-muted); text-decoration: line-through; }
.preview-val-new { color: var(--success); font-weight: 600; font-family: 'IBM Plex Mono', monospace; }

/* Status Badges */
.badge {
    padding: 2px 8px;
    border-radius: 4px;
    font-size: 0.65rem;
    font-weight: 700;
    text-transform: uppercase;
}
.badge-red { background: #fee2e2; color: #dc2626; }
.badge-green { background: #dcfce7; color: #16a34a; }
.badge-blue { background: #dbeafe; color: #1d4ed8; }

/* Animations */
@keyframes fadeIn {
    from { opacity: 0; transform: translateY(4px); }
    to { opacity: 1; transform: translateY(0); }
}

.animate-fade {
    animation: fadeIn 0.3s ease-out;
}

.field-selector.selected {
    border-color: #2563eb;
    background: #eff6ff;
}

.field-selector.mandatory {
    border-left: 4px solid #dc2626;
}

.field-badge {
    display: inline-block;
    padding: 2px 8px;
    border-radius: 12px;
    font-size: 0.7rem;
    font-weight: 600;
    margin-left: 8px;
}

.badge-mandatory { background: #fee2e2; color: #dc2626; }
.badge-optional { background: #f0fdf4; color: #16a34a; }
.badge-mapped { background: #dcfce7; color: #16a34a; }

.transformation-section {
    background: #fafbfc;
    border: 1px solid #e2e8f0;
    border-radius: 6px;
    padding: 12px;
    margin-bottom: 12px;
}

.transformation-section h4 {
    margin: 0 0 12px 0;
    font-size: 0.9rem;
    color: #374151;
    font-weight: 600;
}

.value-mapping-row {
    display: flex;
    gap: 8px;
    align-items: center;
    margin-bottom: 8px;
}

.arrow-icon {
    color: #6b7280;
    font-weight: 600;
}

.live-indicator {
    display: inline-flex;
    align-items: center;
    gap: 6px;
    color: #16a34a;
    font-size: 0.8rem;
    font-weight: 600;
}

.live-indicator::before {
    content: "";
    width: 8px;
    height: 8px;
    background: #16a34a;
    border-radius: 50%;
    animation: pulse 2s infinite;
}

@keyframes pulse {
    0% { opacity: 1; }
    50% { opacity: 0.5; }
    100% { opacity: 1; }
}

.validation-alert {
    padding: 10px 14px;
    border-radius: 6px;
    margin: 8px 0;
    font-size: 0.85rem;
}

.validation-alert.error {
    background: #fef2f2;
    border-left: 4px solid #dc2626;
    color: #dc2626;
}

.validation-alert.warning {
    background: #fffbeb;
    border-left: 4px solid #d97706;
    color: #d97706;
}

.enterprise-header {
    background: linear-gradient(135deg, #1e3a8a 0%, #7c3aed 100%);
    border-radius: 12px;
    padding: 24px 32px;
    margin-bottom: 20px;
    color: white;
}

.enterprise-header h1 {
    font-size: 1.6rem;
    font-weight: 800;
    margin: 0;
    color: white;
}

.stat-card {
    background: var(--surface);
    border: 1px solid var(--border);
    border-radius: var(--radius);
    padding: 16px;
    box-shadow: var(--shadow);
}

.stat-card .val {
    font-size: 1.8rem;
    font-weight: 700;
    font-family: 'IBM Plex Mono', monospace;
}

.validation-badge {
    padding: 3px 10px;
    border-radius: 12px;
    font-size: 0.75rem;
    font-weight: 600;
    display: inline-block;
}

.badge-passed { background: #dcfce7; color: #16a34a; }
.badge-failed { background: #fee2e2; color: #dc2626; }
.badge-not-tested { background: #f3f4f6; color: #6b7280; }

.enterprise-table {
    width: 100%;
    border-collapse: collapse;
    font-size: 0.85rem;
}

.enterprise-table th {
    background: #f8fafc;
    padding: 10px;
    text-align: left;
    font-weight: 700;
    font-size: 0.72rem;
    text-transform: uppercase;
    letter-spacing: 0.05em;
    color: #64748b;
    border-bottom: 2px solid #e2e8f0;
}

.enterprise-table td {
    padding: 10px;
    border-bottom: 1px solid #f1f5f9;
}

.enterprise-table tr:hover {
    background: #fafbff;
}

.error-box {
    background: var(--error-light);
    border: 1px solid #fca5a5;
    border-left: 4px solid var(--error);
    padding: 12px 16px;
    border-radius: 6px;
    margin: 8px 0;
}

.success-box {
    background: var(--success-light);
    border: 1px solid #86efac;
    border-left: 4px solid var(--success);
    padding: 12px 16px;
    border-radius: 6px;
    margin: 8px 0;
}

.info-box {
    background: var(--primary-light);
    border: 1px solid #bfdbfe;
    border-left: 4px solid var(--primary);
    padding: 12px 16px;
    border-radius: 6px;
    margin: 8px 0;
}

.test-panel {
    background: #f0fdf4;
    border: 2px solid #86efac;
    border-radius: 8px;
    padding: 20px;
    margin: 16px 0;
}

.validation-section {
    background: #fafbfc;
    border: 1px solid #e2e8f0;
    border-radius: 8px;
    padding: 16px;
    margin: 12px 0;
}
</style>
""", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
# INIT
# ══════════════════════════════════════════════════════════════════════════════
def init():
    """🔥 FORCE STATE PERSISTENCE: Always try to restore state on page load"""
    
    # ✅ CRITICAL: Always try to load state first, even if eng exists
    # This ensures data persists across page refreshes
    if "eng" not in st.session_state or st.session_state.get("_force_reload", False):
        loaded = load_app_state()
        if not loaded:
            st.session_state.eng = MappingEngine()
        st.session_state["_force_reload"] = False
    
    # ✅ FORCE: If state file exists but eng is empty, reload it
    if os.path.exists(STATE_FILE):
        eng = st.session_state.get("eng")
        if eng and not eng.tpl_fields and not eng.src_fields:
            # Engine exists but is empty - try to reload from disk
            load_app_state()
    
    defaults = {
        "tpl_ok": False,
        "src_ok": False,
        "suggestions": [],
        "show_test_field": None,
        "enterprise_rules": {},
        "test_panel_open": {},
        "selected_target_field": None,
        "mapping_panel_open": False,
        "transform_configs": {},
        "preview_data": {},
        "value_mappings_temp": [],
        "conditional_rules_temp": [],
        "_last_auto_save": datetime.now(),
        "_auto_save_enabled": True,
        # ✅ File persistence flags
        "template_file_persisted": False,
        "source_file_persisted": False,
        "template_file_name": "",
        "source_file_name": "",
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v
    
    # ✅ RESTORE FILE PERSISTENCE FLAGS from loaded state
    if st.session_state.get("tpl_ok") and "_tpl_bytes" in st.session_state:
        st.session_state["template_file_persisted"] = True
    if st.session_state.get("src_ok") and "_src_bytes" in st.session_state:
        st.session_state["source_file_persisted"] = True
    
    # ✅ FORCE AUTO-SAVE: Save every 30 seconds if data exists
    if st.session_state.get("_auto_save_enabled", True):
        last_save = st.session_state.get("_last_auto_save", datetime.now())
        if (datetime.now() - last_save).seconds > 30:
            eng = st.session_state.get("eng")
            if eng and (st.session_state.get("tpl_ok") or st.session_state.get("src_ok")):
                save_app_state()
                st.session_state["_last_auto_save"] = datetime.now()

# ══════════════════════════════════════════════════════════════════════════════
# SIDEBAR
# ══════════════════════════════════════════════════════════════════════════════
def sidebar():
    """🔥 ENHANCED: Sidebar with state persistence indicators"""
    eng = st.session_state.eng
    with st.sidebar:
        st.markdown("""
<div style="padding:12px 0 8px; border-bottom:1px solid #1e293b; margin-bottom:12px;">
  <div style="font-size:1.1rem;font-weight:800;color:#fff;">🔮 Fusion Mapper</div>
  <div style="font-size:0.7rem;color:#64748b;margin-top:2px;">Enterprise Edition v7.0 RT</div>
</div>""", unsafe_allow_html=True)

        ai_ok = eng.ai.available
        
        # ✅ Test actual AI connection
        if ai_ok and "ai_connection_tested" not in st.session_state:
            with st.spinner("Testing AI connection..."):
                ai_connected = eng.ai.test_connection()
                st.session_state["ai_connection_tested"] = True
                st.session_state["ai_connection_status"] = ai_connected
        else:
            ai_connected = st.session_state.get("ai_connection_status", False)
        
        # Show connection status
        if ai_ok and ai_connected:
            status_color = "#dcfce7"
            border_color = "#86efac"
            status_text = "🟢 Connected & Tested"
        elif ai_ok and not ai_connected:
            status_color = "#fef3c7"
            border_color = "#fcd34d"
            status_text = "🟡 Configured (Not Tested)"
        else:
            status_color = "#fee2e2"
            border_color = "#fca5a5"
            status_text = "🔴 Not Configured"
        
        st.markdown(f"""
<div style="margin:10px 0 14px;">
  <div style="font-size:0.68rem;text-transform:uppercase;color:#64748b;margin-bottom:5px;">Azure AI Engine</div>
  <div style="padding:6px 12px;background:{status_color};
       border:1px solid {border_color};border-radius:6px;font-size:0.78rem;">
    {status_text}
  </div>
</div>""", unsafe_allow_html=True)
        
        # Add test button if configured but not tested
        if ai_ok and not ai_connected:
            if st.button("🔄 Test Connection", use_container_width=True, type="secondary"):
                with st.spinner("Testing..."):
                    result = eng.ai.test_connection()
                    st.session_state["ai_connection_status"] = result
                    if result:
                        st.success("✅ Connection successful!")
                    else:
                        st.error("❌ Connection failed. Check credentials.")
                    st.rerun()

        # ✅ AI Configuration from secrets.toml (no UI needed)
        if ai_ok:
            st.markdown("""
<div style="padding:6px 10px;background:#dcfce7;border:1px solid #86efac;border-radius:6px;font-size:0.7rem;margin:8px 0;">
  ✅ AI configured from secrets.toml
</div>""", unsafe_allow_html=True)
        else:
            st.markdown("""
<div style="padding:6px 10px;background:#fef3c7;border:1px solid #fcd34d;border-radius:6px;font-size:0.7rem;margin:8px 0;">
  ⚠️ Configure Azure OpenAI in .streamlit/secrets.toml
</div>""", unsafe_allow_html=True)

        st.markdown("### Stats")
        c1, c2 = st.columns(2)
        c1.metric("Template Fields", len(eng.tpl_fields))
        c2.metric("Source Fields", len(eng.src_fields))
        c1.metric("Mappings", len(eng.mappings))
        mt = sum(1 for f in eng.tpl_fields.values() if f.mandatory_status == "Mandatory")
        mm = sum(1 for f in eng.tpl_fields.values() if f.mandatory_status == "Mandatory" and f.field_id in eng.mappings)
        c2.metric("Mandatory", f"{mm}/{mt}")
        
        st.markdown("---")
        
        # Clear Application Button
        if st.button("🗑️ Clear Application", use_container_width=True, type="secondary"):
            if st.session_state.get("_confirm_clear", False):
                clear_application()
                st.success("✅ Application cleared!")
                st.rerun()
            else:
                st.session_state["_confirm_clear"] = True
                st.warning("⚠️ Click again to confirm")
        
        if st.session_state.get("_confirm_clear", False):
            if st.button("Cancel Clear", use_container_width=True):
                st.session_state["_confirm_clear"] = False
                st.rerun()

# ══════════════════════════════════════════════════════════════════════════════
# TAB 1 — UPLOAD
# ══════════════════════════════════════════════════════════════════════════════
def tab_upload():
    eng = st.session_state.eng

    st.markdown("**File Upload & Configuration**")
    col1, col2 = st.columns(2, gap="large")

    with col1:
        st.markdown("**🗂️ Oracle Fusion Template**")
        
        # ✅ Check if we have a persisted template
        if "template_file_persisted" in st.session_state and st.session_state["template_file_persisted"]:
            st.success(f"✅ Template loaded: {st.session_state.get('template_file_name', 'Unknown')}")
            if st.button("🔄 Upload Different Template", use_container_width=True):
                # Clear persisted template
                st.session_state["template_file_persisted"] = False
                st.rerun()
        else:
            # Show file uploader
            tf = st.file_uploader("Template file", type=["xlsx","xls","xlsm"], key="tup", label_visibility="collapsed")
            
            if tf:
                # ✅ Persist the uploaded file immediately
                persist_uploaded_file(tf, "template_file")
                
                if st.button("🔍 Analyse Template", type="primary", use_container_width=True):
                    with st.spinner("Extracting fields, LOVs, AI mandatory detection…"):
                        try:
                            # Load template from persisted bytes
                            res = eng.load_template(tf)
                            st.session_state.tpl_ok = True
                            st.session_state["template_file_persisted"] = True
                            st.session_state["template_file_name"] = tf.name
                            
                            # ✅ Save state immediately
                            save_app_state()
                            st.rerun()
                        except Exception as e:
                            st.error(f"Failed: {e}")

        if st.session_state.tpl_ok:
            st.caption(f"{len(eng.all_sheets)} sheets · {len(eng.lovs)} LOV tables")

            if eng.all_sheets:
                chosen = st.selectbox("Select sheet to map", eng.all_sheets,
                                       index=eng.all_sheets.index(eng.active) if eng.active in eng.all_sheets else 0)

                import io
                raw_prev = pd.read_excel(io.BytesIO(st.session_state["_tpl_bytes"]),
                                          sheet_name=chosen, header=None, engine="openpyxl", nrows=8)
                st.dataframe(raw_prev.fillna(""), use_container_width=True, height=180)

                det = PatternDetector()
                _, auto_hdr, _, _ = det.detect(raw_prev)
                hdr_row = st.number_input(f"Header row (auto: {auto_hdr})", 0,
                                           min(10, len(raw_prev)-1), auto_hdr)

                if st.button(f"✅ Confirm & Analyse: {chosen}", type="primary", use_container_width=True):
                    with st.spinner(f"Analysing {chosen} with AI mandatory detection…"):
                        eng.active = chosen
                        p = eng.analyse_sheet(chosen, forced_hdr=int(hdr_row))
                        mn = sum(1 for f in p.fields.values() if f.mandatory_status == "Mandatory")
                        ai_detected = sum(1 for f in p.fields.values() if f.ai_mandatory_confidence > 0)
                        
                        if ai_detected > 0:
                            st.success(f"✅ {len(p.fields)} fields · {mn} mandatory ({ai_detected} AI-detected)")
                        else:
                            st.success(f"✅ {len(p.fields)} fields · {mn} mandatory")
                        
                        # ✅ Save state immediately
                        save_app_state()
                        st.rerun()

    with col2:
        st.markdown("**📋 Source Dataset**")
        
        # ✅ Check if we have a persisted source
        if "source_file_persisted" in st.session_state and st.session_state["source_file_persisted"]:
            st.success(f"✅ Source loaded: {st.session_state.get('source_file_name', 'Unknown')}")
            if st.button("🔄 Upload Different Source", use_container_width=True):
                # Clear persisted source
                st.session_state["source_file_persisted"] = False
                st.rerun()
        else:
            # Show file uploader
            sf = st.file_uploader("Source file", type=["xlsx","xls","csv"], key="sup", label_visibility="collapsed")
            
            if sf:
                # ✅ Persist the uploaded file immediately
                persist_uploaded_file(sf, "source_file")
                
                import io
                sb = sf.read(); sf.seek(0)
                is_csv = sf.name.lower().endswith(".csv")

                if is_csv:
                    pv = pd.read_csv(io.BytesIO(sb), nrows=5)
                    src_sheets = ["CSV"]; chosen_s = "CSV"; src_hdr = 0
                else:
                    xl = pd.ExcelFile(io.BytesIO(sb), engine="openpyxl")
                    src_sheets = xl.sheet_names
                    chosen_s = st.selectbox("Source sheet", src_sheets) if len(src_sheets)>1 else src_sheets[0]
                    src_hdr = st.number_input("Header row index", 0, 10, 0)
                    pv = pd.read_excel(io.BytesIO(sb), sheet_name=chosen_s, header=src_hdr, engine="openpyxl", nrows=5)

                all_cols = list(pv.columns)
                sel = st.multiselect("Select columns", all_cols, default=all_cols)

                if sel:
                    st.dataframe(pv[sel].fillna(""), use_container_width=True, height=160)

                if st.button("📊 Load Source Data", type="primary", use_container_width=True):
                    with st.spinner("Loading…"):
                        try:
                            sf.seek(0)
                            r = eng.load_source(sf, chosen_s, src_hdr, sel)
                            st.session_state.src_ok = True
                            st.session_state["source_file_persisted"] = True
                            st.session_state["source_file_name"] = sf.name
                            
                            st.success(f"✅ {r['fields']} fields · {r['records']:,} records loaded")
                            
                            # ✅ Save state immediately
                            save_app_state()
                            st.rerun()
                        except Exception as e:
                            st.error(f"Load failed: {e}")

# ══════════════════════════════════════════════════════════════════════════════
# TAB 2 — TEMPLATE STRUCTURE (UNCHANGED)
# ══════════════════════════════════════════════════════════════════════════════
def tab_template_structure():
    """Excel-style Template Structure tab with inline rule editing"""
    eng = st.session_state.eng
    if not eng.tpl_fields:
        st.info("ℹ️ Analyse a template sheet first (Upload tab)")
        return

    st.markdown(f"""
    <div class="mapping-header animate-fade">
        <h1 style="margin:0">📋 Template Structure — Field Master Grid</h1>
        <p style="margin:5px 0 0;font-size:0.9rem;opacity:0.8;">Define validation rules and structure for {eng.active}</p>
    </div>
    """, unsafe_allow_html=True)

    # Grid Header
    h_col1, h_col2, h_col3, h_col4, h_col5 = st.columns([3, 2, 1, 3, 3])
    with h_col1: st.markdown("**Field Name**")
    with h_col2: st.markdown("**Data Type**")
    with h_col3: st.markdown("**Mandatory**")
    with h_col4: st.markdown("**Rule Name**")
    with h_col5: st.markdown("**Actions**")
    st.markdown("---")

    all_f = list(eng.tpl_fields.values())
    
    for f in all_f:
        if not f.enterprise_rule:
            f.enterprise_rule = EnterpriseValidationRule(field_id=f.field_id, field_name=f.name, data_type=f.data_type)
        
        rule = f.enterprise_rule
        row_id = f.field_id
        
        # Row Container
        with st.container():
            col1, col2, col3, col4, col5 = st.columns([3, 2, 1, 3, 3])
            
            with col1:
                st.markdown(f"**{f.name}**")
                if f.mandatory_status == "Mandatory":
                    st.markdown('<span style="color:#ef4444;font-size:0.7rem;">REQUIRED *</span>', unsafe_allow_html=True)
            
            with col2:
                old_data_type = rule.data_type
                rule.data_type = st.selectbox(
                    "Data Type", 
                    options=["Text", "Number", "Date", "Email", "Currency", "Boolean"],
                    index=["Text", "Number", "Date", "Email", "Currency", "Boolean"].index(rule.data_type) if rule.data_type in ["Text", "Number", "Date", "Email", "Currency", "Boolean"] else 0,
                    key=f"type_{row_id}",
                    label_visibility="collapsed"
                )
                # ✅ AUTO-SAVE: Detect data type change
                if old_data_type != rule.data_type:
                    save_app_state()
            
            with col3:
                old_mandatory = f.is_mandatory
                f.is_mandatory = st.checkbox("", value=f.is_mandatory, key=f"mand_{row_id}", label_visibility="collapsed")
                f.mandatory_status = "Mandatory" if f.is_mandatory else "Optional"
                if f.enterprise_rule:
                    f.enterprise_rule.mandatory_status = f.mandatory_status
                # ✅ AUTO-SAVE: Detect mandatory status change
                if old_mandatory != f.is_mandatory:
                    save_app_state()
            
            with col4:
                rule.rule_name = st.text_input("Rule Name", value=rule.rule_name, key=f"rname_{row_id}", label_visibility="collapsed", placeholder="Enter rule name...")
            
            with col5:
                act_c1, act_c2, act_c3 = st.columns(3)
                is_expanded = st.session_state.get(f"expanded_{row_id}", False)
                if act_c1.button("📐 Set", key=f"set_{row_id}", use_container_width=True, type="secondary" if not is_expanded else "primary"):
                    st.session_state[f"expanded_{row_id}"] = not is_expanded
                    st.rerun()
                
                if act_c2.button("🧪 Test", key=f"test_{row_id}", use_container_width=True):
                    st.session_state[f"test_open_{row_id}"] = True
                
                if act_c3.button("💾 Save", key=f"save_{row_id}", use_container_width=True, type="primary"):
                    save_app_state() # Persist to disk
                    st.success(f"Saved: {f.name}")

            # 📐 Inline Rule Editor (Expandable)
            if st.session_state.get(f"expanded_{row_id}", False):
                with st.container():
                    st.markdown(f"""<div style="background:#f8fafc; padding:15px; border-radius:8px; border:1px solid #e2e8f0; margin:10px 0;">""", unsafe_allow_html=True)
                    st.markdown(f"**Configure Rules for {f.name}**")
                    
                    if rule.data_type == "Text":
                        c1, c2, c3 = st.columns(3)
                        rule.min_length = c1.number_input("Min Length", 0, 999, rule.min_length or 0, key=f"min_l_{row_id}")
                        rule.max_length = c2.number_input("Max Length", 0, 999, rule.max_length or 0, key=f"max_l_{row_id}")
                        rule.no_special_chars = c3.checkbox("No Special Characters", value=rule.no_special_chars, key=f"no_spec_{row_id}")
                        
                        rc1, rc2 = st.columns(2)
                        rule.regex_pattern = rc1.text_input("Regex Pattern", value=rule.regex_pattern, key=f"regex_{row_id}")
                        allowed_str = rc2.text_input("Allowed Values (comma separated)", value=",".join(rule.allowed_values) if rule.allowed_values else "", key=f"allow_{row_id}")
                        if allowed_str: rule.allowed_values = [x.strip() for x in allowed_str.split(",")]
                    
                    elif rule.data_type == "Number":
                        c1, c2, c3 = st.columns(3)
                        rule.min_value = c1.number_input("Min Value", value=float(rule.min_value) if rule.min_value is not None else 0.0, key=f"min_v_{row_id}")
                        rule.max_value = c2.number_input("Max Value", value=float(rule.max_value) if rule.max_value is not None else 0.0, key=f"max_v_{row_id}")
                        rule.decimal_allowed = c3.checkbox("Decimal Allowed", value=rule.decimal_allowed, key=f"dec_{row_id}")
                        rule.duplicate_check = st.checkbox("Check for Duplicates", value=rule.duplicate_check, key=f"dup_{row_id}")

                    elif rule.data_type == "Date":
                        c1, c2 = st.columns(2)
                        rule.accept_all_formats = c1.checkbox("Accept All Date Formats", value=rule.accept_all_formats, key=f"all_fmt_{row_id}")
                        if not rule.accept_all_formats:
                            rule.date_format = c2.selectbox("Specific Date Format", options=ALL_DATE_FORMATS, index=ALL_DATE_FORMATS.index(rule.date_format) if rule.date_format in ALL_DATE_FORMATS else 0, key=f"fmt_{row_id}")
                        
                        st.markdown('<div style="font-size:0.75rem; color:#64748b;">Supports: DD/MM/YYYY, ISO 8601, Enterprise Timestamps, T-Minus Arithmetic</div>', unsafe_allow_html=True)

                    st.markdown("</div>", unsafe_allow_html=True)

            # 🧪 Inline Test Popup
            if st.session_state.get(f"test_open_{row_id}", False):
                with st.expander(f"🧪 Test Validation: {f.name}", expanded=True):
                    test_val = st.text_input("Enter sample value to test", key=f"tval_{row_id}")
                    if st.button("Run Test", key=f"trun_{row_id}"):
                        is_valid, msg = eng._validator.test_sample_value(test_val, rule)
                        if is_valid:
                            st.success("✅ Validation Passed!")
                        else:
                            st.error(f"❌ {msg}")
                    if st.button("Close Test", key=f"tclose_{row_id}"):
                        st.session_state[f"test_open_{row_id}"] = False
                        st.rerun()

        st.markdown('<div style="height:1px; background:#f1f5f9; margin:8px 0;"></div>', unsafe_allow_html=True)

    # ✅ NEW: DATA QUALITY RULES SECTION
    st.markdown("---")
    st.markdown(f"""
    <div class="mapping-header animate-fade" style="margin-top:30px;">
        <h1 style="margin:0">🛡️ Data Quality Rules Configuration</h1>
        <p style="margin:5px 0 0;font-size:0.9rem;opacity:0.8;">Configure template-driven data quality and transformation rules</p>
    </div>
    """, unsafe_allow_html=True)
    
    # Quality Rules for each field
    for f in all_f:
        with st.expander(f"🛡️ Data Quality Rules: {f.name}", expanded=False):
            render_quality_rules_panel(f)
    
    # ✅ NEW: SOURCE DATA TRANSFORMATION PANEL
    st.markdown("---")
    st.markdown(f"""
    <div class="mapping-header animate-fade" style="margin-top:30px;">
        <h1 style="margin:0">🔄 Source Data Transformation</h1>
        <p style="margin:5px 0 0;font-size:0.9rem;opacity:0.8;">Apply global transformations to source data before mapping</p>
    </div>
    """, unsafe_allow_html=True)
    
    with st.expander("🔄 Global Source Data Transformations", expanded=False):
        st.markdown("**Apply these transformations to all source data:**")
        
        col1, col2, col3 = st.columns(3)
        
        global_remove_special = col1.checkbox("☐ Remove special characters", key="global_remove_special")
        global_replace_underscore = col2.checkbox("☐ Replace underscore with space", key="global_replace_underscore")
        global_trim_spaces = col3.checkbox("☐ Trim spaces", key="global_trim_spaces")
        
        col4, col5, col6 = st.columns(3)
        global_uppercase = col4.checkbox("☐ Convert to uppercase", key="global_uppercase")
        global_lowercase = col5.checkbox("☐ Convert to lowercase", key="global_lowercase")
        global_regex_transform = col6.checkbox("☐ Apply regex transformation", key="global_regex_transform")
        
        if global_regex_transform:
            st.markdown("**Custom Regex Transformation:**")
            regex_col1, regex_col2 = st.columns(2)
            global_regex_pattern = regex_col1.text_input("Regex Pattern", key="global_regex_pattern", placeholder="e.g., [^A-Za-z0-9 ]")
            global_regex_replacement = regex_col2.text_input("Replacement", key="global_regex_replacement", placeholder="e.g., (space)")
        
        if any([global_remove_special, global_replace_underscore, global_trim_spaces, 
                global_uppercase, global_lowercase, global_regex_transform]):
            
            if st.button("🔄 Apply Global Transformations", type="primary"):
                if eng.source_df is not None:
                    # Apply transformations to all string columns
                    for col in eng.source_df.columns:
                        if eng.source_df[col].dtype == 'object':  # String columns
                            if global_trim_spaces:
                                eng.source_df[col] = eng.source_df[col].astype(str).str.strip()
                            if global_uppercase:
                                eng.source_df[col] = eng.source_df[col].astype(str).str.upper()
                            if global_lowercase:
                                eng.source_df[col] = eng.source_df[col].astype(str).str.lower()
                            if global_remove_special:
                                eng.source_df[col] = eng.source_df[col].astype(str).str.replace(r'[^A-Za-z0-9\s]', '', regex=True)
                            if global_replace_underscore:
                                eng.source_df[col] = eng.source_df[col].astype(str).str.replace('_', ' ')
                            if global_regex_transform and global_regex_pattern:
                                try:
                                    eng.source_df[col] = eng.source_df[col].astype(str).str.replace(global_regex_pattern, global_regex_replacement, regex=True)
                                except Exception as e:
                                    st.error(f"Regex error: {str(e)}")
                    
                    save_app_state()
                    st.success("✅ Global transformations applied to source data!")
                    st.rerun()
                else:
                    st.warning("⚠️ No source data loaded. Upload and analyze source data first.")
    
    # ✅ NEW: COUNTRY STANDARDIZATION MANAGEMENT
    st.markdown("---")
    st.markdown(f"""
    <div class="mapping-header animate-fade" style="margin-top:30px;">
        <h1 style="margin:0">🌍 Country Standardization</h1>
        <p style="margin:5px 0 0;font-size:0.9rem;opacity:0.8;">Manage country name standardization mappings</p>
    </div>
    """, unsafe_allow_html=True)
    
    with st.expander("🌍 Country Standardization Mappings", expanded=False):
        # Show existing mappings
        st.markdown("**Current Mappings:**")
        
        existing_mappings = eng._quality_engine.country_mappings
        if existing_mappings:
            for mapping in existing_mappings:
                if mapping.is_active:
                    col1, col2, col3 = st.columns([4, 2, 1])
                    col1.write(f"**{mapping.target_value}**: {', '.join(mapping.source_values[:3])}{'...' if len(mapping.source_values) > 3 else ''}")
                    col2.write("✅ Active")
                    if col3.button("🗑️", key=f"del_country_{mapping.mapping_id}"):
                        eng._quality_engine.remove_country_mapping(mapping.mapping_id)
                        save_app_state()
                        st.success("Country mapping removed!")
                        st.rerun()
        
        st.markdown("---")
        st.markdown("**Add New Country Mapping:**")
        
        new_col1, new_col2 = st.columns(2)
        new_target = new_col1.text_input("Target Country Name", placeholder="e.g., UAE", key="new_country_target")
        new_sources = new_col2.text_input("Source Values (comma separated)", 
                                         placeholder="e.g., uae, u.a.e, united arab emirates", 
                                         key="new_country_sources")
        
        if st.button("➕ Add Country Mapping", type="secondary"):
            if new_target and new_sources:
                source_list = [s.strip() for s in new_sources.split(',') if s.strip()]
                if source_list:
                    mapping_id = f"custom_{new_target.lower().replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
                    new_mapping = CountryStandardization(mapping_id, source_list, new_target.upper())
                    eng._quality_engine.add_country_mapping(new_mapping)
                    save_app_state()
                    st.success(f"✅ Added country mapping: {new_target}")
                    st.rerun()
                else:
                    st.error("Please provide valid source values")
            else:
                st.error("Please fill in both target and source values")

def render_quality_rules_panel(field: FieldMetadata):
    """Render the data quality rules configuration panel for a field"""
    eng = st.session_state.eng
    field_id = field.field_id
    
    # Get existing quality rules
    existing_rules = eng._quality_engine.get_quality_rules(field_id)
    
    st.markdown(f"**Configure Quality Rules for: {field.name}**")
    
    # Rule configuration tabs
    completeness_tab, format_tab, standardization_tab, regex_tab, preview_tab = st.tabs([
        "📋 Completeness", "📝 Format", "🔄 Standardization", "🔧 Regex", "👁️ Preview"
    ])
    
    with completeness_tab:
        st.markdown("**Completeness Rules**")
        col1, col2 = st.columns(2)
        
        not_null = col1.checkbox("☐ Not Null", key=f"qr_not_null_{field_id}")
        not_blank = col2.checkbox("☐ Must Not Be Blank", key=f"qr_not_blank_{field_id}")
        
        # Uniqueness Rules
        st.markdown("**Uniqueness Rules**")
        col3, col4 = st.columns(2)
        unique_values = col3.checkbox("☐ Unique Values Only", key=f"qr_unique_{field_id}")
        no_duplicates = col4.checkbox("☐ No Duplicates", key=f"qr_no_dup_{field_id}")
        
        if any([not_null, not_blank, unique_values, no_duplicates]):
            rule_id = f"completeness_{field_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            rule = DataQualityRule(
                rule_id=rule_id,
                column_name=field.name,
                field_id=field_id,
                rule_type=QualityRuleType.COMPLETENESS,
                rule_description="Completeness and uniqueness validation",
                not_null=not_null,
                not_blank=not_blank,
                unique_values=unique_values,
                no_duplicates=no_duplicates
            )
            if st.button("Add Completeness Rule", key=f"add_comp_{field_id}"):
                eng._quality_engine.add_quality_rule(field_id, rule)
                st.success("✅ Completeness rule added!")
                save_app_state()
                st.rerun()
    
    with format_tab:
        st.markdown("**Format Rules**")
        col1, col2, col3 = st.columns(3)
        
        alphanumeric_only = col1.checkbox("☐ Alphanumeric Only", key=f"qr_alphanum_{field_id}")
        no_special_chars = col2.checkbox("☐ No Special Characters", key=f"qr_no_special_{field_id}")
        no_latin_chars = col3.checkbox("☐ No Latin Characters (é, ñ, ü)", key=f"qr_no_latin_{field_id}")
        
        # Character Length
        st.markdown("**Character Length**")
        cl1, cl2 = st.columns(2)
        min_length = cl1.number_input("Minimum Length", min_value=0, key=f"qr_min_len_{field_id}")
        max_length = cl2.number_input("Maximum Length", min_value=0, key=f"qr_max_len_{field_id}")
        
        # Date Format
        if field.data_type == "Date":
            st.markdown("**Date Format**")
            date_format = st.selectbox("Date Format Pattern", 
                ["YYYY/MM/DD", "DD/MM/YYYY", "MM/DD/YYYY", "YYYY-MM-DD", "DD-MM-YYYY"],
                key=f"qr_date_fmt_{field_id}")
        else:
            date_format = ""
        
        if any([alphanumeric_only, no_special_chars, no_latin_chars, min_length > 0, max_length > 0, date_format]):
            rule_id = f"format_{field_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            rule = DataQualityRule(
                rule_id=rule_id,
                column_name=field.name,
                field_id=field_id,
                rule_type=QualityRuleType.FORMAT,
                rule_description="Format validation",
                alphanumeric_only=alphanumeric_only,
                no_special_chars=no_special_chars,
                no_latin_chars=no_latin_chars,
                min_length=min_length if min_length > 0 else None,
                max_length=max_length if max_length > 0 else None,
                date_format_pattern=date_format
            )
            if st.button("Add Format Rule", key=f"add_format_{field_id}"):
                eng._quality_engine.add_quality_rule(field_id, rule)
                st.success("✅ Format rule added!")
                save_app_state()
                st.rerun()
    
    with standardization_tab:
        st.markdown("**Standardization Rules**")
        col1, col2, col3 = st.columns(3)
        
        to_uppercase = col1.checkbox("☐ Convert to Uppercase", key=f"qr_upper_{field_id}")
        to_lowercase = col2.checkbox("☐ Convert to Lowercase", key=f"qr_lower_{field_id}")
        trim_spaces = col3.checkbox("☐ Trim Spaces", key=f"qr_trim_{field_id}")
        
        col4, col5 = st.columns(2)
        remove_special_chars = col4.checkbox("☐ Remove Special Characters", key=f"qr_remove_special_{field_id}")
        replace_underscore = col5.checkbox("☐ Replace Underscore with Space", key=f"qr_replace_under_{field_id}")
        
        if any([to_uppercase, to_lowercase, trim_spaces, remove_special_chars, replace_underscore]):
            rule_id = f"standardization_{field_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            rule = DataQualityRule(
                rule_id=rule_id,
                column_name=field.name,
                field_id=field_id,
                rule_type=QualityRuleType.STANDARDIZATION,
                rule_description="Standardization transformation",
                to_uppercase=to_uppercase,
                to_lowercase=to_lowercase,
                trim_spaces=trim_spaces,
                remove_special_chars=remove_special_chars,
                replace_underscore_with_space=replace_underscore
            )
            if st.button("Add Standardization Rule", key=f"add_std_{field_id}"):
                eng._quality_engine.add_quality_rule(field_id, rule)
                st.success("✅ Standardization rule added!")
                save_app_state()
                st.rerun()
    
    with regex_tab:
        st.markdown("**Advanced Regex Rules**")
        
        regex_examples = {
            "Remove special characters": r"[^A-Za-z0-9 ]",
            "Replace underscore with space": r"_",
            "Extract numbers only": r"[^0-9]",
            "Remove extra spaces": r"\s+",
            "Alphanumeric with dash": r"[^A-Za-z0-9\-]"
        }
        
        example_choice = st.selectbox("Choose Example or Custom", 
            ["Custom"] + list(regex_examples.keys()), 
            key=f"qr_regex_example_{field_id}")
        
        if example_choice != "Custom":
            default_pattern = regex_examples[example_choice]
            default_replacement = " " if "space" in example_choice else ""
        else:
            default_pattern = ""
            default_replacement = ""
        
        regex_pattern = st.text_input("Regex Pattern", value=default_pattern, key=f"qr_regex_pattern_{field_id}")
        regex_replacement = st.text_input("Replacement", value=default_replacement, key=f"qr_regex_replace_{field_id}")
        regex_description = st.text_input("Description", key=f"qr_regex_desc_{field_id}")
        
        if regex_pattern:
            rule_id = f"regex_{field_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            rule = DataQualityRule(
                rule_id=rule_id,
                column_name=field.name,
                field_id=field_id,
                rule_type=QualityRuleType.REGEX,
                rule_description=regex_description or "Custom regex rule",
                regex_pattern=regex_pattern,
                regex_replacement=regex_replacement,
                regex_description=regex_description
            )
            if st.button("Add Regex Rule", key=f"add_regex_{field_id}"):
                eng._quality_engine.add_quality_rule(field_id, rule)
                st.success("✅ Regex rule added!")
                save_app_state()
                st.rerun()
    
    with preview_tab:
        st.markdown("**Real-Time Validation Preview**")
        
        if existing_rules:
            # Show existing rules
            st.markdown("**Active Rules:**")
            for rule in existing_rules:
                if rule.is_active:
                    col1, col2, col3 = st.columns([3, 1, 1])
                    col1.write(f"• {rule.rule_type.value}: {rule.rule_description}")
                    if col2.button("🗑️", key=f"del_rule_{rule.rule_id}"):
                        eng._quality_engine.remove_quality_rule(field_id, rule.rule_id)
                        st.success("Rule removed!")
                        save_app_state()
                        st.rerun()
                    col3.write("✅ Active")
        
        # Test with sample data
        if eng.source_df is not None and field_id in eng.mappings:
            mapping = eng.mappings[field_id]
            if mapping.source_field in eng.source_df.columns:
                sample_values = eng.source_df[mapping.source_field].dropna().head(5).tolist()
                
                if sample_values:
                    st.markdown("**Preview with Source Data:**")
                    previews = eng._quality_engine.generate_preview(sample_values, field_id)
                    
                    if previews:
                        preview_df = pd.DataFrame([
                            {
                                "Source Value": p.source_value,
                                "Transformed Value": p.transformed_value,
                                "Status": p.validation_status,
                                "Applied Rules": ", ".join(p.applied_rules[:2]) + ("..." if len(p.applied_rules) > 2 else "")
                            }
                            for p in previews
                        ])
                        st.dataframe(preview_df, use_container_width=True)
                        
                        # ✅ NEW: Quick Test with Custom Value
                        st.markdown("**Test Custom Value:**")
                        test_col1, test_col2 = st.columns([3, 1])
                        test_value = test_col1.text_input("Enter test value", key=f"test_val_{field_id}", placeholder="Enter a value to test...")
                        
                        if test_col2.button("🧪 Test", key=f"test_btn_{field_id}"):
                            if test_value:
                                # Apply transformations
                                transformed_val, applied_rules, errors = eng._quality_engine.apply_transformations(test_value, field_id)
                                # Validate
                                is_valid, validation_errors, warnings = eng._quality_engine.validate_value(transformed_val, field_id)
                                
                                # Show results
                                if is_valid and not errors:
                                    st.success(f"✅ **Valid**: `{test_value}` → `{transformed_val}`")
                                    if applied_rules:
                                        st.info(f"Applied rules: {', '.join(applied_rules)}")
                                else:
                                    st.error(f"❌ **Invalid**: `{test_value}` → `{transformed_val}`")
                                    all_errors = errors + validation_errors
                                    for error in all_errors[:3]:
                                        st.write(f"  • {error}")
                                    if warnings:
                                        for warning in warnings[:2]:
                                            st.warning(f"⚠️ {warning}")
                            else:
                                st.warning("Please enter a test value")
                    else:
                        st.info("No preview data available")
        else:
            st.info("💡 Map this field to source data to see preview")

# ══════════════════════════════════════════════════════════════════════════════
# TAB 3 — REAL-TIME FIELD MAPPING (RE-ENGINEERED)
# ══════════════════════════════════════════════════════════════════════════════
def tab_field_mapping():
    """Enterprise Real-Time Mapping Interface with Advanced Transformation Engine"""
    eng = st.session_state.eng
    
    if not eng.tpl_fields or not eng.src_fields:
        st.info("⬆️ Upload and analyse both template and source files first")
        return

    # 1. Header & Stats
    all_f = list(eng.tpl_fields.values())
    mapped_ids = [fid for fid, m in eng.mappings.items() if m.is_active]
    mand_ids = [f.field_id for f in all_f if f.mandatory_status == "Mandatory"]
    missing_mand = [fid for fid in mand_ids if fid not in mapped_ids]
    
    st.markdown(f"""
    <div class="mapping-header animate-fade">
        <div style="display:flex;justify-content:space-between;align-items:center;">
            <div>
                <h1 style="margin:0">🔄 Advanced Real-Time Mapping</h1>
                <div style="margin-top:4px; opacity:0.8; font-size:0.9rem">
                    {len(mapped_ids)}/{len(all_f)} fields mapped · {len(missing_mand)} mandatory missing
                </div>
            </div>
            <div class="live-indicator">LIVE PREVIEW ACTIVE</div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    # 2. Toolbar
    t1, t2, t3 = st.columns([2, 3, 1.5])
    search_q = t1.text_input("🔍 Search target...", key="map_search", label_visibility="collapsed", placeholder="Search fields...")
    status_filter = t2.segmented_control(
        "Filter", 
        ["All", "Mandatory", "Unmapped", "Mapped"], 
        default="All",
        key="map_filter_seg"
    )
    if t3.button("🤖 AI Auto-Map", use_container_width=True, type="secondary"):
        with st.spinner("🤖 AI analyzing and mapping fields..."):
            suggestions = eng.suggest()
            mapped_count = 0
            
            # Create all mappings
            for s in suggestions:
                try:
                    eng.save_map(s["target_id"], s["source_id"], confidence=s["confidence"], ai_suggested=True)
                    mapped_count += 1
                except Exception as e:
                    st.warning(f"Failed to map {s.get('target_id', 'unknown')}: {e}")
            
            # ✅ INSTANT AUTO-SAVE: Persist immediately after AI mapping
            save_app_state()
            
            # ✅ FORCE COMPLETE DROPDOWN REFRESH: Clear ALL dropdown-related keys
            # This is critical - we need to clear both old and new key formats
            keys_to_clear = []
            for key in list(st.session_state.keys()):
                if key.startswith("src_"):
                    keys_to_clear.append(key)
            
            for k in keys_to_clear:
                del st.session_state[k]
            
            # Show success message
            st.success(f"✅ AI mapped {mapped_count} fields successfully! Dropdowns will update now.")
            
            # ✅ INSTANT UPDATE: Rerun to refresh UI with new mappings
            st.rerun()

    # 3. Filtering
    display_fields = all_f
    if search_q:
        display_fields = [f for f in display_fields if search_q.lower() in f.name.lower()]
    if status_filter == "Mandatory":
        display_fields = [f for f in display_fields if f.field_id in mand_ids]
    elif status_filter == "Unmapped":
        display_fields = [f for f in display_fields if f.field_id not in mapped_ids]
    elif status_filter == "Mapped":
        display_fields = [f for f in display_fields if f.field_id in mapped_ids]

    st.markdown("---")
    
    # 4. Grid Headers
    gh1, gh2, gh3, gh4 = st.columns([2.5, 2.5, 3.5, 1.5])
    gh1.markdown("**Target Field**")
    gh2.markdown("**Source / Logic**")
    gh3.markdown("**Transformation Preview**")
    gh4.markdown("**Actions**")

    # 5. Grid Rows
    for f in display_fields:
        # Get current mapping for this field
        mapping = eng.mappings.get(f.field_id)
        is_mapped = mapping is not None and mapping.is_active
        is_mand = f.field_id in mand_ids
        row_id = f.field_id
        
        # DEBUG: Log mapping state for troubleshooting
        # Uncomment to debug: st.write(f"DEBUG {f.name}: is_mapped={is_mapped}, source={mapping.source_field if mapping else 'None'}")
        
        with st.container():
            col1, col2, col3, col4 = st.columns([2.5, 2.5, 3.5, 1.5])
            
            # Target Field Column
            with col1:
                st.markdown(f"**{f.name}**")
                # Show mapping status with color coding
                if is_mapped:
                    status_text = f"{f.data_type} • Mapped"
                    status_color = "#16a34a"  # Green for mapped
                elif is_mand:
                    status_text = f"{f.data_type} • Mandatory (Unmapped)"
                    status_color = "#ef4444"  # Red for unmapped mandatory
                else:
                    status_text = f"{f.data_type} • Optional"
                    status_color = "#64748b"  # Gray for optional
                
                st.markdown(f'<span style="color:{status_color}; font-size:0.75rem;">{status_text}</span>', unsafe_allow_html=True)

            # Source Selection Column
            with col2:
                src_names = ["-- Unmapped --"] + [sf.name for sf in eng.src_fields.values()]
                
                # Get current mapping and source field - CRITICAL: Must reflect latest state
                if is_mapped and mapping:
                    current_src = mapping.source_field
                    # Verify the source field exists in our list
                    if current_src not in src_names:
                        # Fallback: try to find by source_field_id
                        src_field_obj = eng.src_fields.get(mapping.source_field_id)
                        current_src = src_field_obj.name if src_field_obj else "-- Unmapped --"
                else:
                    current_src = "-- Unmapped --"
                
                # Calculate dropdown index - ensure it reflects current mapping
                try: 
                    s_idx = src_names.index(current_src)
                except ValueError: 
                    # If source not found, default to unmapped
                    s_idx = 0
                    current_src = "-- Unmapped --"
                
                # 🔍 DEBUG: Uncomment to troubleshoot dropdown issues
                # st.caption(f"DEBUG: is_mapped={is_mapped}, current_src={current_src}, s_idx={s_idx}")
                
                # ✅ CRITICAL FIX: Use unique key that changes when mapping changes
                # This forces Streamlit to recreate the widget with the correct index
                dropdown_key = f"src_{row_id}_{current_src.replace(' ', '_')}"
                
                new_src = st.selectbox(
                    f"Src_{row_id}", 
                    src_names, 
                    index=s_idx, 
                    label_visibility="collapsed", 
                    key=dropdown_key  # ✅ Dynamic key forces refresh
                )
                
                # Handle source field changes with INSTANT AUTO-SAVE
                if new_src != current_src:
                    if new_src == "-- Unmapped --": 
                        eng.remove_map(row_id)
                    else:
                        sf_id = next(sf.field_id for sf in eng.src_fields.values() if sf.name == new_src)
                        eng.save_map(row_id, sf_id)
                    
                    # ✅ INSTANT AUTO-SAVE: Persist immediately after manual change
                    save_app_state()
                    
                    # ✅ INSTANT UPDATE: Rerun to refresh UI
                    st.rerun()

            # Preview Column
            with col3:
                if is_mapped:
                    # ✅ Get actual source data value
                    if eng.source_df is not None and mapping.source_field in eng.source_df.columns:
                        # Get first non-null value from source column
                        source_col = eng.source_df[mapping.source_field].dropna()
                        if not source_col.empty:
                            actual_source_value = source_col.iloc[0]
                        else:
                            actual_source_value = ""
                    else:
                        actual_source_value = "No data"
                    
                    # Generate preview with actual data
                    preview = eng._mapping_validator.preview_transformation(
                        mapping, 
                        actual_source_value, 
                        eng.source_df.iloc[0].to_dict() if eng.source_df is not None else {}
                    )
                    val_status = preview.get("validation_status", "Valid")
                    
                    # ✅ NEW: Apply quality rules transformation to preview
                    transformed_value = preview.get('after_transformation', actual_source_value)
                    quality_transformed_value, applied_rules, quality_errors = eng._quality_engine.apply_transformations(
                        transformed_value, f.field_id
                    )
                    
                    # Update validation status based on quality rules
                    if quality_errors:
                        val_status = "Quality Issues"
                        badge_cls = "badge-failed"
                    elif applied_rules:
                        val_status = "Quality Applied"
                        badge_cls = "badge-passed"
                    else:
                        badge_cls = "badge-passed" if "Valid" in val_status else "badge-failed"
                    
                    # Show quality-transformed value
                    display_value = quality_transformed_value if applied_rules else transformed_value
                    
                    # Create tooltip with quality rules info
                    tooltip_info = f"Source: {actual_source_value}"
                    if applied_rules:
                        tooltip_info += f" | Rules: {', '.join(applied_rules[:2])}"
                    if quality_errors:
                        tooltip_info += f" | Errors: {', '.join(quality_errors[:2])}"
                    
                    st.markdown(f"""
                    <div style="display:flex; align-items:center; gap:8px;">
                        <div class="preview-box" style="flex:1" title="{tooltip_info}">{display_value}</div>
                        <span class="validation-badge {badge_cls}">{val_status}</span>
                    </div>
                    """, unsafe_allow_html=True)
                else:
                    st.markdown('<div class="preview-box" style="color:#cbd5e1;">No mapping</div>', unsafe_allow_html=True)

            # Actions Column
            with col4:
                b1, b2 = st.columns(2)
                is_open = st.session_state.get(f"map_panel_{row_id}", False)
                if b1.button("⚙️", key=f"xf_{row_id}", help="Configure transformations", use_container_width=True, type="primary" if is_open else "secondary"):
                    st.session_state[f"map_panel_{row_id}"] = not is_open
                    st.rerun()
                if b2.button("🗑️", key=f"clr_{row_id}", help="Clear mapping", use_container_width=True):
                    eng.remove_map(row_id)
                    st.rerun()

            # 🛠️ ADVANCED TRANSFORMATION PANEL
            if st.session_state.get(f"map_panel_{row_id}", False) and is_mapped:
                render_transformation_wizard(row_id, mapping)

        st.markdown('<div style="height:1px; background:#f1f5f9; margin:8px 0;"></div>', unsafe_allow_html=True)

def render_transformation_wizard(fid: str, mapping: FieldMapping):
    """Deep configuration for field transformations"""
    eng = st.session_state.eng
    if mapping.transform_config is None:
        mapping.transform_config = TransformationConfig()
    config = mapping.transform_config
    
    with st.container():
        st.markdown(f"""<div style="background:#f8fafc; padding:20px; border-radius:12px; border:2px solid #e2e8f0; margin:10px 0;">""", unsafe_allow_html=True)
        st.markdown(f"🚀 **Advanced Transformation Builder: {mapping.target_field_name}**")
        
        mode_tab, text_tab, date_tab, preview_tab = st.tabs(["🔗 Mode & Logic", "🔤 Text Cleanup", "📅 Date/TZ", "👁️ Live Preview"])
        
        with mode_tab:
            config.transformation_type = st.radio("Logic Mode", 
                ["direct", "concatenate", "expression", "condition"], 
                horizontal=True, 
                index=["direct", "concatenate", "expression", "condition"].index(config.transformation_type),
                key=f"mode_{fid}")
            
            if config.transformation_type == "concatenate":
                c1, c2 = st.columns([3, 1])
                src_list = {sf.field_id: sf.name for sf in eng.src_fields.values()}
                config.concatenate_fields = c1.multiselect("Select fields to combine", options=list(src_list.keys()), format_func=lambda x: src_list[x], default=config.concatenate_fields, key=f"concat_f_{fid}")
                config.concatenate_separator = c2.text_input("Separator", value=config.concatenate_separator, key=f"sep_{fid}")
            
            elif config.transformation_type == "expression":
                st.markdown("Build formula using `{field_id}` placeholders. Examples: `{id} + 100`, `round({price} * 1.05, 2)`")
                config.formula_expression = st.text_area("Formula Expression", value=config.formula_expression, key=f"form_{fid}", height=80)
                
            elif config.transformation_type == "condition":
                c1, c2, c3, c4, c5 = st.columns(5)
                config.if_operator = c1.selectbox("If Value...", ["equals", "contains"], key=f"ifop_{fid}")
                config.if_value = c2.text_input("Is...", value=config.if_value, key=f"ifv_{fid}")
                config.then_result = c3.text_input("Then Set...", value=config.then_result, key=f"then_{fid}")
                config.else_result = c4.text_input("Else Set...", value=config.else_result, key=f"else_{fid}")
                config.handle_null = c5.selectbox("Nulls", ["keep", "default", "error"], key=f"null_{fid}")

        with text_tab:
            c1, c2, c3 = st.columns(3)
            config.trim_type = c1.selectbox("Whitespace", ["none", "left", "right", "anywhere"], key=f"trim_{fid}")
            config.case_type = c2.selectbox("Letter Case", ["none", "upper", "lower", "title"], key=f"case_{fid}")
            config.remove_special_chars = c3.checkbox("No Special Chars", value=config.remove_special_chars, key=f"spec_{fid}")
            
            sc1, sc2, sc3 = st.columns(3)
            config.replace_text_source = sc1.text_input("Replace...", value=config.replace_text_source, key=f"reps_{fid}")
            config.replace_text_target = sc2.text_input("With...", value=config.replace_text_target, key=f"rept_{fid}")
            config.split_delimiter = sc3.text_input("Split by (delimiter)", value=config.split_delimiter, key=f"split_{fid}")

        with date_tab:
            dc1, dc2 = st.columns(2)
            config.input_date_format = dc1.selectbox("Input Date Logic", ["auto", "specific"], key=f"datein_{fid}")
            config.output_date_format = dc2.selectbox("Output Format", options=list(EnterpriseValidator.DATE_FORMAT_MAPPING.keys()), key=f"dateout_{fid}")
            
            tz1, tz2, tz3 = st.columns(3)
            config.timezone_source = tz1.text_input("Source TZ", value=config.timezone_source, key=f"tzs_{fid}")
            config.timezone_target = tz2.text_input("Target TZ", value=config.timezone_target, key=f"tzt_{fid}")
            config.add_days = tz3.number_input("Add Days", value=config.add_days, key=f"addd_{fid}")

        with preview_tab:
            if eng.source_df is not None:
                st.markdown("**Real-Time Data Flow (First 5 Rows)**")
                pre_df = []
                for i in range(min(5, len(eng.source_df))):
                    row = eng.source_df.iloc[i].to_dict()
                    src_val = row.get(mapping.source_field_id, "")
                    final_val = EnterpriseTransformation.apply_config(src_val, config, row)
                    pre_df.append({"Source": src_val, "Final Output": final_val})
                st.table(pd.DataFrame(pre_df))

        if st.button("✅ Apply & Save Transformation", type="primary", key=f"save_xf_{fid}"):
            # ✅ INSTANT AUTO-SAVE: Persist transformation configuration
            save_app_state()
            st.success("✅ Transformation saved and persisted!")
            st.rerun()

        st.markdown("</div>", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
# TAB 4 — EXPORT
# ══════════════════════════════════════════════════════════════════════════════
def tab_export():
    """🔥 ENHANCED: Export with exact template preservation"""
    eng = st.session_state.eng
    if not eng.mappings:
        st.info("ℹ️ Configure field mappings first")
        return

    st.markdown(f"""
    <div class="mapping-header animate-fade">
        <h1 style="margin:0">💾 Export Oracle Fusion Data</h1>
        <p style="margin:5px 0 0;font-size:0.9rem;opacity:0.8;">Generate output preserving exact template format</p>
    </div>
    """, unsafe_allow_html=True)

    # Pre-flight checks
    unmapped_mand = [f for f in eng.tpl_fields.values() if f.mandatory_status == "Mandatory" and f.field_id not in eng.mappings]
    invalid_mappings = [m for m in eng.mappings.values() if m.mapping_validation_status == "Invalid"]
    
    col1, col2, col3 = st.columns(3)
    col1.metric("Total Fields", len(eng.tpl_fields))
    col2.metric("Mapped Fields", len(eng.mappings))
    col3.metric("Source Records", len(eng.source_df) if eng.source_df is not None else 0)
    
    st.markdown("---")
    
    # Validation Status
    if unmapped_mand:
        st.error(f"❌ **{len(unmapped_mand)} mandatory fields unmapped**")
        with st.expander("📋 View Unmapped Mandatory Fields"):
            for f in unmapped_mand:
                st.write(f"• {f.name} ({f.data_type})")
        st.warning("⚠️ Cannot export until all mandatory fields are mapped")
        return
    
    if invalid_mappings:
        st.error(f"❌ **{len(invalid_mappings)} mappings have validation errors**")
        with st.expander("📋 Review Validation Errors"):
            for m in invalid_mappings:
                st.write(f"**{m.target_field_name}**")
                for err in m.mapping_validation_errors:
                    st.write(f"  • {err}")
        st.warning("⚠️ Fix validation errors before exporting")
        return

    st.success("✅ All pre-flight checks passed - Ready to export!")
    
    # ✅ NEW: DATA QUALITY VALIDATION SUMMARY
    st.markdown("---")
    st.markdown("### 🛡️ Data Quality Validation Summary")
    
    if eng.source_df is not None and eng.mappings:
        # Run data quality validation
        validation_results = eng._quality_engine.validate_dataset(eng.source_df, eng.mappings)
        
        if validation_results:
            # Summary metrics
            total_fields_with_rules = len(validation_results)
            total_errors = sum(len(result['errors']) for result in validation_results.values())
            total_warnings = sum(len(result['warnings']) for result in validation_results.values())
            
            col1, col2, col3, col4 = st.columns(4)
            col1.metric("Fields with Rules", total_fields_with_rules)
            col2.metric("Total Errors", total_errors)
            col3.metric("Total Warnings", total_warnings)
            col4.metric("Quality Score", f"{max(0, 100 - total_errors * 5)}%")
            
            if total_errors > 0:
                st.error(f"⚠️ Found {total_errors} data quality errors")
                with st.expander("📋 View Data Quality Issues", expanded=False):
                    for field_id, result in validation_results.items():
                        if result['errors'] or result['warnings']:
                            field_name = eng.tpl_fields.get(field_id, {}).name if hasattr(eng.tpl_fields.get(field_id, {}), 'name') else field_id
                            st.markdown(f"**{field_name}:**")
                            
                            for error in result['errors'][:5]:  # Show first 5 errors
                                st.write(f"  ❌ {error}")
                            
                            for warning in result['warnings'][:3]:  # Show first 3 warnings
                                st.write(f"  ⚠️ {warning}")
                            
                            if result['unique_violations']:
                                st.write(f"  🔄 Duplicate values: {len(result['unique_violations'])} found")
            elif total_warnings > 0:
                st.warning(f"⚠️ Found {total_warnings} data quality warnings")
                with st.expander("📋 View Data Quality Warnings", expanded=False):
                    for field_id, result in validation_results.items():
                        if result['warnings']:
                            field_name = eng.tpl_fields.get(field_id, {}).name if hasattr(eng.tpl_fields.get(field_id, {}), 'name') else field_id
                            st.markdown(f"**{field_name}:**")
                            for warning in result['warnings'][:5]:
                                st.write(f"  ⚠️ {warning}")
            else:
                st.success("✅ No data quality issues found!")
        else:
            st.info("ℹ️ No quality rules configured. Configure rules in Template Structure tab for validation.")
    
    st.markdown("---")
    st.markdown("### Export Options")
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.markdown("**📄 Template Preservation Mode** (Recommended)")
        st.info("""
        ✅ Preserves ALL sheets (only modifies selected sheet)  
        ✅ Preserves formatting, colors, styles  
        ✅ Preserves formulas and merged cells  
        ✅ Exact template structure maintained
        """)
        
        if st.button("🚀 Export with Template Format", type="primary", use_container_width=True):
            with st.spinner("Generating output with template preservation..."):
                try:
                    # Use new template preservation method
                    output_bytes = eng.generate_with_template_preservation()
                    
                    fname = f"Fusion_{eng.tpl_type.replace(' ','_')}_{eng.active}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                    
                    st.success(f"✅ Generated {len(eng.source_df):,} records in sheet '{eng.active}'")
                    st.info(f"📊 All other sheets preserved unchanged")
                    
                    # ✅ Use unique key to prevent download pausing
                    download_key = f"download_template_{datetime.now().strftime('%Y%m%d_%H%M%S_%f')}"
                    
                    st.download_button(
                        label=f"📥 Download {fname}",
                        data=output_bytes,
                        file_name=fname,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                        key=download_key
                    )
                except Exception as e:
                    st.error(f"Export failed: {e}")
                    import traceback
                    with st.expander("🔍 Error Details"):
                        st.code(traceback.format_exc())
    
    with col2:
        st.markdown("**📊 Data Only Mode** (Legacy)")
        st.warning("""
        ⚠️ Exports only mapped data as new file  
        ⚠️ No formatting preservation  
        ⚠️ Only active sheet data  
        ⚠️ CSV option available
        """)
        
        fmt = st.selectbox("Output format", ["Excel (.xlsx)", "CSV (.csv)"], key="legacy_fmt")
        
        if st.button("📤 Export Data Only", type="secondary", use_container_width=True):
            with st.spinner("Generating data-only output..."):
                try:
                    out = eng.generate(fix=True)
                    buf = BytesIO()
                    fname = f"Data_{eng.tpl_type.replace(' ','_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"

                    if "Excel" in fmt:
                        out.to_excel(buf, index=False, engine="openpyxl")
                        mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        ext = "xlsx"
                    else:
                        buf.write(out.to_csv(index=False).encode())
                        mime = "text/csv"
                        ext = "csv"

                    buf.seek(0)
                    st.success(f"✅ Generated {len(out):,} records")
                    
                    # ✅ Use unique key to prevent download pausing
                    download_key = f"download_data_{datetime.now().strftime('%Y%m%d_%H%M%S_%f')}"
                    
                    st.download_button(
                        label=f"📥 Download {fname}.{ext}",
                        data=buf,
                        file_name=f"{fname}.{ext}",
                        mime=mime,
                        use_container_width=True,
                        key=download_key
                    )
                except Exception as e:
                    st.error(f"Export failed: {e}")

# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════
def main():
    apply_css()
    init()
    sidebar()

    tabs = st.tabs([
        "📤 Upload & Analyse",
        "📋 Template Structure",
        "⚡ Real-Time Mapping",
        "💾 Export"
    ])
    
    with tabs[0]: tab_upload()
    with tabs[1]: tab_template_structure()
    with tabs[2]: tab_field_mapping()
    with tabs[3]: tab_export()

if __name__ == "__main__":
    main()