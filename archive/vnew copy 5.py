import streamlit as st
import pandas as pd
import numpy as np
import re
import json
import hashlib
import io
from datetime import datetime
from typing import Dict, List, Optional, Tuple, Any, Set
from dataclasses import dataclass, asdict, field
from enum import Enum
import logging
import warnings
warnings.filterwarnings('ignore')

# Setup logging
logging.basicConfig(
    level=logging.DEBUG,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('mapper_debug.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# Page configuration - MUST be first
st.set_page_config(
    page_title="Mapper Enterprise | Data Migration Platform",
    page_icon="🎯",
    layout="wide",
    initial_sidebar_state="expanded"
)

# =============================================================================
# ENTERPRISE CSS THEME
# =============================================================================

st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&display=swap');
    
    * {
        font-family: 'Inter', sans-serif;
    }
    
    /* Main Background */
    .stApp {
        background: linear-gradient(135deg, #0f172a 0%, #1e293b 100%);
    }
    
    /* Header */
    .enterprise-header {
        background: linear-gradient(135deg, #1e3a8a 0%, #3b82f6 50%, #06b6d4 100%);
        padding: 2rem;
        border-radius: 16px;
        margin-bottom: 2rem;
        box-shadow: 0 20px 60px rgba(0,0,0,0.4);
        position: relative;
        overflow: hidden;
    }
    
    .enterprise-header::before {
        content: '';
        position: absolute;
        top: -50%;
        left: -50%;
        width: 200%;
        height: 200%;
        background: radial-gradient(circle, rgba(255,255,255,0.1) 1px, transparent 1px);
        background-size: 20px 20px;
        opacity: 0.3;
    }
    
    .enterprise-title {
        font-size: 3rem;
        font-weight: 800;
        background: linear-gradient(90deg, #ffffff 0%, #bae6fd 100%);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        margin: 0;
        position: relative;
        z-index: 1;
    }
    
    .enterprise-subtitle {
        color: rgba(255,255,255,0.8);
        font-size: 1.1rem;
        margin-top: 0.5rem;
        position: relative;
        z-index: 1;
    }
    
    /* Tabs Styling */
    .stTabs [data-baseweb="tab-list"] {
        gap: 8px;
        background: rgba(30, 41, 59, 0.5);
        padding: 10px;
        border-radius: 12px;
        border: 1px solid rgba(255,255,255,0.1);
    }
    
    .stTabs [data-baseweb="tab"] {
        height: 50px;
        padding: 0 24px;
        background: transparent;
        border-radius: 8px;
        color: #94a3b8;
        font-weight: 600;
        font-size: 0.9rem;
        border: none;
        transition: all 0.3s ease;
    }
    
    .stTabs [aria-selected="true"] {
        background: linear-gradient(135deg, #3b82f6 0%, #06b6d4 100%) !important;
        color: white !important;
        box-shadow: 0 4px 15px rgba(59, 130, 246, 0.4);
    }
    
    /* Cards */
    .metric-card {
        background: rgba(30, 41, 59, 0.6);
        backdrop-filter: blur(10px);
        border: 1px solid rgba(255,255,255,0.1);
        border-radius: 12px;
        padding: 1.5rem;
        margin: 0.5rem 0;
        transition: all 0.3s ease;
    }
    
    .metric-card:hover {
        transform: translateY(-2px);
        border-color: rgba(59, 130, 246, 0.3);
        box-shadow: 0 10px 30px rgba(0,0,0,0.2);
    }
    
    /* Section Headers */
    .section-container {
        background: rgba(30, 41, 59, 0.4);
        border-radius: 12px;
        padding: 1.5rem;
        margin: 1rem 0;
        border: 1px solid rgba(255,255,255,0.05);
    }
    
    .section-title {
        color: #f1f5f9;
        font-size: 1.2rem;
        font-weight: 700;
        margin-bottom: 1rem;
        display: flex;
        align-items: center;
        gap: 0.5rem;
    }
    
    /* Upload Zone */
    .upload-zone {
        background: rgba(59, 130, 246, 0.05);
        border: 2px dashed rgba(59, 130, 246, 0.3);
        border-radius: 12px;
        padding: 2rem;
        text-align: center;
        transition: all 0.3s ease;
    }
    
    .upload-zone:hover {
        background: rgba(59, 130, 246, 0.1);
        border-color: rgba(59, 130, 246, 0.6);
    }
    
    /* Mapping Interface */
    .mapping-container {
        background: rgba(15, 23, 42, 0.6);
        border-radius: 12px;
        padding: 1.5rem;
        border: 1px solid rgba(255,255,255,0.1);
    }
    
    .mapping-row {
        background: rgba(30, 41, 59, 0.6);
        border: 1px solid rgba(255,255,255,0.1);
        border-radius: 8px;
        padding: 1rem;
        margin: 0.5rem 0;
        display: flex;
        align-items: center;
        gap: 1rem;
        transition: all 0.2s ease;
    }
    
    .mapping-row:hover {
        border-color: rgba(59, 130, 246, 0.4);
        background: rgba(59, 130, 246, 0.1);
    }
    
    .mapping-source {
        flex: 1;
        font-weight: 600;
        color: #60a5fa;
    }
    
    .mapping-arrow {
        color: #94a3b8;
        font-size: 1.5rem;
    }
    
    .mapping-target {
        flex: 1;
    }
    
    .mapping-actions {
        display: flex;
        gap: 0.5rem;
    }
    
    /* Validation Badges */
    .badge {
        display: inline-flex;
        align-items: center;
        gap: 0.25rem;
        padding: 0.25rem 0.75rem;
        border-radius: 20px;
        font-size: 0.75rem;
        font-weight: 600;
    }
    
    .badge-mandatory {
        background: linear-gradient(90deg, #ef4444 0%, #f87171 100%);
        color: white;
    }
    
    .badge-optional {
        background: linear-gradient(90deg, #10b981 0%, #34d399 100%);
        color: white;
    }
    
    .badge-rule {
        background: rgba(59, 130, 246, 0.2);
        color: #60a5fa;
        border: 1px solid rgba(59, 130, 246, 0.3);
    }
    
    /* Status Indicators */
    .status-success { color: #34d399; font-weight: 600; }
    .status-error { color: #f87171; font-weight: 600; }
    .status-warning { color: #fbbf24; font-weight: 600; }
    .status-info { color: #60a5fa; font-weight: 600; }
    
    /* Buttons */
    .stButton>button {
        background: linear-gradient(135deg, #3b82f6 0%, #06b6d4 100%);
        color: white;
        border: none;
        border-radius: 8px;
        padding: 0.75rem 1.5rem;
        font-weight: 600;
        transition: all 0.3s ease;
        text-transform: uppercase;
        letter-spacing: 0.5px;
        font-size: 0.85rem;
    }
    
    .stButton>button:hover {
        transform: translateY(-2px);
        box-shadow: 0 10px 25px rgba(59, 130, 246, 0.4);
    }
    
    .stButton>button[kind="secondary"] {
        background: rgba(255,255,255,0.1);
        border: 1px solid rgba(255,255,255,0.2);
    }
    
    /* Data Editor Customization */
    .stDataFrame {
        background: rgba(30, 41, 59, 0.4);
        border-radius: 8px;
        border: 1px solid rgba(255,255,255,0.1);
    }
    
    /* Progress Steps */
    .step-container {
        display: flex;
        justify-content: space-between;
        margin: 2rem 0;
        position: relative;
    }
    
    .step {
        display: flex;
        flex-direction: column;
        align-items: center;
        gap: 0.5rem;
        z-index: 1;
    }
    
    .step-circle {
        width: 40px;
        height: 40px;
        border-radius: 50%;
        display: flex;
        align-items: center;
        justify-content: center;
        font-weight: 700;
        font-size: 1rem;
        transition: all 0.3s ease;
    }
    
    .step-active {
        background: linear-gradient(135deg, #3b82f6 0%, #06b6d4 100%);
        color: white;
        box-shadow: 0 0 20px rgba(59, 130, 246, 0.5);
    }
    
    .step-completed {
        background: #10b981;
        color: white;
    }
    
    .step-pending {
        background: rgba(255,255,255,0.1);
        color: #64748b;
        border: 2px solid rgba(255,255,255,0.2);
    }
    
    .step-label {
        font-size: 0.8rem;
        color: #94a3b8;
        font-weight: 500;
    }
    
    .step-connector {
        position: absolute;
        top: 20px;
        left: 10%;
        right: 10%;
        height: 2px;
        background: rgba(255,255,255,0.1);
        z-index: 0;
    }
    
    /* Alerts */
    .alert {
        padding: 1rem;
        border-radius: 8px;
        margin: 1rem 0;
        border-left: 4px solid;
    }
    
    .alert-error {
        background: rgba(239, 68, 68, 0.1);
        border-color: #ef4444;
        color: #f87171;
    }
    
    .alert-success {
        background: rgba(16, 185, 129, 0.1);
        border-color: #10b981;
        color: #34d399;
    }
    
    .alert-warning {
        background: rgba(251, 191, 36, 0.1);
        border-color: #fbbf24;
        color: #fbbf24;
    }
    
    .alert-info {
        background: rgba(59, 130, 246, 0.1);
        border-color: #3b82f6;
        color: #60a5fa;
    }
    
    /* Sidebar */
    .css-1d391kg {
        background: rgba(15, 23, 42, 0.95);
    }
    
    /* Custom Scrollbar */
    ::-webkit-scrollbar {
        width: 8px;
        height: 8px;
    }
    
    ::-webkit-scrollbar-track {
        background: rgba(0,0,0,0.2);
    }
    
    ::-webkit-scrollbar-thumb {
        background: #3b82f6;
        border-radius: 4px;
    }
    
    /* Expander */
    .streamlit-expanderHeader {
        background: rgba(30, 41, 59, 0.6);
        border-radius: 8px;
        font-weight: 600;
        color: #f1f5f9;
    }
    
    /* Selectbox styling */
    .stSelectbox label {
        color: #94a3b8;
        font-weight: 500;
    }
    
    /* Checkbox styling */
    .stCheckbox label {
        color: #cbd5e1;
    }
    
    /* Number input */
    .stNumberInput label {
        color: #94a3b8;
    }
    
    /* Header row preview styling */
    .header-preview {
        background: rgba(59, 130, 246, 0.1);
        border: 1px solid rgba(59, 130, 246, 0.3);
        border-radius: 6px;
        padding: 0.5rem;
        font-family: monospace;
        font-size: 0.85rem;
        color: #60a5fa;
        margin-top: 0.5rem;
    }
</style>
""", unsafe_allow_html=True)

# =============================================================================
# DATA CLASSES
# =============================================================================

class ValidationType(Enum):
    MANDATORY = "mandatory"
    NOT_NULL = "not_null"
    NO_SPECIAL_CHARS = "no_special_chars"
    MAX_LENGTH = "max_length"
    MIN_LENGTH = "min_length"
    ONLY_CHARACTERS = "only_characters"
    ONLY_NUMBERS = "only_numbers"
    EMAIL_FORMAT = "email_format"
    DATE_FORMAT = "date_format"
    REGEX_PATTERN = "regex_pattern"

@dataclass
class ColumnRule:
    """Validation rule configuration for a column"""
    column_name: str
    is_mandatory: bool = False
    not_null: bool = False
    no_special_chars: bool = False
    max_length: Optional[int] = None
    min_length: Optional[int] = None
    only_characters: bool = False
    only_numbers: bool = False
    email_format: bool = False
    date_format: Optional[str] = None
    regex_pattern: Optional[str] = None
    transform_regex: Optional[str] = None
    default_value: Optional[str] = None
    description: str = ""
    
    def get_active_rules(self) -> List[str]:
        """Get list of active validation rules"""
        rules = []
        if self.is_mandatory:
            rules.append("Mandatory")
        if self.not_null:
            rules.append("Not Null")
        if self.no_special_chars:
            rules.append("No Special Chars")
        if self.max_length:
            rules.append(f"Max Len: {self.max_length}")
        if self.min_length:
            rules.append(f"Min Len: {self.min_length}")
        if self.only_characters:
            rules.append("Chars Only")
        if self.only_numbers:
            rules.append("Numbers Only")
        if self.email_format:
            rules.append("Email")
        if self.date_format:
            rules.append(f"Date: {self.date_format}")
        if self.regex_pattern:
            rules.append("Regex")
        return rules

@dataclass
class ColumnMapping:
    """Mapping between source and target columns"""
    id: str = field(default_factory=lambda: hashlib.md5(str(datetime.now().timestamp()).encode()).hexdigest()[:8])
    source_column: str = ""
    target_column: str = ""
    source_sheet: str = ""
    target_sheet: str = ""
    transform_rules: List[str] = field(default_factory=list)
    validation_errors: List[Dict] = field(default_factory=list)
    is_active: bool = True
    confidence_score: float = 0.0

@dataclass
class FileData:
    """Container for uploaded file data"""
    name: str = ""
    data: Dict[str, pd.DataFrame] = field(default_factory=dict)
    sheets: List[str] = field(default_factory=list)
    columns: Dict[str, List[str]] = field(default_factory=dict)
    selected_sheet: Optional[str] = None
    header_row: int = 0
    raw_data: Optional[bytes] = None

# =============================================================================
# SESSION STATE MANAGEMENT
# =============================================================================

def init_session_state():
    """Initialize session state with all required keys"""
    defaults = {
        # Workflow State
        'active_tab': 0,
        'completed_steps': set(),
        
        # File Data
        'template_file': FileData(),
        'source_file': FileData(),
        
        # Configuration
        'column_rules': {},  # column_name -> ColumnRule
        'mappings': [],  # List[ColumnMapping]
        
        # Processing Results
        'validation_results': {},
        'transformed_data': None,
        'processing_log': [],
        
        # UI State
        'show_mapping_panel': False,
        'auto_map_triggered': False,
        'selected_mapping_id': None,
        
        # Session
        'session_id': hashlib.md5(str(datetime.now()).encode()).hexdigest()[:8]
    }
    
    for key, value in defaults.items():
        if key not in st.session_state:
            st.session_state[key] = value

def save_session_state():
    """Export session state to JSON"""
    state_dict = {
        'column_rules': {k: asdict(v) for k, v in st.session_state.column_rules.items()},
        'mappings': [asdict(m) for m in st.session_state.mappings],
        'template_file': {
            'name': st.session_state.template_file.name,
            'selected_sheet': st.session_state.template_file.selected_sheet,
            'header_row': st.session_state.template_file.header_row,
        },
        'source_file': {
            'name': st.session_state.source_file.name,
            'selected_sheet': st.session_state.source_file.selected_sheet,
            'header_row': st.session_state.source_file.header_row,
        },
        'session_id': st.session_state.session_id
    }
    return json.dumps(state_dict, indent=2, default=str)

def load_session_state(json_data: str):
    """Import session state from JSON"""
    try:
        data = json.loads(json_data)
        
        # Restore column rules
        st.session_state.column_rules = {
            k: ColumnRule(**v) for k, v in data.get('column_rules', {}).items()
        }
        
        # Restore mappings
        st.session_state.mappings = [ColumnMapping(**m) for m in data.get('mappings', [])]
        
        # Restore file metadata (files themselves need to be re-uploaded)
        if 'template_file' in data:
            st.session_state.template_file.name = data['template_file'].get('name', '')
            st.session_state.template_file.selected_sheet = data['template_file'].get('selected_sheet')
            st.session_state.template_file.header_row = data['template_file'].get('header_row', 0)
        
        if 'source_file' in data:
            st.session_state.source_file.name = data['source_file'].get('name', '')
            st.session_state.source_file.selected_sheet = data['source_file'].get('selected_sheet')
            st.session_state.source_file.header_row = data['source_file'].get('header_row', 0)
        
        return True
    except Exception as e:
        st.error(f"Error loading session: {e}")
        return False

# =============================================================================
# FILE HANDLING - FIXED VERSION
# =============================================================================

def detect_file_type(file_name: str) -> str:
    """Detect file type from extension"""
    if file_name.lower().endswith('.csv'):
        return 'csv'
    elif file_name.lower().endswith(('.xlsx', '.xls', '.xlsm')):
        return 'excel'
    return 'unknown'

def clean_column_names(columns, header_row=None):
    """
    Clean column names by handling unnamed columns and stripping whitespace.
    Preserves actual header values from the selected row.
    Only replaces truly empty/NaN values with Column_N placeholders.
    """
    cleaned = []
    for i, col in enumerate(columns):
        col_str = str(col).strip()
        
        # Check if this is a truly empty/invalid value
        is_empty = (
            col_str.lower() in ['nan', 'none', ''] or 
            pd.isna(col) or
            col_str == 'nan' or
            col_str.startswith('Unnamed:')  # Pandas generated placeholder
        )
        
        if is_empty:
            # Replace with generic name only for truly empty values
            cleaned.append(f"Column_{i+1}")
        else:
            # Keep the original value (even if it starts with Column_)
            cleaned.append(col_str)
    return cleaned

def read_uploaded_file(uploaded_file, header_row: int = 0, selected_sheet: str = None) -> Tuple[Dict[str, pd.DataFrame], List[str], Dict[str, List[str]]]:
    """
    Read uploaded file and return data dictionary, sheet names, and column mapping.
    Uses the specified header_row to read column names dynamically.
    Ensures column headers are always captured even if no data rows exist.
    """
    if uploaded_file is None:
        return {}, [], {}
    
    try:
        # Handle both uploaded files and BytesIO objects
        if hasattr(uploaded_file, 'name'):
            file_type = detect_file_type(uploaded_file.name)
            if hasattr(uploaded_file, 'getvalue'):
                file_bytes = uploaded_file.getvalue()
            else:
                file_bytes = uploaded_file.read()
                uploaded_file.seek(0)
        else:
            # For BytesIO objects
            if hasattr(uploaded_file, 'getvalue'):
                file_bytes = uploaded_file.getvalue()
            else:
                file_bytes = uploaded_file.read()
                uploaded_file.seek(0)
            file_type = 'excel'
            # Try to detect from content
            try:
                pd.ExcelFile(io.BytesIO(file_bytes))
                file_type = 'excel'
            except:
                file_type = 'csv'
        
        data = {}
        
        if file_type == 'csv':
            # First, read without headers to get the raw data preview
            raw_df = pd.read_csv(io.BytesIO(file_bytes), header=None, nrows=50)
            
            # Extract column names from the selected header row if it exists
            if header_row < len(raw_df):
                column_names = clean_column_names(raw_df.iloc[header_row].tolist(), header_row)
            else:
                # Fallback to generic names if header row is beyond data
                column_names = [f"Column_{i+1}" for i in range(len(raw_df.columns))]
            
            # Read again with the specified header row
            df = pd.read_csv(io.BytesIO(file_bytes), header=header_row)
            
            # Ensure dataframe has the correct column names
            if len(df) == 0:
                # No data rows - create empty df with extracted column names
                if len(column_names) > 0:
                    df = pd.DataFrame(columns=column_names)
                # else: keep pandas default columns from header row
            else:
                # Has data rows - clean the column names from header row
                df.columns = clean_column_names(df.columns, header_row)
            
            # Drop rows that are completely empty
            df = df.dropna(how='all')
            
            data['Sheet1'] = df
            sheets = ['Sheet1']
            
        elif file_type == 'excel':
            # Read all sheets from Excel
            excel_file = pd.ExcelFile(io.BytesIO(file_bytes))
            all_sheets = excel_file.sheet_names
            
            # If specific sheet requested, only process that one
            sheets_to_process = [selected_sheet] if selected_sheet and selected_sheet in all_sheets else all_sheets
            
            for sheet in sheets_to_process:
                try:
                    logger.info(f"Processing sheet: {sheet}, header_row: {header_row}")
                    
                    # Read the full sheet WITHOUT headers first
                    raw_df = pd.read_excel(io.BytesIO(file_bytes), sheet_name=sheet, header=None)
                    
                    logger.info(f"Raw df shape: {raw_df.shape}")
                    
                    # Get total rows and columns
                    total_rows = len(raw_df)
                    total_cols = len(raw_df.columns)
                    
                    logger.info(f"Total rows: {total_rows}, Total cols: {total_cols}")
                    
                    if total_rows == 0:
                        logger.warning("No rows in sheet, skipping")
                        continue
                    
                    # Extract column names from the selected header row
                    if header_row < total_rows:
                        header_values = raw_df.iloc[header_row].tolist()
                        logger.info(f"Header values count: {len(header_values)}")
                        logger.info(f"First 10 header values: {header_values[:10]}")
                        
                        # Convert to strings and use as column names directly
                        column_names = []
                        for i, val in enumerate(header_values):
                            if pd.isna(val) or str(val).strip() == '' or str(val).lower() == 'nan':
                                column_names.append(f"Column_{i+1}")
                            else:
                                column_names.append(str(val).strip())
                    else:
                        # Fallback if header row is beyond data
                        logger.warning(f"Header row {header_row} beyond total rows {total_rows}")
                        column_names = [f"Column_{i+1}" for i in range(total_cols)]
                    
                    logger.info(f"Column names count: {len(column_names)}")
                    logger.info(f"Column names: {column_names[:10]}...")
                    
                    # Extract data rows (everything after header row)
                    if header_row + 1 < total_rows:
                        data_df = raw_df.iloc[header_row + 1:].copy()
                        data_df.columns = column_names
                        data_df = data_df.reset_index(drop=True)
                        logger.info(f"Data rows extracted: {len(data_df)}")
                    else:
                        # No data rows - create empty dataframe with headers
                        logger.info("No data rows, creating empty df with headers")
                        data_df = pd.DataFrame(columns=column_names)
                    
                    logger.info(f"Data df shape before any cleaning: {data_df.shape}")
                    
                    # Only drop rows that are completely empty, NOT columns
                    # This preserves all columns the user selected
                    data_df = data_df.dropna(how='all')
                    logger.info(f"Final data df shape: {data_df.shape}")
                    logger.info(f"Final columns count: {len(data_df.columns)}")
                    
                    data[sheet] = data_df
                    logger.info(f"Stored sheet with {len(data_df.columns)} columns")
                except Exception as e:
                    logger.error(f"Error reading sheet '{sheet}': {e}", exc_info=True)
                    st.warning(f"⚠️ Error reading sheet '{sheet}': {e}")
                    continue
                    
            sheets = list(data.keys())
        else:
            st.error("❌ Unsupported file format. Please upload CSV or Excel files.")
            return {}, [], {}
        
        # Extract columns for each sheet
        columns = {sheet: list(df.columns) for sheet, df in data.items()}
        
        return data, sheets, columns
        
    except Exception as e:
        st.error(f"❌ Error reading file: {str(e)}")
        import traceback
        st.error(traceback.format_exc())
        return {}, [], []

def get_header_preview(file_bytes: bytes, file_type: str, sheet_name: str = None, max_rows: int = 10) -> Optional[pd.DataFrame]:
    """
    Get raw data preview for header row selection (without header processing).
    Returns dataframe with NO header to show raw values at each row.
    """
    try:
        if file_type == 'csv':
            df = pd.read_csv(io.BytesIO(file_bytes), header=None, nrows=max_rows)
            return df
        else:
            df = pd.read_excel(io.BytesIO(file_bytes), sheet_name=sheet_name, header=None, nrows=max_rows)
            return df
    except Exception as e:
        return None

def auto_detect_header(df: pd.DataFrame, max_rows: int = 10) -> int:
    """
    Auto-detect header row by analyzing first few rows
    Returns the row index most likely to be the header
    """
    if len(df) == 0:
        return 0
    
    best_row = 0
    best_score = 0
    
    for i in range(min(max_rows, len(df))):
        row = df.iloc[i]
        
        # Skip completely empty rows
        if row.isna().all():
            continue
            
        # Score based on multiple factors
        string_count = sum(1 for x in row if isinstance(x, str) and str(x).strip())
        string_ratio = string_count / len(row) if len(row) > 0 else 0
        
        # Check uniqueness (headers usually have unique values)
        non_null_values = [x for x in row if pd.notna(x) and str(x).strip()]
        unique_ratio = len(set(non_null_values)) / len(non_null_values) if non_null_values else 0
        
        # Prefer rows with meaningful text (not just numbers or single characters)
        meaningful_text_count = sum(1 for x in row if isinstance(x, str) and len(str(x).strip()) > 2)
        meaningful_ratio = meaningful_text_count / len(row) if len(row) > 0 else 0
        
        # Check for common header patterns
        header_keywords = ['id', 'name', 'action', 'type', 'code', 'number', 'date', 'status', 'description']
        keyword_matches = sum(1 for x in row if isinstance(x, str) and 
                             any(keyword in str(x).lower() for keyword in header_keywords))
        keyword_ratio = keyword_matches / len(row) if len(row) > 0 else 0
        
        # Combined score: prioritize string content, uniqueness, meaningful text, and header keywords
        score = (string_ratio * 0.3) + (unique_ratio * 0.3) + (meaningful_ratio * 0.2) + (keyword_ratio * 0.2)
        
        if score > best_score:
            best_score = score
            best_row = i
    
    return best_row

def _refresh_template_data():
    """Refresh template data from stored bytes with current header row setting"""
    if st.session_state.template_file.raw_data:
        # Ensure we have bytes and create BytesIO object
        if isinstance(st.session_state.template_file.raw_data, bytes):
            file_obj = io.BytesIO(st.session_state.template_file.raw_data)
        else:
            file_obj = st.session_state.template_file.raw_data
            
        # Pass selected sheet to only read that sheet
        data, sheets, columns = read_uploaded_file(
            file_obj,
            st.session_state.template_file.header_row,
            st.session_state.template_file.selected_sheet
        )
        st.session_state.template_file.data = data
        st.session_state.template_file.sheets = sheets
        st.session_state.template_file.columns = columns
        # Ensure selected_sheet is set to the sheet we actually read
        if sheets and st.session_state.template_file.selected_sheet not in sheets:
            st.session_state.template_file.selected_sheet = sheets[0] if sheets else None

def _refresh_source_data():
    """Refresh source data from stored bytes with current header row setting"""
    if st.session_state.source_file.raw_data:
        # Ensure we have bytes and create BytesIO object
        if isinstance(st.session_state.source_file.raw_data, bytes):
            file_obj = io.BytesIO(st.session_state.source_file.raw_data)
        else:
            file_obj = st.session_state.source_file.raw_data
            
        # Pass selected sheet to only read that sheet
        data, sheets, columns = read_uploaded_file(
            file_obj,
            st.session_state.source_file.header_row,
            st.session_state.source_file.selected_sheet
        )
        st.session_state.source_file.data = data
        st.session_state.source_file.sheets = sheets
        st.session_state.source_file.columns = columns
        # Ensure selected_sheet is set to the sheet we actually read
        if sheets and st.session_state.source_file.selected_sheet not in sheets:
            st.session_state.source_file.selected_sheet = sheets[0] if sheets else None

# =============================================================================
# VALIDATION ENGINE
# =============================================================================

class ValidationEngine:
    """Enterprise-grade validation engine"""
    
    @staticmethod
    def validate_value(value: Any, rule: ColumnRule) -> Tuple[bool, List[str], Any]:
        """
        Validate and transform a single value
        Returns: (is_valid, error_messages, transformed_value)
        """
        errors = []
        
        # Handle null/empty values
        is_empty = pd.isna(value) or str(value).strip() == '' or value is None
        
        if is_empty:
            if rule.not_null:
                errors.append("Value cannot be null/empty")
            if rule.is_mandatory:
                errors.append("Mandatory field is empty")
            # Apply default value if provided
            if rule.default_value and is_empty:
                value = rule.default_value
            return len(errors) == 0, errors, value
        
        str_value = str(value).strip()
        transformed_value = str_value
        
        # Max Length
        if rule.max_length and len(str_value) > rule.max_length:
            errors.append(f"Length {len(str_value)} exceeds maximum {rule.max_length}")
            transformed_value = str_value[:rule.max_length]
        
        # Min Length
        if rule.min_length and len(str_value) < rule.min_length:
            errors.append(f"Length {len(str_value)} below minimum {rule.min_length}")
        
        # Only Characters (no digits)
        if rule.only_characters:
            if any(char.isdigit() for char in str_value):
                errors.append("Contains numeric characters")
            # Remove digits
            transformed_value = ''.join(char for char in str_value if not char.isdigit())
        
        # Only Numbers
        if rule.only_numbers:
            cleaned = str_value.replace(',', '').replace(' ', '').replace('$', '')
            try:
                float(cleaned)
                transformed_value = float(cleaned)
            except ValueError:
                errors.append("Not a valid numeric value")
        
        # No Special Characters (allows Latin/Unicode letters, numbers, spaces)
        if rule.no_special_chars:
            # Allow alphanumeric, spaces, and common punctuation
            allowed_pattern = r'[^\w\s\-\.]'
            special_chars = re.findall(allowed_pattern, str_value)
            if special_chars:
                errors.append(f"Contains special characters: {set(special_chars)}")
            # Remove special chars
            transformed_value = re.sub(allowed_pattern, '', str_value)
        
        # Email Format
        if rule.email_format:
            email_pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
            if not re.match(email_pattern, str_value):
                errors.append("Invalid email format")
        
        # Date Format validation
        if rule.date_format:
            try:
                pd.to_datetime(str_value, format=rule.date_format)
            except:
                try:
                    pd.to_datetime(str_value)  # Try auto-parse
                except:
                    errors.append(f"Invalid date format (expected {rule.date_format})")
        
        # Custom Regex Pattern
        if rule.regex_pattern:
            try:
                if not re.match(rule.regex_pattern, str_value):
                    errors.append(f"Does not match required pattern")
            except re.error:
                errors.append("Invalid regex pattern configured")
        
        # Apply transformation regex if provided
        if rule.transform_regex:
            try:
                match = re.search(rule.transform_regex, str_value)
                if match:
                    transformed_value = match.group(0)
            except re.error:
                pass
        
        return len(errors) == 0, errors, transformed_value

# =============================================================================
# MAPPING ENGINE
# =============================================================================

class MappingEngine:
    """AI-powered column mapping engine"""
    
    @staticmethod
    def calculate_similarity(source: str, target: str) -> float:
        """Calculate similarity score between two column names (0-100)"""
        source_clean = source.lower().replace('_', ' ').replace('-', ' ').strip()
        target_clean = target.lower().replace('_', ' ').replace('-', ' ').strip()
        
        # Exact match
        if source_clean == target_clean:
            return 100.0
        
        # Contains match
        if source_clean in target_clean or target_clean in source_clean:
            return 90.0
        
        # Word overlap
        source_words = set(source_clean.split())
        target_words = set(target_clean.split())
        
        if source_words and target_words:
            intersection = source_words.intersection(target_words)
            union = source_words.union(target_words)
            jaccard = len(intersection) / len(union) if union else 0
            return jaccard * 100
        
        return 0.0
    
    @staticmethod
    def auto_map_columns(source_cols: List[str], target_cols: List[str], 
                        threshold: float = 60.0) -> List[ColumnMapping]:
        """Generate automatic column mappings"""
        mappings = []
        used_targets = set()
        
        for source in source_cols:
            best_match = None
            best_score = 0
            
            for target in target_cols:
                if target in used_targets:
                    continue
                
                score = MappingEngine.calculate_similarity(source, target)
                if score > best_score and score >= threshold:
                    best_score = score
                    best_match = target
            
            if best_match:
                mapping = ColumnMapping(
                    source_column=source,
                    target_column=best_match,
                    confidence_score=best_score / 100
                )
                mappings.append(mapping)
                used_targets.add(best_match)
        
        return mappings

# =============================================================================
# UI COMPONENTS
# =============================================================================

def render_header():
    """Render enterprise header"""
    st.markdown("""
    <div class="enterprise-header">
        <h1 class="enterprise-title">🎯 MAPPER ENTERPRISE</h1>
        <p class="enterprise-subtitle">
            Advanced Data Migration & Validation Platform
        </p>
        <div style="margin-top: 1rem; font-size: 0.8rem; color: rgba(255,255,255,0.6);">
            Session ID: <code>{}</code> | {}
        </div>
    </div>
    """.format(
        st.session_state.session_id,
        datetime.now().strftime("%Y-%m-%d %H:%M")
    ), unsafe_allow_html=True)

def render_progress_steps():
    """Render workflow progress indicator"""
    steps = [
        ("1", "Upload Template", "template_file"),
        ("2", "Upload Source", "source_file"),
        ("3", "Configure Rules", "column_rules"),
        ("4", "Map Columns", "mappings"),
        ("5", "Validate & Export", "validation_results")
    ]
    
    # Determine current step
    current_step = st.session_state.active_tab
    
    # Add some spacing
    st.markdown("<br>", unsafe_allow_html=True)
    
    # Create columns for each step
    cols = st.columns(len(steps))
    
    for i, (num, label, state_key) in enumerate(steps):
        with cols[i]:
            if i < current_step:
                # Completed step
                st.markdown(f"""
                <div style="text-align: center;">
                    <div style="width: 40px; height: 40px; border-radius: 50%; background: #10b981; color: white; 
                                display: flex; align-items: center; justify-content: center; margin: 0 auto; 
                                font-weight: 700; font-size: 1rem;">✓</div>
                    <div style="font-size: 0.8rem; color: #94a3b8; font-weight: 500; margin-top: 0.5rem;">{label}</div>
                </div>
                """, unsafe_allow_html=True)
            elif i == current_step:
                # Active step
                st.markdown(f"""
                <div style="text-align: center;">
                    <div style="width: 40px; height: 40px; border-radius: 50%; 
                                background: linear-gradient(135deg, #3b82f6 0%, #06b6d4 100%); color: white; 
                                display: flex; align-items: center; justify-content: center; margin: 0 auto; 
                                font-weight: 700; font-size: 1rem; box-shadow: 0 0 20px rgba(59, 130, 246, 0.5);">{num}</div>
                    <div style="font-size: 0.8rem; color: #f1f5f9; font-weight: 500; margin-top: 0.5rem;">{label}</div>
                </div>
                """, unsafe_allow_html=True)
            else:
                # Pending step
                st.markdown(f"""
                <div style="text-align: center;">
                    <div style="width: 40px; height: 40px; border-radius: 50%; 
                                background: rgba(255,255,255,0.1); color: #64748b; 
                                border: 2px solid rgba(255,255,255,0.2);
                                display: flex; align-items: center; justify-content: center; margin: 0 auto; 
                                font-weight: 700; font-size: 1rem;">{num}</div>
                    <div style="font-size: 0.8rem; color: #64748b; font-weight: 500; margin-top: 0.5rem;">{label}</div>
                </div>
                """, unsafe_allow_html=True)
    
    # Add separator
    st.markdown("<hr style='margin: 2rem 0; border: 1px solid rgba(255,255,255,0.1);'>", unsafe_allow_html=True)

def render_sidebar():
    """Render sidebar controls"""
    with st.sidebar:
        st.markdown("## 🎛️ Control Panel")
        
        # Session Management
        with st.expander("💾 Session Management", expanded=True):
            if st.button("📥 Save Session State", use_container_width=True):
                state_json = save_session_state()
                st.download_button(
                    "Download Session JSON",
                    state_json,
                    file_name=f"mapper_session_{st.session_state.session_id}.json",
                    mime="application/json",
                    use_container_width=True
                )
            
            uploaded_session = st.file_uploader(
                "📤 Load Session",
                type=['json'],
                key="session_loader"
            )
            
            if uploaded_session:
                if st.button("Restore Session", use_container_width=True):
                    content = uploaded_session.read().decode()
                    if load_session_state(content):
                        st.success("✅ Session restored!")
                        st.rerun()
        
        # Quick Stats
        st.markdown("---")
        st.markdown("### 📊 Quick Stats")
        
        col1, col2 = st.columns(2)
        with col1:
            st.metric(
                "Template Sheets",
                len(st.session_state.template_file.sheets),
                delta=None
            )
        with col2:
            st.metric(
                "Source Sheets",
                len(st.session_state.source_file.sheets),
                delta=None
            )
        
        st.metric("Active Mappings", len(st.session_state.mappings))
        st.metric("Validation Rules", len(st.session_state.column_rules))
        
        # Reset
        st.markdown("---")
        if st.button("🔄 Reset Application", type="secondary", use_container_width=True):
            for key in list(st.session_state.keys()):
                del st.session_state[key]
            init_session_state()
            st.rerun()
        
        # Footer
        st.markdown("---")
        st.markdown("""
        <div style='font-size: 0.75rem; color: rgba(255,255,255,0.4); text-align: center;'>
            <strong>Mapper Enterprise v2.0</strong><br>
            Built with Streamlit<br>
            &copy; 2024 Enterprise Data Solutions
        </div>
        """, unsafe_allow_html=True)

# =============================================================================
# TAB 1: TEMPLATE FILE UPLOAD - FIXED
# =============================================================================

def render_template_tab():
    """Render Template File Upload Tab - v5 Style"""
    st.markdown('<div class="section-title">📋 Step 1: Upload Template File</div>', 
                unsafe_allow_html=True)

    st.markdown("""
    <div class="alert alert-info">
        <strong>Template File</strong> defines your target data structure. 
        Select the sheet and header row to extract column names.
    </div>
    """, unsafe_allow_html=True)

    # File Upload
    uploaded = st.file_uploader(
        "Upload Template (Excel or CSV)",
        type=['csv', 'xlsx', 'xls', 'xlsm'],
        key='template_upload',
        help="Upload your template file with the desired column structure"
    )

    if uploaded:
        # Store raw file for re-reading with different headers
        file_bytes = uploaded.getvalue()

        # Check if new file
        if st.session_state.template_file.name != uploaded.name:
            st.session_state.template_file = FileData(
                name=uploaded.name,
                raw_data=file_bytes
            )
            # Auto-read to get available sheets only
            _refresh_template_data()

        # Configuration Panel
        st.markdown('<div class="section-container">', unsafe_allow_html=True)
        st.markdown("**⚙️ Configuration**")

        # Sheet selection first (like v5)
        if len(st.session_state.template_file.sheets) > 1:
            selected_sheet = st.selectbox(
                "Select Sheet",
                st.session_state.template_file.sheets,
                index=st.session_state.template_file.sheets.index(
                    st.session_state.template_file.selected_sheet
                ) if st.session_state.template_file.selected_sheet else 0,
                key='template_sheet_select'
            )
            if selected_sheet != st.session_state.template_file.selected_sheet:
                st.session_state.template_file.selected_sheet = selected_sheet
                # Don't reload yet - wait for confirm button
        elif st.session_state.template_file.sheets:
            st.session_state.template_file.selected_sheet = st.session_state.template_file.sheets[0]
            st.info(f"📄 Sheet: {st.session_state.template_file.sheets[0]}")

        selected_sheet = st.session_state.template_file.selected_sheet

        if selected_sheet and st.session_state.template_file.raw_data:
            try:
                file_type = detect_file_type(st.session_state.template_file.name)

                # Show raw preview (like v5 - no header processing)
                raw_prev = get_header_preview(
                    st.session_state.template_file.raw_data,
                    file_type,
                    selected_sheet,
                    max_rows=8
                )

                if raw_prev is not None and len(raw_prev) > 0:
                    st.markdown("**📋 Raw Data Preview (first 8 rows):**")
                    st.dataframe(raw_prev.fillna(""), use_container_width=True, height=180)

                    # Auto-detect header row
                    detected = auto_detect_header(raw_prev)

                    # Header row number input (user selects freely)
                    max_rows = len(raw_prev) - 1 if len(raw_prev) > 1 else 10
                    header_row = st.number_input(
                        "Header row index (0 = first row)",
                        min_value=0,
                        max_value=max(10, max_rows),
                        value=detected,
                        key='template_header_row'
                    )

                    # Show what columns will be extracted
                    if header_row < len(raw_prev):
                        header_values = raw_prev.iloc[int(header_row)].tolist()
                        st.markdown(f"**✅ Will extract columns from Row {header_row}:**")
                        preview_cols = st.columns(min(4, len(header_values)))
                        for i, (col, val) in enumerate(zip(preview_cols, header_values[:4])):
                            with col:
                                val_str = str(val)[:25] + "..." if len(str(val)) > 25 else str(val)
                                st.markdown(f"<small><b>Col {i+1}:</b> {val_str}</small>", unsafe_allow_html=True)
                        if len(header_values) > 4:
                            st.markdown(f"<small><i>... and {len(header_values)-4} more columns</i></small>", unsafe_allow_html=True)

                    # Confirm button (like v5)
                    if st.button(f"✅ Confirm & Extract Columns", type="primary", use_container_width=True):
                        try:
                            with st.spinner(f"Extracting columns from {selected_sheet}..."):
                                # Update header row
                                st.session_state.template_file.header_row = int(header_row)
                                # Reload with correct header
                                _refresh_template_data()
                                
                                # Verify data was loaded
                                df_check = st.session_state.template_file.data.get(selected_sheet, pd.DataFrame())
                                if len(df_check.columns) == 0:
                                    st.error(f"❌ No columns extracted. The file may not have valid data at row {header_row}.")
                                else:
                                    st.success(f"✅ {len(df_check.columns)} columns extracted from Row {header_row}")
                                    st.rerun()
                        except Exception as e:
                            st.error(f"❌ Error extracting columns: {str(e)}")
                            import traceback
                            st.code(traceback.format_exc())
                else:
                    st.info("No preview available")

            except Exception as e:
                st.error(f"Error loading preview: {e}")

        st.markdown('</div>', unsafe_allow_html=True)

        # Debug info (remove in production)
        if st.session_state.template_file.selected_sheet:
            with st.expander("🔧 Debug Info", expanded=False):
                st.write(f"Selected sheet: {st.session_state.template_file.selected_sheet}")
                st.write(f"Header row: {st.session_state.template_file.header_row}")
                st.write(f"Available sheets: {st.session_state.template_file.sheets}")
                st.write(f"Data keys: {list(st.session_state.template_file.data.keys())}")
                df_debug = st.session_state.template_file.data.get(st.session_state.template_file.selected_sheet, pd.DataFrame())
                st.write(f"DataFrame shape: {df_debug.shape}")
                st.write(f"Total columns: {len(df_debug.columns)}")
                st.write(f"All columns:")
                for i, col in enumerate(df_debug.columns):
                    st.write(f"  {i+1}. {col}")

        # Show extracted columns (after confirmation)
        if st.session_state.template_file.selected_sheet:
            df = st.session_state.template_file.data.get(
                st.session_state.template_file.selected_sheet,
                pd.DataFrame()
            )

            if len(df.columns) > 0:
                st.markdown('<div class="section-container">', unsafe_allow_html=True)

                generic_columns = [col for col in df.columns if col.startswith('Column_') or col.startswith('Unnamed:')]
                if generic_columns:
                    st.warning(f"⚠️ Found {len(generic_columns)} generic column names out of {len(df.columns)} total columns")

                st.metric("Template Columns", len(df.columns))

                with st.expander("📋 View All Columns", expanded=True):
                    cols_list = list(df.columns)
                    for i in range(0, len(cols_list), 4):
                        cols = st.columns(4)
                        for j, col in enumerate(cols_list[i:i+4]):
                            with cols[j]:
                                st.markdown(f"<code>{col}</code>", unsafe_allow_html=True)

                st.markdown('</div>', unsafe_allow_html=True)

                # Initialize column rules if not present (needed for mapping)
                for col in df.columns:
                    if col not in st.session_state.column_rules:
                        st.session_state.column_rules[col] = ColumnRule(column_name=col)

                # Mark step as complete
                if 'template_file' not in st.session_state.completed_steps:
                    st.session_state.completed_steps.add('template_file')

                # Next step button
                st.success("✅ Template configured successfully!")
                if st.button("➡️ Proceed to Data Source", type="primary", use_container_width=True):
                    st.session_state.active_tab = 1
                    st.rerun()

# =============================================================================
# TAB 2: DATA SOURCE FILE UPLOAD - FIXED
# =============================================================================

def render_source_tab():
    """Render Data Source File Upload Tab - v5 Style"""
    st.markdown('<div class="section-title">📤 Step 2: Upload Data Source File</div>', 
                unsafe_allow_html=True)

    st.markdown("""
    <div class="alert alert-info">
        <strong>Data Source File</strong> contains your raw data that needs to be 
        mapped and transformed to match the template structure.
    </div>
    """, unsafe_allow_html=True)

    # File Upload
    uploaded = st.file_uploader(
        "Upload Data Source (Excel or CSV)",
        type=['csv', 'xlsx', 'xls', 'xlsm'],
        key='source_upload',
        help="Upload your source data file"
    )

    if uploaded:
        file_bytes = uploaded.getvalue()

        # Check if new file
        if st.session_state.source_file.name != uploaded.name:
            st.session_state.source_file = FileData(
                name=uploaded.name,
                raw_data=file_bytes
            )
            # Reset data until user confirms
            st.session_state.source_file.data = {}
            st.session_state.source_file.columns = {}

        # Configuration Panel
        st.markdown('<div class="section-container">', unsafe_allow_html=True)
        st.markdown("**⚙️ Configuration**")

        is_csv = uploaded.name.lower().endswith('.csv')

        if is_csv:
            # CSV handling (like v5)
            st.info("📄 CSV file detected")

            try:
                # Show preview with header selection
                raw_prev = pd.read_csv(io.BytesIO(file_bytes), nrows=5)
                all_cols = list(raw_prev.columns)

                # Column selection (like v5)
                sel_cols = st.multiselect("Select columns to load", all_cols, default=all_cols)

                if sel_cols:
                    st.dataframe(raw_prev[sel_cols].fillna(""), use_container_width=True, height=160)

                # Load button (like v5)
                if st.button("📊 Load Source Data", type="primary", use_container_width=True):
                    with st.spinner("Loading..."):
                        # Read full CSV with selected columns
                        df = pd.read_csv(io.BytesIO(file_bytes), usecols=sel_cols if sel_cols else None)
                        df = df.dropna(how='all')

                        st.session_state.source_file.data = {'CSV': df}
                        st.session_state.source_file.sheets = ['CSV']
                        st.session_state.source_file.columns = {'CSV': list(df.columns)}
                        st.session_state.source_file.selected_sheet = 'CSV'

                        st.success(f"✅ {len(df.columns)} fields · {len(df):,} records loaded")
                        st.rerun()

            except Exception as e:
                st.error(f"Error loading CSV: {e}")

        else:
            # Excel handling (like v5)
            try:
                xl = pd.ExcelFile(io.BytesIO(file_bytes))
                src_sheets = xl.sheet_names

                # Sheet selection (like v5)
                chosen_s = st.selectbox("Source sheet", src_sheets,
                                        index=src_sheets.index(st.session_state.source_file.selected_sheet) 
                                        if st.session_state.source_file.selected_sheet in src_sheets else 0,
                                        key='source_sheet_select') if len(src_sheets) > 1 else src_sheets[0]

                if len(src_sheets) == 1:
                    st.info(f"📄 Sheet: {src_sheets[0]}")
                    chosen_s = src_sheets[0]

                # Header row selection (like v5)
                src_hdr = st.number_input("Header row index", 0, 10, 
                                          st.session_state.source_file.header_row,
                                          key='source_header_row')

                # Show preview
                pv = pd.read_excel(io.BytesIO(file_bytes), sheet_name=chosen_s, 
                                   header=src_hdr, engine="openpyxl", nrows=5)
                all_cols = list(pv.columns)

                # Column selection (like v5)
                sel = st.multiselect("Select columns", all_cols, default=all_cols)

                if sel:
                    st.dataframe(pv[sel].fillna(""), use_container_width=True, height=160)

                # Load button (like v5)
                if st.button("📊 Load Source Data", type="primary", use_container_width=True):
                    with st.spinner("Loading..."):
                        try:
                            # Read full sheet with selected columns
                            df = pd.read_excel(io.BytesIO(file_bytes), sheet_name=chosen_s,
                                              header=src_hdr, engine="openpyxl", usecols=sel if sel else None)
                            df = df.dropna(how='all')

                            st.session_state.source_file.data = {chosen_s: df}
                            st.session_state.source_file.sheets = [chosen_s]
                            st.session_state.source_file.columns = {chosen_s: list(df.columns)}
                            st.session_state.source_file.selected_sheet = chosen_s
                            st.session_state.source_file.header_row = src_hdr

                            st.success(f"✅ {len(df.columns)} fields · {len(df):,} records loaded")
                            st.rerun()
                        except Exception as e:
                            st.error(f"Load failed: {e}")

            except Exception as e:
                st.error(f"Error reading Excel: {e}")

        st.markdown('</div>', unsafe_allow_html=True)

        # Data Preview (after load)
        if st.session_state.source_file.selected_sheet:
            df = st.session_state.source_file.data.get(
                st.session_state.source_file.selected_sheet,
                pd.DataFrame()
            )

            if len(df.columns) > 0:
                st.markdown('<div class="section-container">', unsafe_allow_html=True)

                cols = st.columns(4)
                cols[0].metric("Rows", len(df))
                cols[1].metric("Columns", len(df.columns))
                cols[2].metric("Memory", f"{df.memory_usage(deep=True).sum() / 1024:.1f} KB")
                cols[3].metric("Null Values", df.isnull().sum().sum())

                st.markdown("**📊 Data Preview**")

                display_df = df.head(100) if len(df) > 0 else pd.DataFrame(columns=df.columns)

                column_config = {}
                for col in df.columns:
                    column_config[col] = st.column_config.TextColumn(
                        col,
                        help=f"Type: {df[col].dtype}",
                        width="medium"
                    )

                st.data_editor(
                    display_df,
                    column_config=column_config,
                    use_container_width=True,
                    height=400,
                    key='source_preview_editor',
                    disabled=True
                )

                if len(df) == 0:
                    st.info("ℹ️ Source file has header row but no data rows.")

                with st.expander("📋 View All Columns"):
                    cols_list = list(df.columns)
                    for i in range(0, len(cols_list), 4):
                        cols = st.columns(4)
                        for j, col in enumerate(cols_list[i:i+4]):
                            with cols[j]:
                                st.markdown(f"<code>{col}</code>", unsafe_allow_html=True)

                st.markdown('</div>', unsafe_allow_html=True)

                if 'source_file' not in st.session_state.completed_steps:
                    st.session_state.completed_steps.add('source_file')

                st.success("✅ Data Source file loaded successfully!")

                col1, col2 = st.columns(2)
                with col1:
                    if st.button("⬅️ Back to Template", use_container_width=True):
                        st.session_state.active_tab = 0
                        st.rerun()
                with col2:
                    if st.button("➡️ Configure Rules", type="primary", use_container_width=True):
                        st.session_state.active_tab = 2
                        st.rerun()

# =============================================================================
# TAB 3: VALIDATION RULES CONFIGURATION
# =============================================================================

def render_rules_tab():
    """Render Validation Rules Configuration Tab"""
    st.markdown('<div class="section-title">🛡️ Step 3: Configure Validation Rules</div>', 
                unsafe_allow_html=True)
    
    if not st.session_state.template_file.selected_sheet:
        st.warning("⚠️ Please upload Template file first")
        return
    
    template_cols = st.session_state.template_file.columns.get(
        st.session_state.template_file.selected_sheet, []
    )
    
    if not template_cols:
        st.error("❌ No columns found in template")
        return
    
    st.markdown("""
    <div class="alert alert-info">
        Define validation rules for each <strong>Template Column</strong>. 
        These rules will be applied when mapping data from the source.
    </div>
    """, unsafe_allow_html=True)
    
    # Summary stats
    total_rules = sum(
        1 for rule in st.session_state.column_rules.values() 
        if rule.get_active_rules()
    )
    st.metric("Active Rules", f"{total_rules}/{len(template_cols)}")
    
    # Column selection for rule editing
    selected_col = st.selectbox(
        "Select Column to Configure",
        template_cols,
        key='rule_column_select'
    )
    
    if selected_col:
        rule = st.session_state.column_rules.get(
            selected_col, 
            ColumnRule(column_name=selected_col)
        )
        
        st.markdown('<div class="section-container">', unsafe_allow_html=True)
        
        # Two column layout
        col1, col2 = st.columns(2)
        
        with col1:
            st.markdown("**📋 Basic Validation**")
            
            rule.is_mandatory = st.toggle(
                "Mandatory Field",
                value=rule.is_mandatory,
                help="Column must be mapped from source",
                key=f"mand_{selected_col}"
            )
            
            rule.not_null = st.toggle(
                "Not NULL / Empty",
                value=rule.not_null,
                help="Values cannot be empty or null",
                key=f"null_{selected_col}"
            )
            
            rule.no_special_chars = st.toggle(
                "No Special Characters",
                value=rule.no_special_chars,
                help="Removes special chars, keeps alphanumeric and spaces",
                key=f"spec_{selected_col}"
            )
            
            rule.only_characters = st.toggle(
                "Characters Only (No Numbers)",
                value=rule.only_characters,
                help="No numeric characters allowed",
                key=f"char_{selected_col}"
            )
            
            rule.only_numbers = st.toggle(
                "Numbers Only",
                value=rule.only_numbers,
                help="Must be numeric value (int or float)",
                key=f"num_{selected_col}"
            )
            
            rule.email_format = st.toggle(
                "Email Format",
                value=rule.email_format,
                help="Must be valid email address",
                key=f"email_{selected_col}"
            )
        
        with col2:
            st.markdown("**⚙️ Advanced Settings**")
            
            length_cols = st.columns(2)
            with length_cols[0]:
                rule.min_length = st.number_input(
                    "Min Length",
                    min_value=0,
                    max_value=1000,
                    value=rule.min_length or 0,
                    key=f"minlen_{selected_col}"
                )
                if rule.min_length == 0:
                    rule.min_length = None
            
            with length_cols[1]:
                rule.max_length = st.number_input(
                    "Max Length",
                    min_value=0,
                    max_value=1000,
                    value=rule.max_length or 0,
                    key=f"maxlen_{selected_col}"
                )
                if rule.max_length == 0:
                    rule.max_length = None
            
            rule.date_format = st.text_input(
                "Date Format (e.g., %Y-%m-%d)",
                value=rule.date_format or "",
                help="Python strptime format",
                key=f"date_{selected_col}"
            )
            if rule.date_format == "":
                rule.date_format = None
            
            rule.regex_pattern = st.text_input(
                "Validation Regex Pattern",
                value=rule.regex_pattern or "",
                help="Custom regex pattern for validation",
                key=f"regex_{selected_col}"
            )
            if rule.regex_pattern == "":
                rule.regex_pattern = None
            
            rule.transform_regex = st.text_input(
                "Transform/Extract Regex",
                value=rule.transform_regex or "",
                help="Extract specific pattern from value",
                key=f"trans_{selected_col}"
            )
            if rule.transform_regex == "":
                rule.transform_regex = None
            
            rule.default_value = st.text_input(
                "Default Value (if empty)",
                value=rule.default_value or "",
                key=f"default_{selected_col}"
            )
            if rule.default_value == "":
                rule.default_value = None
            
            rule.description = st.text_area(
                "Description",
                value=rule.description,
                height=68,
                key=f"desc_{selected_col}"
            )
        
        # Save rule
        st.session_state.column_rules[selected_col] = rule
        
        # Show active rules as badges
        active_rules = rule.get_active_rules()
        if active_rules:
            st.markdown("**Active Rules:**")
            badges_html = "".join([
                f'<span class="badge badge-rule">{r}</span>' 
                for r in active_rules
            ])
            st.markdown(badges_html, unsafe_allow_html=True)
        
        st.markdown('</div>', unsafe_allow_html=True)
        
        # Rules overview table
        with st.expander("📊 View All Column Rules", expanded=False):
            rules_data = []
            for col in template_cols:
                r = st.session_state.column_rules.get(col, ColumnRule(column_name=col))
                rules_data.append({
                    "Column": col,
                    "Mandatory": "✓" if r.is_mandatory else "",
                    "Not Null": "✓" if r.not_null else "",
                    "Rules": ", ".join(r.get_active_rules())
                })
            
            st.dataframe(
                pd.DataFrame(rules_data),
                use_container_width=True,
                hide_index=True
            )
        
        # Navigation
        st.markdown("---")
        col1, col2 = st.columns(2)
        with col1:
            if st.button("⬅️ Back to Data Source", use_container_width=True):
                st.session_state.active_tab = 1
                st.rerun()
        with col2:
            if st.button("➡️ Proceed to Mapping", type="primary", use_container_width=True):
                st.session_state.active_tab = 3
                st.rerun()

# =============================================================================
# TAB 4: COLUMN MAPPING
# =============================================================================

def render_mapping_tab():
    """Render Column Mapping Tab with Left-Right UI"""
    st.markdown('<div class="section-title">🗺️ Step 4: Column Mapping</div>', 
                unsafe_allow_html=True)
    
    if not st.session_state.template_file.selected_sheet or not st.session_state.source_file.selected_sheet:
        st.warning("⚠️ Please upload both Template and Data Source files first")
        return
    
    template_cols = st.session_state.template_file.columns.get(
        st.session_state.template_file.selected_sheet, []
    )
    source_cols = st.session_state.source_file.columns.get(
        st.session_state.source_file.selected_sheet, []
    )
    
    if not template_cols or not source_cols:
        st.error("❌ No columns available for mapping")
        return
    
    # Show current sheets and headers info
    st.markdown(f"""
    <div style="background: rgba(59, 130, 246, 0.1); border-radius: 8px; padding: 1rem; margin-bottom: 1rem;">
        <div style="display: flex; justify-content: space-between; align-items: center;">
            <div>
                <b>📋 Template:</b> {st.session_state.template_file.selected_sheet} 
                <span style="color: #94a3b8;">({len(template_cols)} columns)</span>
            </div>
            <div style="color: #60a5fa; font-size: 1.5rem;">↔</div>
            <div>
                <b>📤 Source:</b> {st.session_state.source_file.selected_sheet}
                <span style="color: #94a3b8;">({len(source_cols)} columns)</span>
            </div>
        </div>
    </div>
    """, unsafe_allow_html=True)
    
    st.markdown("""
    <div class="alert alert-info">
        Map columns from <strong>Source</strong> (left) to <strong>Template</strong> (right). 
        Select which source column should populate each template column.
    </div>
    """, unsafe_allow_html=True)
    
    # Auto-map controls
    col1, col2, col3 = st.columns([1, 1, 2])
    with col1:
        if st.button("🤖 Auto-Map Columns", use_container_width=True):
            with st.spinner("Analyzing column similarities..."):
                auto_mappings = MappingEngine.auto_map_columns(source_cols, template_cols)
                
                # Clear existing and add new auto-mappings
                st.session_state.mappings = []
                for mapping in auto_mappings:
                    mapping.source_sheet = st.session_state.source_file.selected_sheet
                    mapping.target_sheet = st.session_state.template_file.selected_sheet
                    st.session_state.mappings.append(mapping)
                
                st.session_state.auto_map_triggered = True
                st.success(f"✅ Created {len(auto_mappings)} automatic mappings")
                st.rerun()
    
    with col2:
        if st.button("🧹 Clear All Mappings", use_container_width=True):
            st.session_state.mappings = []
            st.session_state.auto_map_triggered = False
            st.rerun()
    
    with col3:
        mapped_count = len([m for m in st.session_state.mappings if m.is_active])
        st.metric("Active Mappings", f"{mapped_count}/{len(template_cols)}")
    
    # Left-Right Mapping Interface
    st.markdown('<div class="mapping-container">', unsafe_allow_html=True)
    
    # Build current mappings lookup
    current_mappings = {}
    for m in st.session_state.mappings:
        current_mappings[m.target_column] = m.source_column
    
    # Create the mapping interface - Left (Source) to Right (Template)
    for target in template_cols:
        rule = st.session_state.column_rules.get(target, ColumnRule(column_name=target))
        
        # Show badges for validation rules
        badges = []
        if rule.is_mandatory:
            badges.append('<span class="badge badge-mandatory">M</span>')
        if rule.not_null:
            badges.append('<span class="badge badge-rule">N</span>')
        
        badges_html = " ".join(badges)
        
        col_left, col_arrow, col_right = st.columns([2, 0.5, 2])
        
        with col_left:
            # Source column selection
            current_value = current_mappings.get(target, "")
            
            # Find index of current value
            options = ["-- Select Source Column --"] + source_cols
            try:
                current_index = options.index(current_value) if current_value in options else 0
            except:
                current_index = 0
            
            selected = st.selectbox(
                f"Source for '{target}'",
                options,
                index=current_index,
                key=f"map_{target}",
                label_visibility="collapsed"
            )
            
            # Update mapping if changed
            selected_clean = selected if selected != "-- Select Source Column --" else ""
            if selected_clean != current_value:
                # Remove old mapping for this target
                st.session_state.mappings = [m for m in st.session_state.mappings if m.target_column != target]
                # Add new mapping if selected
                if selected_clean:
                    new_mapping = ColumnMapping(
                        target_column=target,
                        source_column=selected_clean,
                        target_sheet=st.session_state.template_file.selected_sheet,
                        source_sheet=st.session_state.source_file.selected_sheet
                    )
                    st.session_state.mappings.append(new_mapping)
        
        with col_arrow:
            st.markdown(f"""
            <div style="text-align: center; padding-top: 0.5rem; color: #3b82f6; font-weight: bold; font-size: 1.2rem;">
                →
            </div>
            """, unsafe_allow_html=True)
        
        with col_right:
            st.markdown(f"""
            <div style="background: rgba(30, 41, 59, 0.8); border: 1px solid rgba(59, 130, 246, 0.3); 
                        border-radius: 8px; padding: 0.75rem; margin-bottom: 0.5rem;">
                <div style="font-weight: 600; color: #60a5fa;">{target}</div>
                <div style="margin-top: 0.25rem;">{badges_html}</div>
            </div>
            """, unsafe_allow_html=True)
    
    st.markdown('</div>', unsafe_allow_html=True)
    
    # Validation check for mandatory columns
    unmapped_mandatory = []
    for target in template_cols:
        rule = st.session_state.column_rules.get(target, ColumnRule(column_name=target))
        if rule.is_mandatory:
            if not any(m.target_column == target for m in st.session_state.mappings):
                unmapped_mandatory.append(target)
    
    if unmapped_mandatory:
        st.markdown(
            f'<div class="alert alert-warning">'
            f'<strong>⚠️ Warning:</strong> Mandatory columns not mapped: '
            f'{", ".join(unmapped_mandatory)}</div>',
            unsafe_allow_html=True
        )
    
    # Show current mappings summary
    with st.expander("📋 View Current Mappings", expanded=False):
        if st.session_state.mappings:
            mapping_summary = []
            for m in st.session_state.mappings:
                mapping_summary.append({
                    "Source Column": m.source_column,
                    "→": "→",
                    "Target Column": m.target_column,
                    "Status": "✓ Active" if m.is_active else "✗ Inactive"
                })
            st.dataframe(pd.DataFrame(mapping_summary), use_container_width=True, hide_index=True)
        else:
            st.info("No mappings configured yet")
    
    # Navigation
    st.markdown("---")
    col1, col2 = st.columns(2)
    with col1:
        if st.button("⬅️ Back to Rules", use_container_width=True):
            st.session_state.active_tab = 2
            st.rerun()
    with col2:
        if st.button("➡️ Validate & Export", type="primary", use_container_width=True):
            if len(st.session_state.mappings) == 0:
                st.error("❌ Please create at least one mapping")
            else:
                st.session_state.active_tab = 4
                st.rerun()

# =============================================================================
# TAB 5: VALIDATION & EXPORT
# =============================================================================

def render_validation_tab():
    """Render Validation and Export Tab"""
    st.markdown('<div class="section-title">✅ Step 5: Validation & Export</div>', 
                unsafe_allow_html=True)
    
    if not st.session_state.mappings:
        st.warning("⚠️ Please configure column mappings first")
        return
    
    # Execute validation button
    if st.button("🚀 Execute Validation & Transform", type="primary", use_container_width=True):
        with st.spinner("Processing data... This may take a moment"):
            execute_validation_pipeline()
    
    # Display results if available
    if st.session_state.validation_results:
        results = st.session_state.validation_results
        
        # Summary cards
        st.markdown('<div class="section-container">', unsafe_allow_html=True)
        
        cols = st.columns(4)
        with cols[0]:
            st.metric(
                "Total Rows",
                results.get('total_rows', 0)
            )
        with cols[1]:
            valid_pct = results.get('valid_percentage', 0)
            st.metric(
                "Valid Rows",
                results.get('valid_rows', 0),
                delta=f"{valid_pct:.1f}%"
            )
        with cols[2]:
            error_pct = results.get('error_percentage', 0)
            st.metric(
                "Rows with Errors",
                results.get('error_rows', 0),
                delta=f"{error_pct:.1f}%",
                delta_color="inverse"
            )
        with cols[3]:
            st.metric(
                "Total Errors",
                results.get('total_errors', 0)
            )
        
        st.markdown('</div>', unsafe_allow_html=True)
        
        # Error details
        if results.get('errors'):
            st.markdown('<div class="section-container">', unsafe_allow_html=True)
            st.markdown("**❌ Validation Errors**")
            
            error_df = pd.DataFrame(results['errors'])
            
            # Filter options
            filter_col = st.selectbox(
                "Filter by Column",
                ["All"] + list(error_df['target_column'].unique())
            )
            
            if filter_col != "All":
                error_df = error_df[error_df['target_column'] == filter_col]
            
            st.dataframe(
                error_df,
                use_container_width=True,
                height=300,
                column_config={
                    "row": st.column_config.NumberColumn("Row #"),
                    "source_column": st.column_config.TextColumn("Source"),
                    "target_column": st.column_config.TextColumn("Target"),
                    "value": st.column_config.TextColumn("Value"),
                    "error": st.column_config.TextColumn("Error")
                }
            )
            st.markdown('</div>', unsafe_allow_html=True)
        
        # Transformed data preview
        if st.session_state.transformed_data is not None:
            st.markdown('<div class="section-container">', unsafe_allow_html=True)
            st.markdown("**📊 Transformed Data Preview**")
            
            df = st.session_state.transformed_data
            
            # Show sample with color coding for errors
            st.dataframe(
                df.head(100),
                use_container_width=True,
                height=400
            )
            
            # Export options
            st.markdown("**📥 Export Options**")
            
            col1, col2 = st.columns(2)
            
            with col1:
                # Excel export - preserve entire template structure, replace only selected sheet
                if st.session_state.template_file.raw_data:
                    try:
                        # Import openpyxl for preserving formatting
                        from openpyxl import load_workbook
                        from openpyxl.utils.dataframe import dataframe_to_rows
                        
                        # Load the original template workbook to preserve all sheets and formatting
                        wb = load_workbook(io.BytesIO(st.session_state.template_file.raw_data))
                        
                        # Get the selected sheet name
                        selected_sheet = st.session_state.template_file.selected_sheet
                        
                        # Check if the selected sheet exists in the workbook
                        if selected_sheet in wb.sheetnames:
                            # Clear existing data in selected sheet but preserve formatting
                            ws = wb[selected_sheet]
                            
                            # Clear all data but keep formatting (clear from row 2 onwards, keep headers if they exist)
                            # Find the last row and column
                            max_row = ws.max_row
                            max_col = ws.max_column
                            
                            # Clear data rows (keep header row if it exists)
                            for row in range(2, max_row + 1):
                                for col in range(1, max_col + 1):
                                    ws.cell(row=row, column=col).value = None
                            
                            # Add the mapped data starting from row 2 (after header)
                            if len(df) > 0:
                                # Convert dataframe to rows
                                rows = dataframe_to_rows(df, index=False, header=False)
                                
                                # Write data starting from row 2
                                for r_idx, row in enumerate(rows, 2):
                                    for c_idx, value in enumerate(row, 1):
                                        if c_idx <= max_col:  # Don't exceed original column count
                                            ws.cell(row=r_idx, column=c_idx).value = value
                            
                            # Save the modified workbook
                            output = io.BytesIO()
                            wb.save(output)
                            
                            # Use original filename
                            original_filename = st.session_state.template_file.name
                            if not original_filename.lower().endswith('.xlsx'):
                                original_filename = original_filename.rsplit('.', 1)[0] + '_export.xlsx'
                            
                            st.download_button(
                                label="📄 Download Excel (.xlsx)",
                                data=output.getvalue(),
                                file_name=original_filename,
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                use_container_width=True
                            )
                        else:
                            # Fallback if sheet not found
                            st.error(f"❌ Selected sheet '{selected_sheet}' not found in template")
                            output = io.BytesIO()
                            with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                                df.to_excel(writer, sheet_name='Mapped_Data', index=False)
                                if results.get('errors'):
                                    error_df = pd.DataFrame(results['errors'])
                                    error_df.to_excel(writer, sheet_name='Validation_Errors', index=False)
                            
                            st.download_button(
                                label="📄 Download Excel (.xlsx)",
                                data=output.getvalue(),
                                file_name=f"template_export_{st.session_state.session_id}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                use_container_width=True
                            )
                            
                    except ImportError:
                        st.warning("⚠️ openpyxl not available, using fallback export")
                        # Fallback without formatting preservation
                        output = io.BytesIO()
                        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                            df.to_excel(writer, sheet_name='Mapped_Data', index=False)
                            if results.get('errors'):
                                error_df = pd.DataFrame(results['errors'])
                                error_df.to_excel(writer, sheet_name='Validation_Errors', index=False)
                        
                        st.download_button(
                            label="📄 Download Excel (.xlsx)",
                            data=output.getvalue(),
                            file_name=f"template_export_{st.session_state.session_id}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True
                        )
                    except Exception as e:
                        st.error(f"❌ Error preserving template format: {str(e)}")
                        # Ultimate fallback
                        output = io.BytesIO()
                        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                            df.to_excel(writer, sheet_name='Mapped_Data', index=False)
                            if results.get('errors'):
                                error_df = pd.DataFrame(results['errors'])
                                error_df.to_excel(writer, sheet_name='Validation_Errors', index=False)
                        
                        st.download_button(
                            label="📄 Download Excel (.xlsx)",
                            data=output.getvalue(),
                            file_name=f"template_export_{st.session_state.session_id}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True
                        )
                else:
                    # No original template available
                    st.warning("⚠️ Original template not available for format preservation")
                    output = io.BytesIO()
                    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                        df.to_excel(writer, sheet_name='Mapped_Data', index=False)
                        if results.get('errors'):
                            error_df = pd.DataFrame(results['errors'])
                            error_df.to_excel(writer, sheet_name='Validation_Errors', index=False)
                    
                    st.download_button(
                        label="📄 Download Excel (.xlsx)",
                        data=output.getvalue(),
                        file_name=f"template_export_{st.session_state.session_id}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
            
            with col2:
                # CSV export
                csv = df.to_csv(index=False)
                st.download_button(
                    label="📄 Download CSV",
                    data=csv,
                    file_name=f"mapped_data_{st.session_state.session_id}.csv",
                    mime="text/csv",
                    use_container_width=True
                )
            
            st.markdown('</div>', unsafe_allow_html=True)
    
    # Navigation
    st.markdown("---")
    if st.button("⬅️ Back to Mapping", use_container_width=True):
        st.session_state.active_tab = 3
        st.rerun()

def execute_validation_pipeline():
    """Execute the full validation and transformation pipeline"""
    try:
        source_df = st.session_state.source_file.data.get(
            st.session_state.source_file.selected_sheet,
            pd.DataFrame()
        )
        
        template_cols = st.session_state.template_file.columns.get(
            st.session_state.template_file.selected_sheet,
            []
        )
        
        # Initialize target dataframe
        transformed_data = pd.DataFrame(columns=template_cols)
        errors = []
        valid_rows = 0
        error_rows = 0
        
        # Process each row
        for idx, row in source_df.iterrows():
            new_row = {}
            row_has_error = False
            
            for mapping in st.session_state.mappings:
                if not mapping.is_active:
                    continue
                
                source_val = row.get(mapping.source_column, None)
                target_col = mapping.target_column
                rule = st.session_state.column_rules.get(
                    target_col, 
                    ColumnRule(column_name=target_col)
                )
                
                # Validate and transform
                is_valid, val_errors, transformed_val = ValidationEngine.validate_value(
                    source_val, rule
                )
                
                if not is_valid:
                    row_has_error = True
                    for err in val_errors:
                        errors.append({
                            'row': idx + 1,
                            'source_column': mapping.source_column,
                            'target_column': target_col,
                            'value': source_val,
                            'error': err
                        })
                
                new_row[target_col] = transformed_val
            
            # Handle unmapped columns
            for col in template_cols:
                if col not in new_row:
                    rule = st.session_state.column_rules.get(col, ColumnRule(column_name=col))
                    if rule.default_value:
                        new_row[col] = rule.default_value
                    else:
                        new_row[col] = None
                    
                    if rule.is_mandatory:
                        row_has_error = True
                        errors.append({
                            'row': idx + 1,
                            'source_column': 'N/A',
                            'target_column': col,
                            'value': None,
                            'error': 'Mandatory column not mapped'
                        })
            
            # Add row to result
            transformed_data = pd.concat([
                transformed_data, 
                pd.DataFrame([new_row])
            ], ignore_index=True)
            
            if row_has_error:
                error_rows += 1
            else:
                valid_rows += 1
        
        # Update session state
        total_rows = len(source_df)
        st.session_state.transformed_data = transformed_data
        st.session_state.validation_results = {
            'total_rows': total_rows,
            'valid_rows': valid_rows,
            'error_rows': error_rows,
            'valid_percentage': (valid_rows / total_rows * 100) if total_rows > 0 else 0,
            'error_percentage': (error_rows / total_rows * 100) if total_rows > 0 else 0,
            'total_errors': len(errors),
            'errors': errors[:500]  # Limit stored errors
        }
        
        st.success("✅ Validation complete!")
        
    except Exception as e:
        st.error(f"❌ Error during processing: {str(e)}")
        import traceback
        st.error(traceback.format_exc())

# =============================================================================
# MAIN APPLICATION
# =============================================================================

def main():
    """Main application entry point"""
    init_session_state()
    render_header()
    render_sidebar()
    render_progress_steps()
    
    # Main tabs
    tabs = st.tabs([
        "📋 Template Upload",
        "📤 Source Upload", 
        "🛡️ Validation Rules",
        "🗺️ Column Mapping",
        "✅ Validate & Export"
    ])
    
    with tabs[0]:
        render_template_tab()
    
    with tabs[1]:
        render_source_tab()
    
    with tabs[2]:
        render_rules_tab()
    
    with tabs[3]:
        render_mapping_tab()
    
    with tabs[4]:
        render_validation_tab()
    
    # Auto-navigate to active tab (for button navigation)
    if st.session_state.active_tab != 0:
        # Use JavaScript to switch tabs
        tab_names = ["Template Upload", "Source Upload", "Validation Rules", "Column Mapping", "Validate & Export"]
        js = f"""
        <script>
            var tabs = window.parent.document.querySelectorAll('[data-baseweb="tab"]');
            if (tabs.length > {st.session_state.active_tab}) {{
                tabs[{st.session_state.active_tab}].click();
            }}
        </script>
        """
        st.components.v1.html(js, height=0)

if __name__ == "__main__":
    main()