import streamlit as st
import pandas as pd
import numpy as np
import re
import json
import hashlib
import io
from datetime import datetime
from typing import Dict, List, Optional, Tuple, Any, Set
from dataclasses import dataclass, asdict, field
from enum import Enum
import warnings
import time
warnings.filterwarnings('ignore')

# Page configuration - MUST be first
st.set_page_config(
    page_title="Mapper Enterprise | Data Migration Platform",
    page_icon="Target",
    layout="wide",
    initial_sidebar_state="expanded"
)

# =============================================================================
# CUSTOM CSS STYLING
# =============================================================================
st.markdown("""
<style>
    /* Button styling - Blue background with white text */
    .stButton>button {
        background-color: #0000D1 !important;
        color: white !important;
        border: none !important;
        border-radius: 8px !important;
        padding: 10px 24px !important;
        font-weight: 600 !important;
        transition: all 0.3s ease !important;
    }
    
    .stButton>button:hover {
        background-color: #0000a8 !important;
        color: white !important;
    }
    
    .stButton>button:active {
        background-color: #000080 !important;
    }
    
    /* Download button styling */
    .stDownloadButton>button {
        background-color: #0000D1 !important;
        color: white !important;
        border: none !important;
        border-radius: 8px !important;
        padding: 10px 24px !important;
        font-weight: 600 !important;
    }
    
    .stDownloadButton>button:hover {
        background-color: #0000a8 !important;
        color: white !important;
    }
    
    /* Tab styling - Active tab green with cool effects */
    .stTabs [data-baseweb="tab-list"] {
        gap: 8px !important;
    }
    
    .stTabs [data-baseweb="tab"] {
        border-radius: 12px !important;
        padding: 12px 24px !important;
        transition: all 0.3s ease !important;
        border: 2px solid transparent !important;
    }
    
    .stTabs [aria-selected="true"] {
        background: linear-gradient(135deg, #10b981 0%, #059669 100%) !important;
        color: white !important;
        border-radius: 12px !important;
        box-shadow: 0 4px 15px rgba(16, 185, 129, 0.4) !important;
        border: 2px solid rgba(255, 255, 255, 0.3) !important;
        transform: translateY(-2px) !important;
        font-weight: 600 !important;
    }
    
    .stTabs [aria-selected="false"] {
        background-color: rgba(255, 255, 255, 0.05) !important;
        color: #94a3b8 !important;
        border-radius: 12px !important;
    }
    
    .stTabs [aria-selected="false"]:hover {
        background-color: rgba(16, 185, 129, 0.1) !important;
        color: #10b981 !important;
        border: 2px solid rgba(16, 185, 129, 0.3) !important;
    }
</style>
""", unsafe_allow_html=True)

# Loading spinner
with st.spinner('🔄 Loading application...'):
    time.sleep(0.5)  # Brief pause to show spinner

# =============================================================================
# DATA CLASSES
# =============================================================================

class ValidationType(Enum):
    MANDATORY = "mandatory"
    NOT_NULL = "not_null"
    NO_SPECIAL_CHARS = "no_special_chars"
    MAX_LENGTH = "max_length"
    MIN_LENGTH = "min_length"
    ONLY_CHARACTERS = "only_characters"
    ONLY_NUMBERS = "only_numbers"
    EMAIL_FORMAT = "email_format"
    DATE_FORMAT = "date_format"
    PHONE_FORMAT = "phone_format"
    URL_FORMAT = "url_format"
    POSTAL_CODE = "postal_code"
    SSN_FORMAT = "ssn_format"
    CREDIT_CARD = "credit_card"
    IP_ADDRESS = "ip_address"
    CURRENCY_FORMAT = "currency_format"
    PERCENTAGE = "percentage"
    BOOLEAN_FORMAT = "boolean_format"
    NUMERIC_RANGE = "numeric_range"
    UPPERCASE_ONLY = "uppercase_only"
    LOWERCASE_ONLY = "lowercase_only"
    TITLE_CASE = "title_case"
    SENTENCE_CASE = "sentence_case"
    CAMEL_CASE = "camel_case"
    LOWER_CAMEL_CASE = "lower_camel_case"
    SNAKE_CASE = "snake_case"
    TITLE_CASE_STRICT = "title_case_strict"
    STARTS_WITH = "starts_with"
    ENDS_WITH = "ends_with"
    CONTAINS = "contains"
    UNIQUE_VALUE = "unique_value"
    CHECKSUM_VALIDATION = "checksum_validation"
    AGE_VALIDATION = "age_validation"
    REGEX_PATTERN = "regex_pattern"

@dataclass
class ColumnRule:
    """Validation rule configuration for a column"""
    column_name: str
    is_mandatory: bool = False
    not_null: bool = False
    no_special_chars: bool = False
    max_length: Optional[int] = None
    min_length: Optional[int] = None
    only_characters: bool = False
    only_numbers: bool = False
    email_format: bool = False
    date_format: Optional[str] = None
    phone_format: bool = False
    url_format: bool = False
    postal_code: bool = False
    ssn_format: bool = False
    credit_card: bool = False
    ip_address: bool = False
    currency_format: bool = False
    percentage: bool = False
    boolean_format: bool = False
    numeric_range_min: Optional[float] = None
    numeric_range_max: Optional[float] = None
    uppercase_only: bool = False
    lowercase_only: bool = False
    title_case: bool = False
    sentence_case: bool = False
    camel_case: bool = False
    lower_camel_case: bool = False
    snake_case: bool = False
    title_case_strict: bool = False
    starts_with: Optional[str] = None
    ends_with: Optional[str] = None
    contains: Optional[str] = None
    unique_value: bool = False
    checksum_validation: bool = False
    age_validation_min: Optional[int] = None
    age_validation_max: Optional[int] = None
    regex_pattern: Optional[str] = None
    default_value: Optional[str] = None
    description: str = ""
    
    def get_active_rules(self) -> List[str]:
        """Get list of active validation rules"""
        rules = []
        if self.is_mandatory:
            rules.append("Mandatory")
        if self.not_null:
            rules.append("Not Null")
        if self.no_special_chars:
            rules.append("No Special Chars")
        if self.max_length:
            rules.append(f"Max Len: {self.max_length}")
        if self.min_length:
            rules.append(f"Min Len: {self.min_length}")
        if self.only_characters:
            rules.append("Chars Only")
        if self.only_numbers:
            rules.append("Numbers Only")
        if self.email_format:
            rules.append("Email")
        if self.date_format:
            rules.append(f"Date: {self.date_format}")
        if self.phone_format:
            rules.append("Phone")
        if self.url_format:
            rules.append("URL")
        if self.postal_code:
            rules.append("Postal Code")
        if self.ssn_format:
            rules.append("SSN")
        if self.credit_card:
            rules.append("Credit Card")
        if self.ip_address:
            rules.append("IP Address")
        if self.currency_format:
            rules.append("Currency")
        if self.percentage:
            rules.append("Percentage")
        if self.boolean_format:
            rules.append("Boolean")
        if self.numeric_range_min is not None or self.numeric_range_max is not None:
            min_val = self.numeric_range_min if self.numeric_range_min is not None else ""
            max_val = self.numeric_range_max if self.numeric_range_max is not None else ""
            rules.append(f"Range: {min_val}-{max_val}")
        if self.uppercase_only:
            rules.append("Uppercase")
        if self.lowercase_only:
            rules.append("Lowercase")
        if self.title_case:
            rules.append("Title Case")
        if self.sentence_case:
            rules.append("Sentence Case")
        if self.camel_case:
            rules.append("CamelCase")
        if self.lower_camel_case:
            rules.append("lowerCamelCase")
        if self.snake_case:
            rules.append("snake_case")
        if self.title_case_strict:
            rules.append("Title Case (Strict)")
        if self.starts_with:
            rules.append(f"Starts: '{self.starts_with}'")
        if self.ends_with:
            rules.append(f"Ends: '{self.ends_with}'")
        if self.contains:
            rules.append(f"Contains: '{self.contains}'")
        if self.unique_value:
            rules.append("Unique")
        if self.checksum_validation:
            rules.append("Checksum")
        if self.age_validation_min is not None or self.age_validation_max is not None:
            min_age = self.age_validation_min if self.age_validation_min is not None else ""
            max_age = self.age_validation_max if self.age_validation_max is not None else ""
            rules.append(f"Age: {min_age}-{max_age}")
        if self.regex_pattern:
            rules.append("Regex")
        return rules

@dataclass
class ColumnMapping:
    """Mapping between source and target columns"""
    id: str = field(default_factory=lambda: hashlib.md5(str(datetime.now().timestamp()).encode()).hexdigest()[:8])
    source_column: str = ""
    target_column: str = ""
    source_sheet: str = ""
    target_sheet: str = ""
    transform_rules: List[str] = field(default_factory=list)
    validation_errors: List[Dict] = field(default_factory=list)
    is_active: bool = True
    confidence_score: float = 0.0

@dataclass
class FileData:
    """Container for uploaded file data"""
    name: str = ""
    data: Dict[str, pd.DataFrame] = field(default_factory=dict)
    sheets: List[str] = field(default_factory=list)
    all_sheets: List[str] = field(default_factory=list)  # Store all available sheet names
    columns: Dict[str, List[str]] = field(default_factory=dict)
    selected_sheet: Optional[str] = None
    header_row: int = 0
    raw_data: Optional[bytes] = None

# =============================================================================
# SESSION STATE MANAGEMENT
# =============================================================================

def init_session_state():
    """Initialize session state with all required keys"""
    defaults = {
        # Workflow State
        'active_tab': 0,
        'completed_steps': set(),
        
        # File Data
        'template_file': FileData(),
        'source_file': FileData(),
        
        # Configuration
        'column_rules': {},  # column_name -> ColumnRule
        'mappings': [],  # List[ColumnMapping]
        
        # Processing Results
        'validation_results': {},
        'transformed_data': None,
        'processing_log': [],
        
        # UI State
        'show_mapping_panel': False,
        'auto_map_triggered': False,
        'selected_mapping_id': None,
        
        # Session
        'session_id': hashlib.md5(str(datetime.now()).encode()).hexdigest()[:8]
    }
    
    for key, value in defaults.items():
        if key not in st.session_state:
            st.session_state[key] = value

def save_session_state():
    """Export session state to JSON"""
    state_dict = {
        'column_rules': {k: asdict(v) for k, v in st.session_state.column_rules.items()},
        'mappings': [asdict(m) for m in st.session_state.mappings],
        'template_file': {
            'name': st.session_state.template_file.name,
            'selected_sheet': st.session_state.template_file.selected_sheet,
            'header_row': st.session_state.template_file.header_row,
        },
        'source_file': {
            'name': st.session_state.source_file.name,
            'selected_sheet': st.session_state.source_file.selected_sheet,
            'header_row': st.session_state.source_file.header_row,
        },
        'session_id': st.session_state.session_id
    }
    return json.dumps(state_dict, indent=2, default=str)

def load_session_state(json_data: str):
    """Import session state from JSON"""
    try:
        data = json.loads(json_data)
        
        # Restore column rules
        st.session_state.column_rules = {
            k: ColumnRule(**v) for k, v in data.get('column_rules', {}).items()
        }
        
        # Restore mappings
        st.session_state.mappings = [ColumnMapping(**m) for m in data.get('mappings', [])]
        
        # Restore file metadata (files themselves need to be re-uploaded)
        if 'template_file' in data:
            st.session_state.template_file.name = data['template_file'].get('name', '')
            st.session_state.template_file.selected_sheet = data['template_file'].get('selected_sheet')
            st.session_state.template_file.header_row = data['template_file'].get('header_row', 0)
        
        if 'source_file' in data:
            st.session_state.source_file.name = data['source_file'].get('name', '')
            st.session_state.source_file.selected_sheet = data['source_file'].get('selected_sheet')
            st.session_state.source_file.header_row = data['source_file'].get('header_row', 0)
        
        return True
    except Exception as e:
        st.error(f"Error loading session: {e}")
        return False

# =============================================================================
# FILE HANDLING - FIXED VERSION
# =============================================================================

def detect_file_type(file_name: str) -> str:
    """Detect file type from extension"""
    if file_name.lower().endswith('.csv'):
        return 'csv'
    elif file_name.lower().endswith(('.xlsx', '.xls', '.xlsm')):
        return 'excel'
    return 'unknown'

def clean_column_names(columns, header_row=None):
    """
    Clean column names by handling unnamed columns and stripping whitespace.
    Preserves actual header values from the selected row.
    Only replaces truly empty/NaN values with Column_N placeholders.
    """
    cleaned = []
    for i, col in enumerate(columns):
        col_str = str(col).strip()
        
        # Check if this is a truly empty/invalid value
        is_empty = (
            col_str.lower() in ['nan', 'none', ''] or 
            pd.isna(col) or
            col_str == 'nan' or
            col_str.startswith('Unnamed:')  # Pandas generated placeholder
        )
        
        if is_empty:
            # Replace with generic name only for truly empty values
            cleaned.append(f"Column_{i+1}")
        else:
            # Keep the original value (even if it starts with Column_)
            cleaned.append(col_str)
    return cleaned

def is_generic_column(column_name: str) -> bool:
    """
    Check if a column name is auto-generated placeholder (Column_X, Unnamed:).
    Returns True only for pandas-generated placeholders, not actual column names.
    """
    col_str = str(column_name).strip()
    
    # Only check for pandas auto-generated patterns
    generic_patterns = [
        r'^Column_\d+$',           # Column_1, Column_2, Column_92, etc. (pandas generated)
        r'^Unnamed:\s*\d+$',       # Unnamed: 0, Unnamed: 1, etc. (pandas generated)
    ]
    
    for pattern in generic_patterns:
        if re.match(pattern, col_str, re.IGNORECASE):
            return True
    
    return False

def filter_generic_columns(columns: List[str]) -> List[str]:
    """
    Filter out only pandas auto-generated column names.
    Returns all actual columns from the file.
    """
    return [col for col in columns if not is_generic_column(col)]

def filter_empty_and_generic_columns(df: pd.DataFrame, threshold: float = 0.95) -> pd.DataFrame:
    """
    Filter out columns that are:
    1. Auto-generated placeholders (Column_X, Unnamed:)
    2. Completely empty (100% null values)
    
    Keeps all actual columns from the file, even if they have names like "Role 1", "Field 1", etc.
    
    Args:
        df: DataFrame to filter
        threshold: Not used anymore, kept for compatibility
    
    Returns:
        Filtered DataFrame with only actual columns from the file
    """
    if df.empty or len(df.columns) == 0:
        return df
    
    valid_columns = []
    
    for col in df.columns:
        # Check if column name is auto-generated placeholder
        if is_generic_column(col):
            continue
        
        # Check if column is completely empty (all null)
        if df[col].isna().all():
            continue
        
        # Keep this column - it's an actual column from the file
        valid_columns.append(col)
    
    # Return filtered dataframe
    if len(valid_columns) > 0:
        return df[valid_columns]
    else:
        # If no valid columns found, return empty dataframe
        return pd.DataFrame()

def parse_dates_in_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """
    Automatically detect and parse date/datetime columns in various formats.
    Supports: ISO 8601, date-only, datetime, timestamps, and common date formats.
    """
    for col in df.columns:
        # Skip if already datetime
        if pd.api.types.is_datetime64_any_dtype(df[col]):
            continue
            
        # Only try to parse object/string columns
        if df[col].dtype == 'object' or df[col].dtype == 'string':
            try:
                # Try to parse as datetime (pandas now infers format by default)
                # This handles ISO 8601, common formats, timestamps, etc.
                parsed = pd.to_datetime(df[col], errors='coerce')
                
                # Only convert if at least 50% of non-null values were successfully parsed
                non_null_count = df[col].notna().sum()
                parsed_count = parsed.notna().sum()
                
                if non_null_count > 0 and (parsed_count / non_null_count) >= 0.5:
                    df[col] = parsed
            except:
                # If parsing fails, leave as is
                pass
    
    return df

def read_uploaded_file(uploaded_file, header_row: int = 0, selected_sheet: str = None) -> Tuple[Dict[str, pd.DataFrame], List[str], Dict[str, List[str]]]:
    """
    Read uploaded file and return data dictionary, sheet names, and column mapping.
    Uses the specified header_row to read column names dynamically.
    Ensures column headers are always captured even if no data rows exist.
    """
    if uploaded_file is None:
        return {}, [], {}
    
    try:
        # Handle both uploaded files and BytesIO objects
        if hasattr(uploaded_file, 'name'):
            file_type = detect_file_type(uploaded_file.name)
            if hasattr(uploaded_file, 'getvalue'):
                file_bytes = uploaded_file.getvalue()
            else:
                file_bytes = uploaded_file.read()
                uploaded_file.seek(0)
        else:
            # For BytesIO objects
            if hasattr(uploaded_file, 'getvalue'):
                file_bytes = uploaded_file.getvalue()
            else:
                file_bytes = uploaded_file.read()
                uploaded_file.seek(0)
            file_type = 'excel'
            # Try to detect from content
            try:
                pd.ExcelFile(io.BytesIO(file_bytes))
                file_type = 'excel'
            except:
                file_type = 'csv'
        
        data = {}
        
        if file_type == 'csv':
            # First, read without headers to get the raw data preview
            raw_df = pd.read_csv(io.BytesIO(file_bytes), header=None, nrows=50)
            
            # Extract column names from the selected header row if it exists
            if header_row < len(raw_df):
                column_names = clean_column_names(raw_df.iloc[header_row].tolist(), header_row)
            else:
                # Fallback to generic names if header row is beyond data
                column_names = [f"Column_{i+1}" for i in range(len(raw_df.columns))]
            
            # Read again with the specified header row
            df = pd.read_csv(io.BytesIO(file_bytes), header=header_row)
            
            # Ensure dataframe has the correct column names
            if len(df) == 0:
                # No data rows - create empty df with extracted column names
                if len(column_names) > 0:
                    df = pd.DataFrame(columns=column_names)
                # else: keep pandas default columns from header row
            else:
                # Has data rows - clean the column names from header row
                df.columns = clean_column_names(df.columns, header_row)
            
            # Drop rows that are completely empty
            df = df.dropna(how='all')
            
            # Parse date columns automatically
            df = parse_dates_in_dataframe(df)
            
            data['Sheet1'] = df
            sheets = ['Sheet1']
            
        elif file_type == 'excel':
            # Read all sheets from Excel
            excel_file = pd.ExcelFile(io.BytesIO(file_bytes))
            all_sheets = excel_file.sheet_names
            
            # If specific sheet requested, only process that one
            sheets_to_process = [selected_sheet] if selected_sheet and selected_sheet in all_sheets else all_sheets
            
            for sheet in sheets_to_process:
                try:
                    
                    # Read the full sheet WITHOUT headers first
                    raw_df = pd.read_excel(io.BytesIO(file_bytes), sheet_name=sheet, header=None)
                    
                    
                    # Get total rows and columns
                    total_rows = len(raw_df)
                    total_cols = len(raw_df.columns)
                    
                    
                    if total_rows == 0:
                        continue
                    
                    # Extract column names from the selected header row
                    if header_row < total_rows:
                        header_values = raw_df.iloc[header_row].tolist()
                        
                        # Convert to strings and use as column names directly
                        column_names = []
                        for i, val in enumerate(header_values):
                            if pd.isna(val) or str(val).strip() == '' or str(val).lower() == 'nan':
                                column_names.append(f"Column_{i+1}")
                            else:
                                column_names.append(str(val).strip())
                    else:
                        # Fallback if header row is beyond data
                        column_names = [f"Column_{i+1}" for i in range(total_cols)]
                    
                    # Extract data rows (everything after header row)
                    if header_row + 1 < total_rows:
                        data_df = raw_df.iloc[header_row + 1:].copy()
                        data_df.columns = column_names
                        data_df = data_df.reset_index(drop=True)
                    else:
                        # No data rows - create empty dataframe with headers
                        data_df = pd.DataFrame(columns=column_names)
                    
                    # Only drop rows that are completely empty, NOT columns
                    # This preserves all columns the user selected
                    data_df = data_df.dropna(how='all')
                    
                    # Parse date columns automatically
                    data_df = parse_dates_in_dataframe(data_df)
                    
                    data[sheet] = data_df
                except Exception as e:
                    st.warning(f" Error reading sheet '{sheet}': {e}")
                    continue
                    
            sheets = list(data.keys())
        else:
            st.error(" Unsupported file format. Please upload CSV or Excel files.")
            return {}, [], {}
        
        # Extract columns for each sheet and filter out auto-generated placeholders
        columns = {}
        for sheet, df in data.items():
            # Apply filtering: remove Column_X, Unnamed:, and completely empty columns
            filtered_df = filter_empty_and_generic_columns(df)
            valid_cols = list(filtered_df.columns)
            columns[sheet] = valid_cols
            # Update the dataframe to only include valid columns
            data[sheet] = filtered_df
        
        return data, sheets, columns
        
    except Exception as e:
        st.error(f" Error reading file: {str(e)}")
        import traceback
        st.error(traceback.format_exc())
        return {}, [], []

def get_header_preview(file_bytes: bytes, file_type: str, sheet_name: str = None, max_rows: int = 10) -> Optional[pd.DataFrame]:
    """
    Get raw data preview for header row selection (without header processing).
    Returns dataframe with NO header to show raw values at each row.
    """
    try:
        if file_type == 'csv':
            df = pd.read_csv(io.BytesIO(file_bytes), header=None, nrows=max_rows)
            return df
        else:
            df = pd.read_excel(io.BytesIO(file_bytes), sheet_name=sheet_name, header=None, nrows=max_rows)
            return df
    except Exception as e:
        return None

def auto_detect_header(df: pd.DataFrame, max_rows: int = 10) -> int:
    """
    Auto-detect header row by analyzing first few rows
    Returns the row index most likely to be the header
    """
    if len(df) == 0:
        return 0
    
    best_row = 0
    best_score = 0
    
    for i in range(min(max_rows, len(df))):
        row = df.iloc[i]
        
        # Skip completely empty rows
        if row.isna().all():
            continue
            
        # Score based on multiple factors
        string_count = sum(1 for x in row if isinstance(x, str) and str(x).strip())
        string_ratio = string_count / len(row) if len(row) > 0 else 0
        
        # Check uniqueness (headers usually have unique values)
        non_null_values = [x for x in row if pd.notna(x) and str(x).strip()]
        unique_ratio = len(set(non_null_values)) / len(non_null_values) if non_null_values else 0
        
        # Prefer rows with meaningful text (not just numbers or single characters)
        meaningful_text_count = sum(1 for x in row if isinstance(x, str) and len(str(x).strip()) > 2)
        meaningful_ratio = meaningful_text_count / len(row) if len(row) > 0 else 0
        
        # Check for common header patterns
        header_keywords = ['id', 'name', 'action', 'type', 'code', 'number', 'date', 'status', 'description']
        keyword_matches = sum(1 for x in row if isinstance(x, str) and 
                             any(keyword in str(x).lower() for keyword in header_keywords))
        keyword_ratio = keyword_matches / len(row) if len(row) > 0 else 0
        
        # Combined score: prioritize string content, uniqueness, meaningful text, and header keywords
        score = (string_ratio * 0.3) + (unique_ratio * 0.3) + (meaningful_ratio * 0.2) + (keyword_ratio * 0.2)
        
        if score > best_score:
            best_score = score
            best_row = i
    
    return best_row

def _refresh_template_data():
    """Refresh template data from stored bytes with current header row setting"""
    if st.session_state.template_file.raw_data:
        # Ensure we have bytes and create BytesIO object
        if isinstance(st.session_state.template_file.raw_data, bytes):
            file_obj = io.BytesIO(st.session_state.template_file.raw_data)
        else:
            file_obj = st.session_state.template_file.raw_data
            
        # Pass selected sheet to only read that sheet
        data, sheets, columns = read_uploaded_file(
            file_obj,
            st.session_state.template_file.header_row,
            st.session_state.template_file.selected_sheet
        )
        st.session_state.template_file.data = data
        st.session_state.template_file.sheets = sheets
        st.session_state.template_file.columns = columns
        # Ensure selected_sheet is set to the sheet we actually read
        if sheets and st.session_state.template_file.selected_sheet not in sheets:
            st.session_state.template_file.selected_sheet = sheets[0] if sheets else None

def _refresh_source_data():
    """Refresh source data from stored bytes with current header row setting"""
    if st.session_state.source_file.raw_data:
        # Ensure we have bytes and create BytesIO object
        if isinstance(st.session_state.source_file.raw_data, bytes):
            file_obj = io.BytesIO(st.session_state.source_file.raw_data)
        else:
            file_obj = st.session_state.source_file.raw_data
            
        # Pass selected sheet to only read that sheet
        data, sheets, columns = read_uploaded_file(
            file_obj,
            st.session_state.source_file.header_row,
            st.session_state.source_file.selected_sheet
        )
        st.session_state.source_file.data = data
        st.session_state.source_file.sheets = sheets
        st.session_state.source_file.columns = columns
        # Ensure selected_sheet is set to the sheet we actually read
        if sheets and st.session_state.source_file.selected_sheet not in sheets:
            st.session_state.source_file.selected_sheet = sheets[0] if sheets else None

# =============================================================================
# UNIFIED REGEX ENGINE
# =============================================================================

import re

class RegexEngine:
    """Enterprise-grade unified regex engine for validation, extraction, transformation, and more."""
    
    def __init__(self, pattern: str):
        self.pattern = pattern
        try:
            self.compiled = re.compile(pattern)
            self.is_valid = True
        except re.error as e:
            self.compiled = None
            self.is_valid = False
            self.error = str(e)
    
    def detect_operation(self) -> str:
        """Automatically detect the intended regex operation based on pattern structure."""
        if not self.pattern:
            return "search"
        
        pattern = self.pattern.strip()
        
        # Validation: starts with ^ and ends with $
        if pattern.startswith("^") and pattern.endswith("$"):
            return "validate"
        
        # Remove characters: starts with [^
        if pattern.startswith("[^"):
            return "remove"
        
        # Split: contains | or , as delimiter
        if "|" in pattern or ("," in pattern and not pattern.startswith("[")):
            return "split"
        
        # Extraction: contains capture groups or common extraction patterns
        if "(" in pattern or "\\d" in pattern or "\\w" in pattern:
            return "extract"
        
        # Default fallback: search
        return "search"
    
    def process(self, value: Any) -> Dict[str, Any]:
        """
        Process a value using the appropriate regex operation.
        Returns a dict with operation type and result.
        """
        if not self.is_valid:
            return {
                "operation": "error",
                "error": self.error,
                "result": str(value) if value else ""
            }
        
        if value is None:
            value = ""
        
        value_str = str(value)
        operation = self.detect_operation()
        
        try:
            if operation == "validate":
                # Full match validation
                match = self.compiled.fullmatch(value_str)
                return {
                    "operation": "validate",
                    "result": bool(match),
                    "matched": bool(match)
                }
            
            elif operation == "remove":
                # Remove matching characters
                cleaned = self.compiled.sub("", value_str)
                return {
                    "operation": "transform",
                    "result": cleaned,
                    "original": value_str
                }
            
            elif operation == "extract":
                # Extract matching patterns
                matches = self.compiled.findall(value_str)
                # If single capture group, flatten the list
                if matches and isinstance(matches[0], tuple) and len(matches[0]) == 1:
                    matches = [m[0] for m in matches]
                return {
                    "operation": "extract",
                    "result": matches,
                    "count": len(matches)
                }
            
            elif operation == "split":
                # Split by pattern
                parts = self.compiled.split(value_str)
                return {
                    "operation": "split",
                    "result": parts,
                    "parts_count": len(parts)
                }
            
            elif operation == "search":
                # Search for first match
                match = self.compiled.search(value_str)
                return {
                    "operation": "search",
                    "result": match.group(0) if match else None,
                    "found": bool(match)
                }
            
            else:
                return {
                    "operation": "unknown",
                    "result": value_str,
                    "error": f"Unknown operation: {operation}"
                }
                
        except Exception as e:
            return {
                "operation": operation,
                "error": str(e),
                "result": value_str
            }


# =============================================================================
# VALIDATION ENGINE
# =============================================================================

class ValidationEngine:
    """Enterprise-grade validation engine"""
    
    @staticmethod
    def validate_value(value: Any, rule: ColumnRule) -> Tuple[bool, List[str], Any]:
        """
        Validate and transform a single value
        Returns: (is_valid, error_messages, transformed_value)
        """
        errors = []
        
        # Handle null/empty values
        is_empty = pd.isna(value) or str(value).strip() == '' or value is None
        
        if is_empty:
            if rule.not_null:
                errors.append("Value cannot be null/empty")
            if rule.is_mandatory:
                errors.append("Mandatory field is empty")
            # Apply default value if provided
            if rule.default_value and is_empty:
                value = rule.default_value
            return len(errors) == 0, errors, value
        
        str_value = str(value).strip()
        transformed_value = str_value
        
        # Max Length
        if rule.max_length and len(str_value) > rule.max_length:
            errors.append(f"Length {len(str_value)} exceeds maximum {rule.max_length}")
            transformed_value = str_value[:rule.max_length]
        
        # Min Length
        if rule.min_length and len(str_value) < rule.min_length:
            errors.append(f"Length {len(str_value)} below minimum {rule.min_length}")
        
        # Only Characters (no digits)
        if rule.only_characters:
            if any(char.isdigit() for char in str_value):
                errors.append("Contains numeric characters")
            # Remove digits
            transformed_value = ''.join(char for char in str_value if not char.isdigit())
        
        # Only Numbers
        if rule.only_numbers:
            cleaned = str_value.replace(',', '').replace(' ', '').replace('$', '')
            try:
                float(cleaned)
                transformed_value = float(cleaned)
            except ValueError:
                errors.append("Not a valid numeric value")
        
        # No Special Characters (allows Latin/Unicode letters, numbers, spaces)
        if rule.no_special_chars:
            # Allow alphanumeric, spaces, and common punctuation if required make changes here
            allowed_pattern = r'[^\w\s\-\&]'
            special_chars = re.findall(allowed_pattern, str_value)
            if special_chars:
                errors.append(f"Contains special characters: {set(special_chars)}")
            # Remove special chars
            transformed_value = re.sub(allowed_pattern, '', str_value)
        
        # Email Format
        if rule.email_format:
            email_pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
            if not re.match(email_pattern, str_value):
                errors.append("Invalid email format")
        
        # Date Format validation
        if rule.date_format:
            try:
                pd.to_datetime(str_value, format=rule.date_format)
            except:
                try:
                    pd.to_datetime(str_value)  # Try auto-parse
                except:
                    errors.append(f"Invalid date format (expected {rule.date_format})")
        
        # Unified Regex Engine - handles validation, extraction, transformation
        if rule.regex_pattern:
            try:
                engine = RegexEngine(rule.regex_pattern)
                result = engine.process(str_value)
                
                if result.get("operation") == "validate":
                    # Validation mode: check if pattern matches
                    if not result.get("matched", False):
                        errors.append(f"Does not match required pattern: {rule.regex_pattern}")
                        
                elif result.get("operation") == "transform":
                    # Transform mode: clean/modify the value
                    transformed_value = result.get("result", str_value)
                    
                elif result.get("operation") == "extract":
                    # Extract mode: get matches and transform
                    matches = result.get("result", [])
                    if matches:
                        # If matches is a list of tuples (multiple capture groups)
                        if isinstance(matches[0], tuple):
                            # Join all capture groups with space or format as needed
                            # For date ranges like (05, 21, 1991, 10, 15, 1995)
                            # Convert to readable format
                            if len(matches[0]) == 6:  # Date range pattern
                                # Format: MM/DD/YYYY to MM/DD/YYYY
                                start_date = f"{matches[0][0]}/{matches[0][1]}/{matches[0][2]}"
                                end_date = f"{matches[0][3]}/{matches[0][4]}/{matches[0][5]}"
                                transformed_value = f"{start_date} to {end_date}"
                            elif len(matches[0]) == 3:  # Single date pattern
                                # Format: MM/DD/YYYY
                                transformed_value = f"{matches[0][0]}/{matches[0][1]}/{matches[0][2]}"
                            else:
                                # Join all groups
                                transformed_value = " ".join(str(m) for m in matches[0] if m)
                        else:
                            # Simple list of matches
                            transformed_value = str(matches[0]) if matches else str_value
                    else:
                        # No matches found - reject the value
                        errors.append(f"Pattern not found in value: {rule.regex_pattern}")
                        
                elif result.get("operation") == "search":
                    # Search mode: find and extract first match
                    found_value = result.get("result")
                    if found_value:
                        transformed_value = found_value
                    else:
                        # No match found - reject the value
                        errors.append(f"Pattern not found in value: {rule.regex_pattern}")
                        
                elif result.get("operation") == "error":
                    # Invalid regex pattern
                    errors.append(f"Invalid regex pattern: {result.get('error', 'Unknown error')}")
                    
            except Exception as e:
                errors.append(f"Regex processing error: {str(e)}")
        
        # Phone Format
        if rule.phone_format:
            phone_pattern = r'^\+?[\d\s\-\(\)\.]{10,15}$'
            if not re.match(phone_pattern, str_value):
                errors.append("Invalid phone number format")

        # URL Format
        if rule.url_format:
            url_pattern = r'^(https?://)?(www\.)?([a-zA-Z0-9-]+\.)+[a-zA-Z]{2,}(/.*)?$'
            if not re.match(url_pattern, str_value):
                errors.append("Invalid URL format")

        # Postal Code (US ZIP)
        if rule.postal_code:
            postal_pattern = r'^\d{5}(-\d{4})?$'
            if not re.match(postal_pattern, str_value):
                errors.append("Invalid postal code format (expected: 12345 or 12345-6789)")

        # SSN Format
        if rule.ssn_format:
            ssn_pattern = r'^\d{3}-\d{2}-\d{4}$'
            if not re.match(ssn_pattern, str_value):
                errors.append("Invalid SSN format (expected: 123-45-6789)")

        # Credit Card
        if rule.credit_card:
            # Remove spaces and dashes
            clean_card = re.sub(r'[-\s]', '', str_value)
            if not clean_card.isdigit():
                errors.append("Credit card must contain only numbers")
            elif len(clean_card) < 13 or len(clean_card) > 19:
                errors.append("Credit card number length invalid")
            else:
                # Luhn algorithm validation
                def luhn_checksum(card_num):
                    def digits_of(n):
                        return [int(d) for d in str(n)]
                    digits = digits_of(card_num)
                    odd_digits = digits[-1::-2]
                    even_digits = digits[-2::-2]
                    checksum = sum(odd_digits)
                    for d in even_digits:
                        checksum += sum(digits_of(d*2))
                    return checksum % 10
                if luhn_checksum(int(clean_card)) != 0:
                    errors.append("Invalid credit card number")

        # IP Address
        if rule.ip_address:
            ip_pattern = r'^(\d{1,3}\.){3}\d{1,3}$'
            if not re.match(ip_pattern, str_value):
                errors.append("Invalid IP address format")
            else:
                # Check each octet
                octets = str_value.split('.')
                for octet in octets:
                    if not octet.isdigit() or not 0 <= int(octet) <= 255:
                        errors.append("Invalid IP address range")
                        break

        # Currency Format
        if rule.currency_format:
            currency_pattern = r'^\$?\d{1,3}(,\d{3})*(\.\d{2})?$'
            if not re.match(currency_pattern, str_value):
                errors.append("Invalid currency format (expected: $1,234.56 or 1234.56)")

        # Percentage
        if rule.percentage:
            percentage_pattern = r'^\d+(\.\d+)?%?$'
            if not re.match(percentage_pattern, str_value):
                errors.append("Invalid percentage format (expected: 25 or 25.5 or 25%)")

        # Boolean Format
        if rule.boolean_format:
            boolean_values = ['true', 'false', 'yes', 'no', '1', '0', 'y', 'n', 't', 'f']
            if str_value.lower() not in boolean_values:
                errors.append("Invalid boolean value (expected: true/false, yes/no, 1/0, y/n, t/f)")

        # Numeric Range
        if rule.numeric_range_min is not None or rule.numeric_range_max is not None:
            try:
                num_value = float(str_value)
                if rule.numeric_range_min is not None and num_value < rule.numeric_range_min:
                    errors.append(f"Value {num_value} below minimum {rule.numeric_range_min}")
                if rule.numeric_range_max is not None and num_value > rule.numeric_range_max:
                    errors.append(f"Value {num_value} above maximum {rule.numeric_range_max}")
            except ValueError:
                errors.append("Value must be numeric for range validation")

        # Case validation and transformation
        if rule.uppercase_only and str_value != str_value.upper():
            errors.append("Must be uppercase only")
            transformed_value = str_value.upper()

        if rule.lowercase_only and str_value != str_value.lower():
            errors.append("Must be lowercase only")
            transformed_value = str_value.lower()
        
        if rule.title_case:
            # Title Case: Capitalize first/last words and major words
            transformed_value = ValidationEngine._apply_title_case(str_value)
        
        if rule.sentence_case:
            # Sentence case: Capitalize only the first word
            transformed_value = ValidationEngine._apply_sentence_case(str_value)
        
        if rule.camel_case:
            # CamelCase: No spaces, capitalize first letter of each word
            transformed_value = ValidationEngine._apply_camel_case(str_value)
        
        if rule.lower_camel_case:
            # lowerCamelCase: First word lowercase, subsequent words capitalized
            transformed_value = ValidationEngine._apply_lower_camel_case(str_value)
        
        if rule.snake_case:
            # snake_case: Lowercase with underscores
            transformed_value = ValidationEngine._apply_snake_case(str_value)
        
        if rule.title_case_strict:
            # Title Case w/o Small Words: Stricter on keeping short words lowercase
            transformed_value = ValidationEngine._apply_title_case_strict(str_value)

        # String matching
        if rule.starts_with and not str_value.startswith(rule.starts_with):
            errors.append(f"Must start with '{rule.starts_with}'")

        if rule.ends_with and not str_value.endswith(rule.ends_with):
            errors.append(f"Must end with '{rule.ends_with}'")

        if rule.contains and rule.contains not in str_value:
            errors.append(f"Must contain '{rule.contains}'")

        # Unique value validation (checked at pipeline level)
        if rule.unique_value:
            errors.append("UNIQUE_CHECK")  # Placeholder for pipeline-level uniqueness check

        # Checksum validation (basic mod 10 for numeric strings)
        if rule.checksum_validation:
            if not str_value.isdigit():
                errors.append("Checksum validation requires numeric value")
            else:
                digits = [int(d) for d in str_value]
                checksum = sum(digits) % 10
                if checksum != 0:
                    errors.append("Checksum validation failed")

        # Age validation (requires date format)
        if rule.age_validation_min is not None or rule.age_validation_max is not None:
            if rule.date_format:
                try:
                    birth_date = pd.to_datetime(str_value, format=rule.date_format)
                    age = (pd.Timestamp.now() - birth_date).days // 365
                    if rule.age_validation_min is not None and age < rule.age_validation_min:
                        errors.append(f"Age {age} below minimum {rule.age_validation_min}")
                    if rule.age_validation_max is not None and age > rule.age_validation_max:
                        errors.append(f"Age {age} above maximum {rule.age_validation_max}")
                except:
                    errors.append("Invalid date for age validation")
            else:
                errors.append("Age validation requires date format to be set")

        return len(errors) == 0, errors, transformed_value
    
    @staticmethod
    def _apply_title_case(text: str) -> str:
        """Apply Title Case: Capitalize first/last words and major words"""
        # Small words that should remain lowercase (except at start/end)
        small_words = {'a', 'an', 'and', 'as', 'at', 'but', 'by', 'for', 'in', 'of', 'on', 'or', 'the', 'to', 'with'}
        
        words = text.split()
        if not words:
            return text
        
        result = []
        for i, word in enumerate(words):
            # Always capitalize first and last word
            if i == 0 or i == len(words) - 1:
                result.append(word.capitalize())
            # Keep small words lowercase unless they're at boundaries
            elif word.lower() in small_words:
                result.append(word.lower())
            else:
                result.append(word.capitalize())
        
        return ' '.join(result)
    
    @staticmethod
    def _apply_sentence_case(text: str) -> str:
        """Apply Sentence case: Capitalize only the first word"""
        if not text:
            return text
        # Capitalize first character, lowercase the rest
        return text[0].upper() + text[1:].lower() if len(text) > 1 else text.upper()
    
    @staticmethod
    def _apply_camel_case(text: str) -> str:
        """Apply CamelCase: No spaces, capitalize first letter of each word"""
        # Remove special characters and split by spaces/underscores
        words = re.split(r'[\s_-]+', text)
        # Capitalize each word and join without spaces
        return ''.join(word.capitalize() for word in words if word)
    
    @staticmethod
    def _apply_lower_camel_case(text: str) -> str:
        """Apply lowerCamelCase: First word lowercase, subsequent words capitalized"""
        # Remove special characters and split by spaces/underscores
        words = re.split(r'[\s_-]+', text)
        words = [word for word in words if word]
        
        if not words:
            return text
        
        # First word lowercase, rest capitalized
        result = [words[0].lower()]
        result.extend(word.capitalize() for word in words[1:])
        return ''.join(result)
    
    @staticmethod
    def _apply_snake_case(text: str) -> str:
        """Apply snake_case: Lowercase with underscores"""
        # Replace spaces and hyphens with underscores
        text = re.sub(r'[\s-]+', '_', text)
        # Handle camelCase by inserting underscores before capitals
        text = re.sub(r'([a-z])([A-Z])', r'\1_\2', text)
        # Convert to lowercase
        return text.lower()
    
    @staticmethod
    def _apply_title_case_strict(text: str) -> str:
        """Apply Title Case w/o Small Words: Stricter on keeping short words lowercase"""
        # Extended list of small words to keep lowercase
        small_words = {
            'a', 'an', 'and', 'as', 'at', 'but', 'by', 'for', 'from', 'in', 'into',
            'of', 'on', 'or', 'the', 'to', 'with', 'via', 'per', 'nor', 'yet', 'so'
        }
        
        words = text.split()
        if not words:
            return text
        
        result = []
        for i, word in enumerate(words):
            # Always capitalize first and last word
            if i == 0 or i == len(words) - 1:
                result.append(word.capitalize())
            # Keep small words lowercase
            elif word.lower() in small_words:
                result.append(word.lower())
            else:
                result.append(word.capitalize())
        
        return ' '.join(result)

# =============================================================================
# MAPPING ENGINE
# =============================================================================

class MappingEngine:
    """AI-powered column mapping engine"""
    
    @staticmethod
    def calculate_similarity(source: str, target: str) -> float:
        """Calculate similarity score between two column names (0-100)"""
        source_clean = source.lower().replace('_', ' ').replace('-', ' ').strip()
        target_clean = target.lower().replace('_', ' ').replace('-', ' ').strip()
        
        # Exact match
        if source_clean == target_clean:
            return 100.0
        
        # Contains match
        if source_clean in target_clean or target_clean in source_clean:
            return 90.0
        
        # Word overlap
        source_words = set(source_clean.split())
        target_words = set(target_clean.split())
        
        if source_words and target_words:
            intersection = source_words.intersection(target_words)
            union = source_words.union(target_words)
            jaccard = len(intersection) / len(union) if union else 0
            return jaccard * 100
        
        return 0.0
    
    @staticmethod
    def auto_map_columns(source_cols: List[str], target_cols: List[str], 
                        threshold: float = 60.0) -> List[ColumnMapping]:
        """Generate automatic column mappings"""
        mappings = []
        used_targets = set()
        
        for source in source_cols:
            best_match = None
            best_score = 0
            
            for target in target_cols:
                if target in used_targets:
                    continue
                
                score = MappingEngine.calculate_similarity(source, target)
                if score > best_score and score >= threshold:
                    best_score = score
                    best_match = target
            
            if best_match:
                mapping = ColumnMapping(
                    source_column=source,
                    target_column=best_match,
                    confidence_score=best_score / 100
                )
                mappings.append(mapping)
                used_targets.add(best_match)
        
        return mappings

# =============================================================================
# UI COMPONENTS
# =============================================================================

def render_header():
    """Render enterprise header"""
def render_progress_steps():
    """Render workflow progress indicator with tick marks for completed steps"""
    steps = [
        ("1", "Upload Template", "template_file", ""),
        ("2", "Upload Source", "source_file", ""),
        ("3", "Configure Rules", "column_rules", ""),
        ("4", "Map Columns", "mappings", ""),
        ("5", "Validate & Export", "validation_results", "")
    ]
    
    # Determine current step
    current_step = st.session_state.active_tab
    
    # Add some spacing
    st.markdown("<br>", unsafe_allow_html=True)
    
    # Create columns for each step
    cols = st.columns(len(steps))
    
    for i, (num, label, state_key, icon) in enumerate(steps):
        with cols[i]:
            # Check if step is completed (has required data)
            is_completed = False
            if state_key == "template_file" and st.session_state.template_file.selected_sheet:
                is_completed = True
            elif state_key == "source_file" and st.session_state.source_file.selected_sheet:
                is_completed = True
            elif state_key == "column_rules" and st.session_state.column_rules:
                # Check if any rules are actually configured (not just default empty rules)
                has_configured_rules = any(
                    rule.get_active_rules() 
                    for rule in st.session_state.column_rules.values()
                )
                if has_configured_rules:
                    is_completed = True
            elif state_key == "mappings" and st.session_state.mappings:
                is_completed = True
            elif state_key == "validation_results" and st.session_state.validation_results:
                is_completed = True
            
            if i < current_step or is_completed:
                # Completed step - show tick mark
                st.markdown(f"""
                <div style="text-align: center;">
                    <div style="width: 48px; height: 48px; border-radius: 50%; background: linear-gradient(135deg, #10b981 0%, #059669 100%); 
                                color: white; display: flex; align-items: center; justify-content: center; margin: 0 auto; 
                                font-weight: 700; font-size: 1.2rem; box-shadow: 0 4px 15px rgba(16, 185, 129, 0.3);
                                border: 3px solid rgba(255,255,255,0.2);">
                        &#10004;
                    </div>
                    <div style="font-size: 0.85rem; color: #10b981; font-weight: 600; margin-top: 0.75rem; line-height: 1.2;">
                        {label}
                    </div>
                </div>
                """, unsafe_allow_html=True)
            elif i == current_step:
                # Active step
                st.markdown(f"""
                <div style="text-align: center;">
                    <div style="width: 48px; height: 48px; border-radius: 50%; 
                                background: linear-gradient(135deg, #3b82f6 0%, #06b6d4 100%); color: white; 
                                display: flex; align-items: center; justify-content: center; margin: 0 auto; 
                                font-weight: 700; font-size: 1.2rem; box-shadow: 0 0 25px rgba(59, 130, 246, 0.5); 
                                border: 3px solid rgba(255,255,255,0.3); animation: pulse 2s infinite;">
                        {num}
                    </div>
                    <div style="font-size: 0.85rem; color: #3b82f6; font-weight: 600; margin-top: 0.75rem; line-height: 1.2;">
                        {label}
                    </div>
                </div>
                """, unsafe_allow_html=True)
            else:
                # Pending step
                st.markdown(f"""
                <div style="text-align: center;">
                    <div style="width: 48px; height: 48px; border-radius: 50%; 
                                background: rgba(255,255,255,0.08); color: #64748b; 
                                border: 2px solid rgba(255,255,255,0.15);
                                display: flex; align-items: center; justify-content: center; margin: 0 auto; 
                                font-weight: 600; font-size: 1.1rem;">
                        {num}
                    </div>
                    <div style="font-size: 0.8rem; color: #64748b; font-weight: 500; margin-top: 0.75rem; line-height: 1.2;">
                        {label}
                    </div>
                </div>
                """, unsafe_allow_html=True)
    
    # Add separator with enhanced styling
    st.markdown("""
    <hr style='margin: 2.5rem 0; border: none; height: 2px; background: linear-gradient(90deg, 
                rgba(59, 130, 246, 0.3) 0%, rgba(6, 182, 212, 0.3) 50%, rgba(16, 185, 129, 0.3) 100%); 
                border-radius: 1px;'>
    """, unsafe_allow_html=True)

def render_sidebar():
    """Render sidebar controls"""
    with st.sidebar:
        st.markdown("##  Control Panel")
        
        # Session Management
        with st.expander(" Session Management", expanded=True):
            if st.button(" Save Session State", use_container_width=True):
                state_json = save_session_state()
                st.download_button(
                    "Download Session JSON",
                    state_json,
                    file_name=f"mapper_session_{st.session_state.session_id}.json",
                    mime="application/json",
                    use_container_width=True
                )
            
            uploaded_session = st.file_uploader(
                " Load Session",
                type=['json'],
                key="session_loader"
            )
            
            if uploaded_session:
                if st.button("Restore Session", use_container_width=True):
                    content = uploaded_session.read().decode()
                    if load_session_state(content):
                        st.success(" Session restored!")
                        st.rerun()
        
        # Quick Stats
        st.markdown("---")
        st.markdown("###  Quick Stats")
        
        col1, col2 = st.columns(2)
        with col1:
            st.metric(
                "Template Sheets",
                len(st.session_state.template_file.sheets),
                delta=None
            )
        with col2:
            st.metric(
                "Source Sheets",
                len(st.session_state.source_file.sheets),
                delta=None
            )
        
        st.metric("Active Mappings", len(st.session_state.mappings))
        st.metric("Validation Rules", len(st.session_state.column_rules))
        
        # Reset
        st.markdown("---")
        if st.button(" Reset Application", type="secondary", use_container_width=True):
            for key in list(st.session_state.keys()):
                del st.session_state[key]
            init_session_state()
            st.rerun()
        
        # Footer
        st.markdown("---")
        st.markdown("""
        <div style='font-size: 0.75rem; color: rgba(255,255,255,0.4); text-align: center;'>
            <strong>Mapper Enterprise v2.0</strong><br>
            Built with Streamlit<br>
            &copy; 2024 Enterprise Data Solutions
        </div>
        """, unsafe_allow_html=True)

# =============================================================================
# TAB 1: TEMPLATE FILE UPLOAD - FIXED
# =============================================================================

def render_template_tab():
    # File Upload
    uploaded = st.file_uploader(
        "Upload Template (Excel or CSV)",
        type=['csv', 'xlsx', 'xls', 'xlsm'],
        key='template_upload',
        help="Upload your template file with the desired column structure"
    )

    if uploaded:
        # Store raw file for re-reading with different headers
        file_bytes = uploaded.getvalue()

        # Check if new file
        if st.session_state.template_file.name != uploaded.name:
            # First, get all available sheet names
            try:
                file_type = detect_file_type(uploaded.name)
                if file_type == 'excel':
                    excel_file = pd.ExcelFile(io.BytesIO(file_bytes))
                    all_available_sheets = excel_file.sheet_names
                else:
                    all_available_sheets = ['Sheet1']  # CSV only has one sheet
            except:
                all_available_sheets = []
            
            st.session_state.template_file = FileData(
                name=uploaded.name,
                raw_data=file_bytes,
                all_sheets=all_available_sheets
            )
            # Auto-read to get available sheets only
            _refresh_template_data()

        # Simple Configuration - Auto-detect everything
        st.markdown('<div class="section-container">', unsafe_allow_html=True)

        # Only show sheet selector if multiple sheets
        if len(st.session_state.template_file.sheets) > 1:
            selected_sheet = st.selectbox(
                "📄 Select Sheet",
                st.session_state.template_file.sheets,
                index=st.session_state.template_file.sheets.index(
                    st.session_state.template_file.selected_sheet
                ) if st.session_state.template_file.selected_sheet else 0,
                key='template_sheet_select'
            )
            if selected_sheet != st.session_state.template_file.selected_sheet:
                st.session_state.template_file.selected_sheet = selected_sheet
        elif st.session_state.template_file.sheets:
            st.session_state.template_file.selected_sheet = st.session_state.template_file.sheets[0]

        selected_sheet = st.session_state.template_file.selected_sheet

        if selected_sheet and st.session_state.template_file.raw_data:
            try:
                file_type = detect_file_type(st.session_state.template_file.name)
                raw_prev = get_header_preview(
                    st.session_state.template_file.raw_data,
                    file_type,
                    selected_sheet,
                    max_rows=8
                )

                if raw_prev is not None and len(raw_prev) > 0:
                    # Auto-detect header row
                    detected = auto_detect_header(raw_prev)
                    
                    # Simple expander for advanced users
                    with st.expander("⚙️ Advanced: Change Header Row", expanded=False):
                        st.dataframe(raw_prev.fillna(""), use_container_width=True, height=180)
                        max_rows = len(raw_prev) - 1 if len(raw_prev) > 1 else 10
                        header_row = st.number_input(
                            "Header row (0 = first row)",
                            min_value=0,
                            max_value=max(10, max_rows),
                            value=detected,
                            key='template_header_row'
                        )
                    
                    # Use detected header row if expander is closed
                    if 'template_header_row' not in st.session_state:
                        header_row = detected
                    else:
                        header_row = st.session_state.template_header_row

                    # Single Load button
                    if st.button(f"✅ Load Template", type="primary", use_container_width=True, key='load_template_btn'):
                        try:
                            with st.spinner("Loading..."):
                                st.session_state.template_file.header_row = int(header_row)
                                _refresh_template_data()
                                
                                df_check = st.session_state.template_file.data.get(selected_sheet, pd.DataFrame())
                                if len(df_check.columns) == 0:
                                    st.error(f"❌ No valid columns found")
                                else:
                                    st.success(f"✅ Loaded {len(df_check.columns)} columns")
                                    st.rerun()
                        except Exception as e:
                            st.error(f"❌ Error: {str(e)}")

            except Exception as e:
                st.error(f"❌ Error: {e}")

        st.markdown('</div>', unsafe_allow_html=True)

        # Show loaded data
        if st.session_state.template_file.selected_sheet:
            df = st.session_state.template_file.data.get(
                st.session_state.template_file.selected_sheet,
                pd.DataFrame()
            )

            if len(df.columns) > 0:
                st.markdown('<div class="section-container">', unsafe_allow_html=True)
                
                # Simple sheet switcher
                if len(st.session_state.template_file.all_sheets) > 1:
                    col1, col2 = st.columns([4, 1])
                    with col1:
                        new_sheet = st.selectbox(
                            "📄 Switch Sheet",
                            st.session_state.template_file.all_sheets,
                            index=st.session_state.template_file.all_sheets.index(st.session_state.template_file.selected_sheet) if st.session_state.template_file.selected_sheet in st.session_state.template_file.all_sheets else 0,
                            key='template_sheet_changer'
                        )
                    with col2:
                        st.markdown("<br>", unsafe_allow_html=True)
                        if new_sheet != st.session_state.template_file.selected_sheet:
                            if st.button("🔄", key='switch_template_sheet', use_container_width=True):
                                st.session_state.template_file.selected_sheet = new_sheet
                                _refresh_template_data()
                                st.rerun()

                st.metric("✅ Valid Columns", len(df.columns))
                
                if len(df.columns) == 0:
                    st.error("❌ No valid columns found")
                else:
                    with st.expander("📋 View Columns", expanded=False):
                        cols_list = list(df.columns)
                        for i in range(0, len(cols_list), 4):
                            cols = st.columns(4)
                            for j, col in enumerate(cols_list[i:i+4]):
                                with cols[j]:
                                    st.markdown(f"<code>{col}</code>", unsafe_allow_html=True)

                st.markdown('</div>', unsafe_allow_html=True)

                # Initialize column rules if not present (needed for mapping)
                for col in df.columns:
                    if col not in st.session_state.column_rules:
                        st.session_state.column_rules[col] = ColumnRule(column_name=col)

                # Mark step as complete
                if 'template_file' not in st.session_state.completed_steps:
                    st.session_state.completed_steps.add('template_file')
                    
                st.success("✅ Template loaded successfully!")

        # Navigation button - always at bottom of tab
        st.markdown("---")
        if st.session_state.template_file.selected_sheet and len(st.session_state.template_file.data.get(st.session_state.template_file.selected_sheet, pd.DataFrame()).columns) > 0:
            if st.button("➡️ Proceed to Data Source", type="primary", use_container_width=True, key='btn_template_to_source'):
                st.session_state.active_tab = 1
                st.rerun()
        else:
            st.info("ℹ️ Load template file to proceed")

# =============================================================================
# TAB 2: DATA SOURCE FILE UPLOAD - FIXED
# =============================================================================

def render_source_tab():
    """Render Data Source File Upload Tab"""
    st.markdown('<div class="section-title">📤 Step 2: Upload Data Source File</div>', 
                unsafe_allow_html=True)

    st.markdown("""
    <div class="alert alert-info">
        <strong>Data Source File</strong> contains your raw data that needs to be 
        mapped and transformed to match the template structure.
    </div>
    """, unsafe_allow_html=True)

    # File Upload - Always visible
    st.markdown('<div class="section-container">', unsafe_allow_html=True)
    st.markdown("**📁 File Selection**")
    
    uploaded = st.file_uploader(
        "Choose your source data file",
        type=['csv', 'xlsx', 'xls', 'xlsm'],
        key='source_upload',
        help="Upload Excel or CSV file"
    )

    if uploaded:
        file_bytes = uploaded.getvalue()

        # Check if new file
        if st.session_state.source_file.name != uploaded.name:
            # First, get all available sheet names
            try:
                file_type = detect_file_type(uploaded.name)
                if file_type == 'excel':
                    excel_file = pd.ExcelFile(io.BytesIO(file_bytes))
                    all_available_sheets = excel_file.sheet_names
                else:
                    all_available_sheets = ['CSV']
            except:
                all_available_sheets = []
            
            st.session_state.source_file = FileData(
                name=uploaded.name,
                raw_data=file_bytes,
                all_sheets=all_available_sheets
            )
            st.session_state.source_file.data = {}
            st.session_state.source_file.columns = {}

        is_csv = uploaded.name.lower().endswith('.csv')

        # Sheet Selection - Always visible for Excel
        if not is_csv:
            try:
                xl = pd.ExcelFile(io.BytesIO(file_bytes))
                src_sheets = xl.sheet_names
                
                st.markdown("**� Sheet Selection**")
                if len(src_sheets) > 1:
                    chosen_s = st.selectbox(
                        "Select sheet to load",
                        src_sheets,
                        index=src_sheets.index(st.session_state.source_file.selected_sheet) 
                        if st.session_state.source_file.selected_sheet in src_sheets else 0,
                        key='source_sheet_select'
                    )
                else:
                    st.info(f"📄 Single sheet: {src_sheets[0]}")
                    chosen_s = src_sheets[0]
            except Exception as e:
                st.error(f"❌ Error reading file: {e}")
                st.markdown('</div>', unsafe_allow_html=True)
                return
        else:
            chosen_s = 'CSV'

        # Preview and Load Section
        st.markdown("**📊 Preview & Load**")
        
        try:
            # Show preview
            if is_csv:
                df_preview = pd.read_csv(io.BytesIO(file_bytes), nrows=5)
            else:
                # Use header row from session state or default to 0
                src_hdr = st.session_state.source_file.header_row if hasattr(st.session_state.source_file, 'header_row') else 0
                
                # Advanced header row option
                with st.expander("⚙️ Advanced: Change Header Row", expanded=False):
                    src_hdr = st.number_input("Header row (0 = first row)", 0, 10, src_hdr, key='source_header_row')
                
                df_preview = pd.read_excel(io.BytesIO(file_bytes), sheet_name=chosen_s, 
                                          header=src_hdr, engine="openpyxl", nrows=5)
            
            st.dataframe(df_preview.fillna(""), use_container_width=True, height=160)

            # Load button
            if st.button("✅ Load Source Data", type="primary", use_container_width=True, key='btn_load_source'):
                with st.spinner("Loading..."):
                    try:
                        if is_csv:
                            df = pd.read_csv(io.BytesIO(file_bytes))
                            src_hdr = 0
                        else:
                            df = pd.read_excel(io.BytesIO(file_bytes), sheet_name=chosen_s,
                                              header=src_hdr, engine="openpyxl")
                        
                        df = df.dropna(how='all')
                        df = parse_dates_in_dataframe(df)
                        df = filter_empty_and_generic_columns(df)
                        valid_columns = list(df.columns)

                        st.session_state.source_file.data = {chosen_s: df}
                        st.session_state.source_file.sheets = [chosen_s]
                        st.session_state.source_file.columns = {chosen_s: valid_columns}
                        st.session_state.source_file.selected_sheet = chosen_s
                        st.session_state.source_file.header_row = src_hdr

                        st.success(f"✅ Loaded {len(valid_columns)} columns, {len(df):,} rows")
                        st.rerun()
                    except Exception as e:
                        st.error(f"❌ Error: {e}")

        except Exception as e:
            st.error(f"❌ Error: {e}")

        st.markdown('</div>', unsafe_allow_html=True)

    # Show loaded data summary
    if st.session_state.source_file.selected_sheet:
        df = st.session_state.source_file.data.get(
            st.session_state.source_file.selected_sheet,
            pd.DataFrame()
        )

        if len(df.columns) > 0:
            st.markdown('<div class="section-container">', unsafe_allow_html=True)
            st.markdown("**✅ Loaded Data Summary**")
            
            # Simple metrics
            cols = st.columns(3)
            cols[0].metric("📊 Rows", f"{len(df):,}")
            cols[1].metric("📋 Columns", len(df.columns))
            cols[2].metric("⚠️ Nulls", f"{df.isnull().sum().sum():,}")

            # Data preview in expander
            with st.expander("📊 View Data Preview", expanded=False):
                    display_df = df.head(100) if len(df) > 0 else pd.DataFrame(columns=df.columns)

                    column_config = {}
                    for col in df.columns:
                        # Determine column type based on actual data
                        dtype = df[col].dtype
                        if pd.api.types.is_integer_dtype(dtype):
                            column_config[col] = st.column_config.NumberColumn(
                                col,
                                help=f"Type: {dtype}",
                                width="medium"
                            )
                        elif pd.api.types.is_float_dtype(dtype):
                            column_config[col] = st.column_config.NumberColumn(
                                col,
                                help=f"Type: {dtype}",
                                width="medium",
                                format="%.2f"
                            )
                        elif pd.api.types.is_bool_dtype(dtype):
                            column_config[col] = st.column_config.CheckboxColumn(
                                col,
                                help=f"Type: {dtype}",
                                width="medium"
                            )
                        elif pd.api.types.is_datetime64_any_dtype(dtype):
                            column_config[col] = st.column_config.DatetimeColumn(
                                col,
                                help=f"Type: {dtype}",
                                width="medium"
                            )
                        else:
                            column_config[col] = st.column_config.TextColumn(
                                col,
                                help=f"Type: {dtype}",
                                width="medium"
                            )

                    st.data_editor(
                        display_df,
                        column_config=column_config,
                        use_container_width=True,
                        height=400,
                        key='source_preview_editor',
                        disabled=True
                    )

            # Columns list in separate expander
            with st.expander("📋 View All Columns", expanded=False):
                cols_list = list(df.columns)
                for i in range(0, len(cols_list), 4):
                    cols = st.columns(4)
                    for j, col in enumerate(cols_list[i:i+4]):
                        with cols[j]:
                            st.markdown(f"<code>{col}</code>", unsafe_allow_html=True)

            st.markdown('</div>', unsafe_allow_html=True)

            if 'source_file' not in st.session_state.completed_steps:
                st.session_state.completed_steps.add('source_file')
                
            st.success("✅ Data Source loaded successfully!")

        # Navigation buttons - always at bottom of tab
        st.markdown("---")
        if st.session_state.source_file.selected_sheet and len(st.session_state.source_file.data.get(st.session_state.source_file.selected_sheet, pd.DataFrame()).columns) > 0:
            col1, col2 = st.columns(2)
            with col1:
                if st.button("⬅️ Back to Template", use_container_width=True, key='btn_source_to_template'):
                    st.session_state.active_tab = 0
                    st.rerun()
            with col2:
                if st.button("➡️ Configure Rules", type="primary", use_container_width=True, key='btn_source_to_rules'):
                    st.session_state.active_tab = 2
                    st.rerun()
        else:
            st.info("ℹ️ Load source file to proceed")

# =============================================================================
# TAB 3: VALIDATION RULES CONFIGURATION
# =============================================================================

def render_rules_tab():
    """Render Validation Rules Configuration Tab"""
    st.markdown('<div class="section-title"> Step 3: Configure Validation Rules</div>', 
                unsafe_allow_html=True)
    
    if not st.session_state.template_file.selected_sheet:
        st.warning(" Please upload Template file first")
        return
    
    template_cols = st.session_state.template_file.columns.get(
        st.session_state.template_file.selected_sheet, []
    )
    
    if not template_cols:
        st.error(" No columns found in template")
        return
    
    st.markdown("""
    <div class="alert alert-info">
        Define validation rules for each <strong>Template Column</strong>. 
        These rules will be applied when mapping data from the source.
    </div>
    """, unsafe_allow_html=True)
    
    # Summary stats
    total_rules = sum(
        1 for rule in st.session_state.column_rules.values() 
        if rule.get_active_rules()
    )
    st.metric("Active Rules", f"{total_rules}/{len(template_cols)}")
    
    # Column selection for rule editing
    selected_col = st.selectbox(
        "Select Column to Configure",
        template_cols,
        key='rule_column_select'
    )
    
    if selected_col:
        rule = st.session_state.column_rules.get(
            selected_col, 
            ColumnRule(column_name=selected_col)
        )
        
        st.markdown('<div class="section-container">', unsafe_allow_html=True)
        
        # Three column layout for better organization
        col1, col2, col3 = st.columns(3)
        
        with col1:
            st.markdown("** Basic Validation**")
            
            rule.is_mandatory = st.toggle(
                "Mandatory Field",
                value=rule.is_mandatory,
                help="Column must be mapped from source",
                key=f"mand_{selected_col}"
            )
            
            rule.not_null = st.toggle(
                "Not NULL / Empty",
                value=rule.not_null,
                help="Values cannot be empty or null",
                key=f"null_{selected_col}"
            )
            
            rule.no_special_chars = st.toggle(
                "No Special Characters",
                value=rule.no_special_chars,
                help="Removes special chars, keeps alphanumeric and spaces",
                key=f"spec_{selected_col}"
            )
            
            rule.only_characters = st.toggle(
                "Characters Only (No Numbers)",
                value=rule.only_characters,
                help="No numeric characters allowed",
                key=f"char_{selected_col}"
            )
            
            rule.only_numbers = st.toggle(
                "Numbers Only",
                value=rule.only_numbers,
                help="Must be numeric value (int or float)",
                key=f"num_{selected_col}"
            )
            
            rule.email_format = st.toggle(
                "Email Format",
                value=rule.email_format,
                help="Must be valid email address",
                key=f"email_{selected_col}"
            )
            
            # New format validations
            rule.phone_format = st.toggle(
                "Phone Format",
                value=rule.phone_format,
                help="Must be valid phone number format",
                key=f"phone_{selected_col}"
            )
            
            rule.url_format = st.toggle(
                "URL Format",
                value=rule.url_format,
                help="Must be valid URL format",
                key=f"url_{selected_col}"
            )
            
            rule.postal_code = st.toggle(
                "Postal Code (US ZIP)",
                value=rule.postal_code,
                help="Must be valid US ZIP code format",
                key=f"postal_{selected_col}"
            )
            
            rule.ssn_format = st.toggle(
                "SSN Format",
                value=rule.ssn_format,
                help="Must be valid SSN format (123-45-6789)",
                key=f"ssn_{selected_col}"
            )

            # Regex and transforms
            st.markdown("**Patterns & Transforms**")
            
            # Azure OpenAI Chat Box
            with st.expander("🤖 AI Regex Generator (Azure OpenAI)", expanded=False):
                st.markdown("""
                <div style="background: rgba(59, 130, 246, 0.15); padding: 1rem; border-radius: 8px; border: 1px solid rgba(59, 130, 246, 0.3);">
                    <p style="margin: 0; color: #cbd5e1; font-size: 0.9rem;">
                        Ask AI to generate regex patterns for your data transformations.
                        <br><small style="color: #94a3b8;">Example: "Create a regex for 10-digit phone numbers with country code"</small>
                    </p>
                </div>
                """, unsafe_allow_html=True)
                
                # Chat input
                user_prompt = st.text_input(
                    "Describe the pattern you need:",
                    placeholder="e.g., Extract email domains from text or validate 10-digit phone numbers",
                    key=f"ai_prompt_{selected_col}"
                )
                
                # Generate button
                if st.button("Generate Regex Pattern", key=f"ai_gen_{selected_col}", use_container_width=True):
                    if user_prompt.strip():
                        with st.spinner("Generating regex pattern with Azure OpenAI..."):
                            try:
                                from openai import AzureOpenAI
                                
                                # Check if secrets exist
                                if not hasattr(st.secrets, 'endpoint') or not st.secrets.endpoint:
                                    st.error("Azure OpenAI credentials not configured in .streamlit/secrets.toml")
                                    st.info("Please add: endpoint, api_key, deployment_name, api_version")
                                    st.stop()
                                
                                client = AzureOpenAI(
                                    azure_endpoint=st.secrets.endpoint,
                                    api_key=st.secrets.api_key,
                                    api_version=st.secrets.api_version
                                )
                                
                                response = client.chat.completions.create(
                                    model=st.secrets.deployment_name,
                                    messages=[
                                        {
                                            "role": "system",
                                            "content": """You are an expert regex pattern generator. 
                                            Return ONLY the regex pattern in a code block format.
                                            If a transform is needed, return the transform regex.
                                            Keep patterns simple and practical for data validation."""
                                        },
                                        {
                                            "role": "user",
                                            "content": user_prompt
                                        }
                                    ],
                                    max_tokens=1000,
                                    temperature=0.1
                                )
                                
                                generated_pattern = response.choices[0].message.content.strip()
                                # Extract pattern from code block if present
                                if generated_pattern.startswith('```'):
                                    generated_pattern = generated_pattern.split('```')[1]
                                    if generated_pattern.startswith('regex') or generated_pattern.startswith('python'):
                                        generated_pattern = generated_pattern.split('\n', 1)[1]
                                    generated_pattern = generated_pattern.strip('`').strip()
                                
                                st.success("Pattern generated successfully!")
                                st.code(generated_pattern, language="regex")
                                
                                # Auto-fill option
                                if st.button("Use This Pattern", key=f"use_pattern_{selected_col}"):
                                    rule.regex_pattern = generated_pattern
                                    st.rerun()
                                    
                            except Exception as e:
                                st.error(f"Error generating pattern: {str(e)}")
                                st.info("Please check your Azure OpenAI credentials in .streamlit/secrets.toml")
                    else:
                        st.warning("Please describe the pattern you need first")
            
            rule.regex_pattern = st.text_input(
                "Regex Pattern",
                value=rule.regex_pattern or "",
                help="Single regex for validation, extraction, transformation, and more. AI-powered generation available.",
                key=f"regex_{selected_col}"
            )
            if rule.regex_pattern == "":
                rule.regex_pattern = None
            
            
        with col2:
            st.markdown("** Advanced Formats**")
            
            rule.credit_card = st.toggle(
                "Credit Card",
                value=rule.credit_card,
                help="Validates credit card number with Luhn algorithm",
                key=f"cc_{selected_col}"
            )
            
            rule.ip_address = st.toggle(
                "IP Address",
                value=rule.ip_address,
                help="Must be valid IPv4 address format",
                key=f"ip_{selected_col}"
            )
            
            rule.currency_format = st.toggle(
                "Currency Format",
                value=rule.currency_format,
                help="Must be valid currency format ($1,234.56)",
                key=f"currency_{selected_col}"
            )
            
            rule.percentage = st.toggle(
                "Percentage",
                value=rule.percentage,
                help="Must be valid percentage format",
                key=f"percent_{selected_col}"
            )
            
            rule.boolean_format = st.toggle(
                "Boolean Format",
                value=rule.boolean_format,
                help="Must be true/false, yes/no, 1/0, y/n, t/f",
                key=f"bool_{selected_col}"
            )
            
            # Case validation
            st.markdown("**Case Validation**")
            
            col1, col2 = st.columns(2)
            
            with col1:
                rule.uppercase_only = st.toggle(
                    "UPPERCASE",
                    value=rule.uppercase_only,
                    help="All letters are capitalized (e.g., THE CAT SAT)",
                    key=f"upper_{selected_col}"
                )
                
                rule.lowercase_only = st.toggle(
                    "lowercase",
                    value=rule.lowercase_only,
                    help="All letters are lowercase (e.g., the cat sat)",
                    key=f"lower_{selected_col}"
                )
                
                rule.title_case = st.toggle(
                    "Title Case",
                    value=rule.title_case,
                    help="Capitalize first/last words and major words (e.g., Lord of the Flies)",
                    key=f"title_{selected_col}"
                )
                
                rule.sentence_case = st.toggle(
                    "Sentence case",
                    value=rule.sentence_case,
                    help="Capitalize only the first word (e.g., The cat sat)",
                    key=f"sentence_{selected_col}"
                )
            
            with col2:
                rule.camel_case = st.toggle(
                    "CamelCase",
                    value=rule.camel_case,
                    help="No spaces, capitalize first letter of each word (e.g., CamelCase)",
                    key=f"camel_{selected_col}"
                )
                
                rule.lower_camel_case = st.toggle(
                    "lowerCamelCase",
                    value=rule.lower_camel_case,
                    help="First word lowercase, subsequent words capitalized (e.g., iPhone)",
                    key=f"lower_camel_{selected_col}"
                )
                
                rule.snake_case = st.toggle(
                    "snake_case",
                    value=rule.snake_case,
                    help="Lowercase with underscores (e.g., snake_case_example)",
                    key=f"snake_{selected_col}"
                )
                
                rule.title_case_strict = st.toggle(
                    "Title Case (Strict)",
                    value=rule.title_case_strict,
                    help="Similar to Title Case but stricter on keeping short words lowercase",
                    key=f"title_strict_{selected_col}"
                )
            
             # Special validations
            st.markdown("**Special Validations**")
            rule.unique_value = st.toggle(
                "Unique Values",
                value=rule.unique_value,
                help="All values in this column must be unique",
                key=f"unique_{selected_col}"
            )
            
            rule.checksum_validation = st.toggle(
                "Checksum Validation",
                value=rule.checksum_validation,
                help="Mod 10 checksum validation for numeric strings",
                key=f"checksum_{selected_col}"
            )
           
            # String matching
            st.markdown("**String Matching**")
            rule.starts_with = st.text_input(
                "Starts With",
                value=rule.starts_with or "",
                help="Must start with this text",
                key=f"starts_{selected_col}"
            )
            if rule.starts_with == "":
                rule.starts_with = None
            
            rule.ends_with = st.text_input(
                "Ends With",
                value=rule.ends_with or "",
                help="Must end with this text",
                key=f"ends_{selected_col}"
            )
            if rule.ends_with == "":
                rule.ends_with = None
            
            rule.contains = st.text_input(
                "Contains",
                value=rule.contains or "",
                help="Must contain this text",
                key=f"contains_{selected_col}"
            )
            if rule.contains == "":
                rule.contains = None
            
            
        with col3:
            st.markdown("** Constraints & Ranges**")
            
            # Length constraints
            length_cols = st.columns(2)
            with length_cols[0]:
                rule.min_length = st.number_input(
                    "Min Length",
                    min_value=0,
                    max_value=1000,
                    value=rule.min_length or 0,
                    key=f"minlen_{selected_col}"
                )
                if rule.min_length == 0:
                    rule.min_length = None
            
            with length_cols[1]:
                rule.max_length = st.number_input(
                    "Max Length",
                    min_value=0,
                    max_value=1000,
                    value=rule.max_length or 0,
                    key=f"maxlen_{selected_col}"
                )
                if rule.max_length == 0:
                    rule.max_length = None
            
            # Numeric range
            st.markdown("**Numeric Range**")
            range_cols = st.columns(2)
            with range_cols[0]:
                rule.numeric_range_min = st.number_input(
                    "Min Value",
                    value=rule.numeric_range_min,
                    key=f"range_min_{selected_col}",
                    help="Minimum numeric value"
                )
            
            with range_cols[1]:
                rule.numeric_range_max = st.number_input(
                    "Max Value", 
                    value=rule.numeric_range_max,
                    key=f"range_max_{selected_col}",
                    help="Maximum numeric value"
                )
            
            # Date and age validation
            st.markdown("**Date & Age**")
            rule.date_format = st.text_input(
                "Date Format",
                value=rule.date_format or "",
                help="Python strptime format (e.g., %Y-%m-%d)",
                key=f"date_{selected_col}"
            )
            if rule.date_format == "":
                rule.date_format = None
            
            age_cols = st.columns(2)
            with age_cols[0]:
                rule.age_validation_min = st.number_input(
                    "Min Age",
                    min_value=0,
                    max_value=150,
                    value=rule.age_validation_min or 0,
                    key=f"age_min_{selected_col}"
                )
                if rule.age_validation_min == 0:
                    rule.age_validation_min = None
            
            with age_cols[1]:
                rule.age_validation_max = st.number_input(
                    "Max Age",
                    min_value=0,
                    max_value=150,
                    value=rule.age_validation_max or 0,
                    key=f"age_max_{selected_col}"
                )
                if rule.age_validation_max == 0:
                    rule.age_validation_max = None
            
            
            
            # Default value and description
            rule.default_value = st.text_input(
                "Default Value",
                value=rule.default_value or "",
                help="Value to use when field is empty",
                key=f"default_{selected_col}"
            )
            if rule.default_value == "":
                rule.default_value = None
            
            rule.description = st.text_area(
                "Description",
                value=rule.description,
                height=60,
                help="Optional description of this validation rule",
                key=f"desc_{selected_col}"
            )
        
        # Save rule
        st.session_state.column_rules[selected_col] = rule
        
        # Show active rules as badges
        active_rules = rule.get_active_rules()
        if active_rules:
            st.markdown("**Active Rules:**")
            badges_html = "".join([
                f'<span class="badge badge-rule">{r}</span>' 
                for r in active_rules
            ])
            st.markdown(badges_html, unsafe_allow_html=True)
        
        st.markdown('</div>', unsafe_allow_html=True)
        
        # Rules overview table
        with st.expander(" View All Column Rules", expanded=False):
            rules_data = []
            for col in template_cols:
                r = st.session_state.column_rules.get(col, ColumnRule(column_name=col))
                rules_data.append({
                    "Column": col,
                    "Mandatory": "" if r.is_mandatory else "",
                    "Not Null": "" if r.not_null else "",
                    "Rules": ", ".join(r.get_active_rules())
                })
            
            st.dataframe(
                pd.DataFrame(rules_data),
                use_container_width=True,
                hide_index=True
            )
        
        # Navigation
        st.markdown("---")
        col1, col2 = st.columns(2)
        with col1:
            if st.button("⬅️ Back to Data Source", use_container_width=True, key='btn_rules_to_source'):
                st.session_state.active_tab = 1
                st.rerun()
        with col2:
            if st.button("➡️ Proceed to Mapping", type="primary", use_container_width=True, key='btn_rules_to_mapping'):
                st.session_state.active_tab = 3
                st.rerun()

# =============================================================================
# TAB 4: COLUMN MAPPING
# =============================================================================

def render_mapping_tab():
    """Render Column Mapping Tab with Left-Right UI"""
    template_cols = st.session_state.template_file.columns.get(
        st.session_state.template_file.selected_sheet, []
    )
    source_cols = st.session_state.source_file.columns.get(
        st.session_state.source_file.selected_sheet, []
    )
    
    if not template_cols or not source_cols:
        st.error("Error No columns available for mapping")
        return
    
    st.markdown('<div class="section-title"> Step 4: Column Mapping</div>', 
                unsafe_allow_html=True)
    
    # Show current sheets and headers info
    st.markdown(f"""
    <div style="background: rgba(59, 130, 246, 0.1); border-radius: 8px; padding: 1rem; margin-bottom: 1rem;">
        <div style="display: flex; justify-content: space-between; align-items: center;">
            <div>
                <b> Template:</b> {st.session_state.template_file.selected_sheet} 
                <span style="color: #94a3b8;">({len(template_cols)} columns)</span>
            </div>
            <div style="color: #60a5fa; font-size: 1.5rem;"></div>
            <div>
                <b> Source:</b> {st.session_state.source_file.selected_sheet}
                <span style="color: #94a3b8;">({len(source_cols)} columns)</span>
            </div>
        </div>
    </div>
    """, unsafe_allow_html=True)
    
    st.markdown("""
    <div class="alert alert-info">
        Map columns from <strong>Source</strong> (left) to <strong>Template</strong> (right). 
        Select which source column should populate each template column.
    </div>
    """, unsafe_allow_html=True)
    
    # Enhanced Filter Controls
    st.markdown(f"""
    <div style="background: linear-gradient(135deg, rgba(59, 130, 246, 0.1) 0%, rgba(6, 182, 212, 0.1) 100%); 
                border-radius: 12px; padding: 1.5rem; margin: 1rem 0; border: 1px solid rgba(59, 130, 246, 0.2);">
        <div style="display: flex; align-items: center; gap: 1rem; margin-bottom: 1rem;">
            <div style="font-size: 1.2rem; font-weight: 700; color: #1e293b;">
                🔍 Mapping Filters
            </div>
            <div style="flex: 1;"></div>
            <div style="display: flex; gap: 0.5rem; align-items: center;">
                <span style="font-size: 0.9rem; color: #64748b;">Quick Stats:</span>
                <span style="background: #10b981; color: white; padding: 0.25rem 0.5rem; border-radius: 12px; font-size: 0.8rem; font-weight: 600;">
                    {len([m for m in st.session_state.mappings if m.is_active])}/{len(template_cols)} Mapped
                </span>
                <span style="background: #f59e0b; color: white; padding: 0.25rem 0.5rem; border-radius: 12px; font-size: 0.8rem; font-weight: 600;">
                    {sum(1 for col in template_cols if st.session_state.column_rules.get(col, ColumnRule(column_name=col)).is_mandatory)} Mandatory
                </span>
            </div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    # Filter Controls
    filter_col1, filter_col2, filter_col3, filter_col4 = st.columns(4)
    
    with filter_col1:
        show_mapped = st.checkbox("✅ Show Mapped", value=True, key="filter_mapped")
        
    with filter_col2:
        show_unmapped = st.checkbox("❌ Show Unmapped", value=True, key="filter_unmapped")
        
    with filter_col3:
        show_mandatory = st.checkbox("⚠️ Show Mandatory", value=True, key="filter_mandatory")
        
    with filter_col4:
        show_ruled = st.checkbox("📋 Show With Rules", value=True, key="filter_ruled")

    # Quick Actions
    st.markdown("---")
    action_col1, action_col2, action_col3, action_col4 = st.columns(4)
    
    with action_col1:
        if st.button("🤖 Auto-Map Similar", use_container_width=True, type="secondary"):
            with st.spinner("Analyzing column similarities..."):
                auto_mappings = MappingEngine.auto_map_columns(source_cols, template_cols)
                
                # Clear existing and add new auto-mappings
                st.session_state.mappings = []
                for mapping in auto_mappings:
                    mapping.source_sheet = st.session_state.source_file.selected_sheet
                    mapping.target_sheet = st.session_state.template_file.selected_sheet
                    st.session_state.mappings.append(mapping)
                
                # Force UI refresh by updating a timestamp
                st.session_state.auto_map_timestamp = time.time()
                
                st.success(f"🤖 Auto-mapped {len(auto_mappings)} columns!")
                # Don't call st.rerun() to avoid page refresh
    
    with action_col2:
        if st.button("🧹 Clear All Mappings", use_container_width=True, type="secondary"):
            st.session_state.mappings = []
            st.session_state.auto_map_triggered = False
            st.success("🧹 All mappings cleared!")
            st.rerun()
    
    with action_col3:
        if st.button("🔄 Reset Filters", use_container_width=True, type="secondary"):
            st.session_state.filter_mapped = True
            st.session_state.filter_unmapped = True
            st.session_state.filter_mandatory = True
            st.session_state.filter_ruled = True
            st.rerun()
    
    with action_col4:
        # Export mapping summary
        if st.button("📊 Export Mapping", use_container_width=True, type="secondary"):
            mapping_data = []
            for target in template_cols:
                rule = st.session_state.column_rules.get(target, ColumnRule(column_name=target))
                source = next((m.source_column for m in st.session_state.mappings if m.target_column == target), "")
                mapping_data.append({
                    "Target Column": target,
                    "Source Column": source,
                    "Mapped": "Yes" if source else "No",
                    "Mandatory": "Yes" if rule.is_mandatory else "No",
                    "Has Rules": "Yes" if rule.get_active_rules() else "No"
                })
            
            df_export = pd.DataFrame(mapping_data)
            csv_data = df_export.to_csv(index=False)
            st.download_button(
                "Download Mapping CSV",
                csv_data,
                "mapping_summary.csv",
                "text/csv",
                use_container_width=True
            )
    
    # Simple Arrow-Based Mapping Interface
    st.markdown("### 📋 Column Mapping")
    
    # Show current sheets info in a simple way
    st.markdown(f"**Source:** {st.session_state.source_file.selected_sheet} ({len(source_cols)} columns) → **Target:** {st.session_state.template_file.selected_sheet} ({len(template_cols)} columns)")
    
    st.markdown("---")
    
    current_mappings = {}
    for m in st.session_state.mappings:
        current_mappings[m.target_column] = m.source_column
    
    # Create reverse mapping for easy lookup
    reverse_mappings = {}
    for target, source in current_mappings.items():
        if source:
            reverse_mappings[source] = target
    
    # Filter columns based on user selections
    filtered_template_cols = []
    for target in template_cols:
        rule = st.session_state.column_rules.get(target, ColumnRule(column_name=target))
        is_mapped = target in current_mappings and current_mappings[target]
        is_mandatory = rule.is_mandatory
        has_rules = bool(rule.get_active_rules())
        
        # Apply filters
        show_this_column = (
            (show_mapped and is_mapped) or
            (show_unmapped and not is_mapped) or
            (show_mandatory and is_mandatory) or
            (show_ruled and has_rules)
        )
        
        # If no filters are active, show all columns
        if not any([show_mapped, show_unmapped, show_mandatory, show_ruled]):
            show_this_column = True
            
        if show_this_column:
            filtered_template_cols.append(target)

    # Show filtered results count
    if len(filtered_template_cols) != len(template_cols):
        st.info(f"Showing {len(filtered_template_cols)} of {len(template_cols)} columns (filtered)")

    # Add CSS for bordered groups
    st.markdown("""
    <style>
    .mapping-group {
        border: 2px solid #e5e7eb;
        border-radius: 8px;
        padding: 16px;
        margin: 8px 0;
        background: white;
    }
    .mapping-group.mapped {
        border-color: #10b981;
        background: linear-gradient(135deg, rgba(16, 185, 129, 0.05) 0%, rgba(16, 185, 129, 0.02) 100%);
        box-shadow: 0 0 8px rgba(16, 185, 129, 0.3);
    }
    .target-side {
        background: linear-gradient(135deg, rgba(16, 185, 129, 0.1) 0%, rgba(16, 185, 129, 0.05) 100%);
        padding: 12px;
        border-radius: 6px;
        border: 1px solid rgba(16, 185, 129, 0.2);
    }
    .source-side {
        background: linear-gradient(135deg, rgba(59, 130, 246, 0.1) 0%, rgba(59, 130, 246, 0.05) 100%);
        padding: 12px;
        border-radius: 6px;
        border: 1px solid rgba(59, 130, 246, 0.2);
    }
    </style>
    """, unsafe_allow_html=True)

    # Display simple arrow-based mapping pairs
    for i, target in enumerate(filtered_template_cols):
        rule = st.session_state.column_rules.get(target, ColumnRule(column_name=target))
        current_source = current_mappings.get(target, "")
        is_mapped = bool(current_source)
        is_mandatory = rule.is_mandatory
        has_rules = bool(rule.get_active_rules())

        # Simple arrow layout: Source → Target
        col1, arrow_col, col2 = st.columns([2, 1, 2])
        
        with col1:
            # Source column selection
            available_sources = ["-- Select Source --"]
            for s in source_cols:
                if s not in reverse_mappings:
                    available_sources.append(s)
                elif reverse_mappings.get(s) == target:
                    available_sources.append(s)
            
            options = available_sources
            current_value = current_source if current_source in options else ""

            try:
                current_index = options.index(current_value) if current_value else 0
            except:
                current_index = 0

            selected = st.selectbox(
                "📊 Source Column",
                options,
                index=current_index,
                key=f"map_{target}_{i}_{st.session_state.get('auto_map_timestamp', 0)}",
                help=f"Choose source column for '{target}'",
                label_visibility="visible"
            )

            # Update mapping if changed
            selected_clean = selected if selected != "-- Select Source --" else ""
            if selected_clean != current_source:
                # Remove old mapping for this target
                st.session_state.mappings = [m for m in st.session_state.mappings if m.target_column != target]
                # Remove old mapping for this source (if it was mapped to another target)
                if selected_clean:
                    st.session_state.mappings = [m for m in st.session_state.mappings if m.source_column != selected_clean]
                # Add new mapping if selected
                if selected_clean:
                    new_mapping = ColumnMapping(
                        target_column=target,
                        source_column=selected_clean,
                        target_sheet=st.session_state.template_file.selected_sheet,
                        source_sheet=st.session_state.source_file.selected_sheet
                    )
                    st.session_state.mappings.append(new_mapping)

        with arrow_col:
            # Arrow indicating direction
            arrow_color = "#10b981" if is_mapped else "#6b7280"
            st.markdown(f"<div style='text-align: center; color: {arrow_color}; font-size: 1.5rem; margin-top: 25px;'>→</div>", unsafe_allow_html=True)

        with col2:
            # Target column display
            target_bg = "#10b981" if is_mapped else "#6b7280"
            st.markdown(f"""
            <div style="background: rgba({144 if is_mapped else 107}, {163 if is_mapped else 114}, {138 if is_mapped else 128}, 0.1); 
                        border: 2px solid {target_bg}; border-radius: 8px; padding: 12px; margin-top: 8px;">
                <div style="font-weight: 600; color: {target_bg};">🎯 {target}</div>
            """, unsafe_allow_html=True)
            
            # Status badges
            badges = []
            if is_mapped:
                badges.append("✅ Mapped")
            if is_mandatory:
                badges.append("⚠️ Required")
            if has_rules:
                badges.append("📋 Has Rules")

            if badges:
                st.markdown(" ".join(badges))

            # Show validation rules preview
            if has_rules:
                active_rules = rule.get_active_rules()[:2]  # Show first 2 rules
                if active_rules:
                    rules_text = ", ".join(active_rules)
                    if len(rule.get_active_rules()) > 2:
                        rules_text += f" +{len(rule.get_active_rules()) - 2} more"
                    st.caption(f"📋 Rules: {rules_text}")

            # Show mapping confidence
            if selected_clean:
                confidence_score = next((m.confidence_score for m in st.session_state.mappings
                                       if m.target_column == target and m.source_column == selected_clean), 0)
                if confidence_score > 0:
                    confidence_pct = int(confidence_score * 100)
                    st.progress(confidence_score, text=f"{confidence_pct}% confidence")

            st.markdown("</div>", unsafe_allow_html=True)

        st.markdown("---")

    # Validation check for mandatory columns
    unmapped_mandatory = []
    for target in template_cols:
        rule = st.session_state.column_rules.get(target, ColumnRule(column_name=target))
        # ... (rest of the code remains the same)
        if rule.is_mandatory:
            if not any(m.target_column == target for m in st.session_state.mappings):
                unmapped_mandatory.append(target)

    if unmapped_mandatory:
        st.warning(f"⚠️ Mandatory columns not mapped: {', '.join(unmapped_mandatory)}")

    # Show current mappings summary
    with st.expander("📋 Current Mappings", expanded=False):
        if st.session_state.mappings:
            mapping_summary = []
            for m in st.session_state.mappings:
                mapping_summary.append({
                    "Source Column": m.source_column,
                    "Target Column": m.target_column,
                    "Status": "Active" if m.is_active else "Inactive"
                })
            st.dataframe(pd.DataFrame(mapping_summary), use_container_width=True, hide_index=True)
        else:
            st.info("No mappings configured yet")

    # Navigation
    st.markdown("---")
    col1, col2 = st.columns(2)
    with col1:
        if st.button("⬅️ Back to Rules", use_container_width=True, key='btn_mapping_to_rules'):
            st.session_state.active_tab = 2
            st.rerun()
    with col2:
        if st.button("✅ Validate & Export", type="primary", use_container_width=True, key='btn_mapping_to_validate'):
            if len(st.session_state.mappings) == 0:
                st.error("Please create at least one mapping")
            else:
                st.session_state.active_tab = 4
                st.rerun()

# =============================================================================
# TAB 5: VALIDATION & EXPORT
# =============================================================================

def render_validation_tab():
    st.markdown('<div class="section-title">Success Step 5: Validation & Export</div>', 
                unsafe_allow_html=True)
    
    if not st.session_state.mappings:
        st.warning("Warning Please configure column mappings first")
        return
    
    # Execute validation button
    if st.button(" Execute Validation & Transform", type="primary", use_container_width=True):
        with st.spinner("Processing data... This may take a moment"):
            execute_validation_pipeline()
    
    # Display results if available
    if st.session_state.validation_results:
        results = st.session_state.validation_results
        
        # Summary cards
        st.markdown('<div class="section-container">', unsafe_allow_html=True)
        
        cols = st.columns(4)
        with cols[0]:
            st.metric(
                "Total Rows",
                results.get('total_rows', 0)
            )
        with cols[1]:
            valid_pct = results.get('valid_percentage', 0)
            st.metric(
                "Valid Rows",
                results.get('valid_rows', 0),
                delta=f"{valid_pct:.1f}%"
            )
        with cols[2]:
            error_pct = results.get('error_percentage', 0)
            st.metric(
                "Rows with Errors",
                results.get('error_rows', 0),
                delta=f"{error_pct:.1f}%",
                delta_color="inverse"
            )
        with cols[3]:
            st.metric(
                "Total Errors",
                results.get('total_errors', 0)
            )
        
        st.markdown('</div>', unsafe_allow_html=True)
        
        # Error details
        if results.get('errors'):
            st.markdown('<div class="section-container">', unsafe_allow_html=True)
            st.markdown("** Validation Errors**")
            
            error_df = pd.DataFrame(results['errors'])
            
            # Filter options
            filter_col = st.selectbox(
                "Filter by Column",
                ["All"] + list(error_df['target_column'].unique())
            )
            
            if filter_col != "All":
                error_df = error_df[error_df['target_column'] == filter_col]
            
            st.dataframe(
                error_df,
                use_container_width=True,
                height=300,
                column_config={
                    "row": st.column_config.NumberColumn("Row #"),
                    "source_column": st.column_config.TextColumn("Source"),
                    "target_column": st.column_config.TextColumn("Target"),
                    "value": st.column_config.TextColumn("Value"),
                    "error": st.column_config.TextColumn("Error")
                }
            )
            st.markdown('</div>', unsafe_allow_html=True)
        
        # Transformed data preview
        if st.session_state.transformed_data is not None:
            st.markdown('<div class="section-container">', unsafe_allow_html=True)
            st.markdown("** Transformed Data Preview**")
            
            df = st.session_state.transformed_data
            
            # Show sample with color coding for errors
            st.dataframe(
                df.head(100),
                use_container_width=True,
                height=400
            )
            
            # Export options
            st.markdown("** Export Options**")
            
            col1, col2, col3 = st.columns(3)
            
            with col1:
                # Excel export - preserve entire template structure, replace only selected sheet
                if st.session_state.template_file.raw_data:
                    try:
                        # Import openpyxl for preserving formatting
                        from openpyxl import load_workbook
                        from openpyxl.utils.dataframe import dataframe_to_rows
                        
                        # Load the original template workbook to preserve all sheets and formatting
                        wb = load_workbook(io.BytesIO(st.session_state.template_file.raw_data))
                        
                        # Get the selected sheet name
                        selected_sheet = st.session_state.template_file.selected_sheet
                        
                        # Check if the selected sheet exists in the workbook
                        if selected_sheet in wb.sheetnames:
                            # Preserve original data and only update mapped columns
                            ws = wb[selected_sheet]
                            
                            # Get template columns and mapped columns
                            template_cols = st.session_state.template_file.columns.get(selected_sheet, [])
                            mapped_cols = [m.target_column for m in st.session_state.mappings if m.is_active and m.target_sheet == selected_sheet]
                            
                            # Get the actual header row position (0-indexed from session state)
                            header_row = st.session_state.template_file.header_row
                            data_start_row = header_row + 2  # Row after headers (1-indexed)
                            
                            # Create column index mapping
                            col_indices = {}
                            for col_idx, col_name in enumerate(template_cols, 1):
                                col_indices[col_name] = col_idx
                            
                            # Find the last row and column
                            max_row = ws.max_row
                            max_col = ws.max_column
                            
                            # Clear data only in mapped columns from data_start_row onwards
                            for row in range(data_start_row, max_row + 1):
                                for mapped_col in mapped_cols:
                                    if mapped_col in col_indices:
                                        col_idx = col_indices[mapped_col]
                                        if col_idx <= max_col:
                                            cell = ws.cell(row=row, column=col_idx)
                                            cell.value = None
                                            # Clear any comments/notes in mapped columns
                                            if cell.comment:
                                                cell.comment = None
                            
                            # Clear comments from header row in mapped columns
                            header_row_1_indexed = header_row + 1
                            for mapped_col in mapped_cols:
                                if mapped_col in col_indices:
                                    col_idx = col_indices[mapped_col]
                                    if col_idx <= max_col:
                                        cell = ws.cell(row=header_row_1_indexed, column=col_idx)
                                        if cell.comment:
                                            cell.comment = None
                            
                            # Hide all comments in the worksheet
                            ws.show_comments = False
                            
                            # Add the mapped data only to mapped columns starting from data_start_row
                            if len(df) > 0:
                                # Write data starting from data_start_row, only to mapped columns
                                for r_idx, row in enumerate(df.itertuples(index=False), data_start_row):
                                    for mapped_col in mapped_cols:
                                        if mapped_col in col_indices:
                                            col_idx = col_indices[mapped_col]
                                            if col_idx <= max_col:
                                                # Get the value for this mapped column
                                                if hasattr(row, mapped_col):
                                                    value = getattr(row, mapped_col)
                                                else:
                                                    # Fallback to column position
                                                    col_pos = template_cols.index(mapped_col) if mapped_col in template_cols else None
                                                    value = row[col_pos] if col_pos is not None and col_pos < len(row) else None
                                                
                                                cell = ws.cell(row=r_idx, column=col_idx)
                                                cell.value = value
                                                # Ensure no comments on new cells
                                                if cell.comment:
                                                    cell.comment = None
                            
                            # Also hide comments for the entire workbook
                            wb.show_comments = False
                            
                            # Save the modified workbook
                            output = io.BytesIO()
                            wb.save(output)
                            
                            # Use original filename
                            original_filename = st.session_state.template_file.name
                            if not original_filename.lower().endswith('.xlsx'):
                                original_filename = original_filename.rsplit('.', 1)[0] + '_export.xlsx'
                            
                            st.download_button(
                                label=" Download Excel (.xlsx)",
                                data=output.getvalue(),
                                file_name=original_filename,
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                use_container_width=True
                            )
                        else:
                            # Fallback if sheet not found
                            st.error(f" Selected sheet '{selected_sheet}' not found in template")
                            output = io.BytesIO()
                            with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                                df.to_excel(writer, sheet_name='Mapped_Data', index=False)
                                if results.get('errors'):
                                    error_df = pd.DataFrame(results['errors'])
                                    error_df.to_excel(writer, sheet_name='Validation_Errors', index=False)
                            
                            st.download_button(
                                label=" Download Excel (.xlsx)",
                                data=output.getvalue(),
                                file_name=f"template_export_{st.session_state.session_id}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                use_container_width=True
                            )
                            
                    except ImportError:
                        st.warning(" openpyxl not available, using fallback export")
                        # Fallback without formatting preservation
                        output = io.BytesIO()
                        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                            df.to_excel(writer, sheet_name='Mapped_Data', index=False)
                            if results.get('errors'):
                                error_df = pd.DataFrame(results['errors'])
                                error_df.to_excel(writer, sheet_name='Validation_Errors', index=False)
                        
                        st.download_button(
                            label=" Download Excel (.xlsx)",
                            data=output.getvalue(),
                            file_name=f"template_export_{st.session_state.session_id}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True
                        )
                else:
                    # No original template available
                    st.warning(" Original template not available for format preservation")
                    output = io.BytesIO()
                    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                        df.to_excel(writer, sheet_name='Mapped_Data', index=False)
                        if results.get('errors'):
                            error_df = pd.DataFrame(results['errors'])
                            error_df.to_excel(writer, sheet_name='Validation_Errors', index=False)
                    
                    st.download_button(
                        label=" Download Excel (.xlsx)",
                        data=output.getvalue(),
                        file_name=f"template_export_{st.session_state.session_id}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
            
            with col2:
                # CSV export
                csv = df.to_csv(index=False)
                st.download_button(
                    label=" Download CSV",
                    data=csv,
                    file_name=f"mapped_data_{st.session_state.session_id}.csv",
                    mime="text/csv",
                    use_container_width=True
                )
            
            with col3:
                # Rejected records export
                if results.get('errors'):
                    error_df = pd.DataFrame(results['errors'])
                    output = io.BytesIO()
                    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                        error_df.to_excel(writer, sheet_name='Rejected_Records', index=False)
                    
                    st.download_button(
                        label=" Download Rejected Records",
                        data=output.getvalue(),
                        file_name=f"rejected_records_{st.session_state.session_id}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
                else:
                    st.info(" No rejected records to export")
            
            st.markdown('</div>', unsafe_allow_html=True)
    
    st.markdown("---")
    if st.button("⬅️ Back to Mapping", use_container_width=True, key='btn_validate_to_mapping'):
        st.session_state.active_tab = 3
        st.rerun()

def execute_validation_pipeline():
    """Execute the full validation and transformation pipeline"""
    try:
        source_df = st.session_state.source_file.data.get(
            st.session_state.source_file.selected_sheet,
            pd.DataFrame()
        )
        
        template_cols = st.session_state.template_file.columns.get(
            st.session_state.template_file.selected_sheet,
            []
        )
        
        # Initialize target dataframe
        transformed_data = pd.DataFrame(columns=template_cols)
        errors = []
        valid_rows = 0
        error_rows = 0
        
        # Process each row
        for idx, row in source_df.iterrows():
            new_row = {}
            row_has_error = False
            
            for mapping in st.session_state.mappings:
                if not mapping.is_active:
                    continue
                
                source_val = row.get(mapping.source_column, None)
                target_col = mapping.target_column
                rule = st.session_state.column_rules.get(
                    target_col, 
                    ColumnRule(column_name=target_col)
                )
                
                # Validate and transform
                is_valid, val_errors, transformed_val = ValidationEngine.validate_value(
                    source_val, rule
                )
                
                if not is_valid:
                    row_has_error = True
                    for err in val_errors:
                        errors.append({
                            'row': idx + 1,
                            'source_column': mapping.source_column,
                            'target_column': target_col,
                            'value': source_val,
                            'error': err
                        })
                
                new_row[target_col] = transformed_val
            
            # Handle unmapped columns
            for col in template_cols:
                if col not in new_row:
                    rule = st.session_state.column_rules.get(col, ColumnRule(column_name=col))
                    if rule.default_value:
                        new_row[col] = rule.default_value
                    else:
                        new_row[col] = None
                    
                    if rule.is_mandatory:
                        row_has_error = True
                        errors.append({
                            'row': idx + 1,
                            'source_column': 'N/A',
                            'target_column': col,
                            'value': None,
                            'error': 'Mandatory column not mapped'
                        })
            
            # Add row to result
            transformed_data = pd.concat([
                transformed_data, 
                pd.DataFrame([new_row])
            ], ignore_index=True)
            
            if row_has_error:
                error_rows += 1
            else:
                valid_rows += 1
        
        # Update session state
        total_rows = len(source_df)
        st.session_state.transformed_data = transformed_data
        st.session_state.validation_results = {
            'total_rows': total_rows,
            'valid_rows': valid_rows,
            'error_rows': error_rows,
            'valid_percentage': (valid_rows / total_rows * 100) if total_rows > 0 else 0,
            'error_percentage': (error_rows / total_rows * 100) if total_rows > 0 else 0,
            'total_errors': len(errors),
            'errors': errors[:500]  # Limit stored errors
        }
        
        st.success(" Validation complete!")
        
    except Exception as e:
        st.error(f" Error during processing: {str(e)}")
        import traceback
        st.error(traceback.format_exc())

# =============================================================================
# MAIN APPLICATION
# =============================================================================

def main():
    """Main application entry point"""
    init_session_state()
    render_header()
    render_sidebar()
    render_progress_steps()
    
    # Main tabs
    tabs = st.tabs([
        "Fusion Template Upload",
        "Upload Source Records", 
        "Apply Validation Rules",
        "Mapping B/W Template  & Source",
        "Validate & Export"
    ])
    
    with tabs[0]:
        render_template_tab()
    
    with tabs[1]:
        render_source_tab()
    
    with tabs[2]:
        render_rules_tab()
    
    with tabs[3]:
        render_mapping_tab()
    
    with tabs[4]:
        render_validation_tab()
    
    # Auto-navigate to active tab (for button navigation)
    # Always run JavaScript to ensure tab switching works
    js = f"""
    <script>
        // Wait for DOM to be ready
        setTimeout(function() {{
            var tabs = window.parent.document.querySelectorAll('[data-baseweb="tab"]');
            if (tabs.length > {st.session_state.active_tab}) {{
                tabs[{st.session_state.active_tab}].click();
            }}
        }}, 100);
    </script>
    """
    st.components.v1.html(js, height=0)

if __name__ == "__main__":
    main()


    